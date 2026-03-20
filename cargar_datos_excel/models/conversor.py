import openpyxl
import os
from models.base_datos_posgrest import save_data

CARACTERES_REEMPLAZO = dict([
    ("datetime.time(", "'"),
    ("datetime.datetime(", "'"),
    (")", "'"),
    ("?", " "),
    ("#", " "),
    ("*", " "),
    (",", "-"),
    ("$", " "),
    ("%", " "),
    ("{", " "),
    ("}", " "),
    ("[", " "),
    ("]", " "),
    ("<", " "),
    (">", " "),
    ("¨", " "),
    ("^", " "),
    ("~", " "),
    ("`", " "),
    ('"', " "),
    ("'", "")
])

def transformar(dato):
    dato = str(dato)
    dato = dato.replace("datetime.time(",  " '")
    dato = dato.replace("?",  ' ')
    dato = dato.replace("#",  ' ')
    # dato = dato.replace("/",  ' ')
    dato = dato.replace("*",  ' ')
    dato = dato.replace(",",  '-')
    dato = dato.replace("datetime.datetime(", " '")
    # dato = dato.replace("(",  ' ')
    dato = dato.replace("),",  "',")
    # dato = dato.replace(")",  ' ')
    dato = dato.replace("$",  ' ')
    dato = dato.replace("%",  ' ')
    dato = dato.replace("{",  ' ')
    dato = dato.replace("}",  ' ')
    dato = dato.replace("[",  ' ')
    dato = dato.replace("]",  ' ')
    dato = dato.replace("<",  ' ')
    dato = dato.replace(">",  ' ')
    dato = dato.replace("¨",  ' ')
    dato = dato.replace("^",  ' ')
    dato = dato.replace("~",  ' ')
    dato = dato.replace("`",  ' ')
    dato = dato.replace("'",  '')
    dato = dato.replace('"',  ' ')
    return dato
    



class conversion:
    def __init__(self):
        self.wb = None
        self.nombre_archivo = ""

    def leer_excel(self, fichero):
        self.wb = openpyxl.load_workbook(fichero, data_only=True)
        self.nombre_archivo = os.path.basename(fichero).lower()

    def detectar_tipo_resultado(self, col_max):
        if "hechos archivo plano" in self.nombre_archivo:
            return "HECHOS"
        elif "resultados aplicativo" in self.nombre_archivo:
            return "RESULTADOS"
        elif col_max == 12:
            return "ERRADICACION"
        else:
            return None

    def convertir(self, hoja):
        sheet = self.wb[hoja]
        fila_max = sheet.max_row
        col_max = sheet.max_column

        resultado = self.detectar_tipo_resultado(col_max)
        if not resultado:
            print(f"No se puede determinar el tipo de resultado para archivo '{self.nombre_archivo}' con {col_max} columnas.")
            return

        data = []

        for row in sheet.iter_rows(min_row=2, max_row=fila_max, max_col=col_max):
            fila = [transformar(cell.value) for cell in row]
            if all(cell is None for cell in fila):
                break
            data.append(tuple(fila))

        if not data:
            print(f"No se encontraron datos válidos en '{self.nombre_archivo}'")
            return

        try:
            save_data(data, resultado)
        except Exception as e:
            print(f"Error al guardar datos: {e}")
