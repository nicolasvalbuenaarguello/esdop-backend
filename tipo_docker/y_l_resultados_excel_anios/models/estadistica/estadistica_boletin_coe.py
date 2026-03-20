# El presente modulo es para el manejo de la estadística de la Dirección de operaciones del Ejército Nacional 
from tipo_docker.y_l_resultados_excel_anios.models.conexion_pos import Databa_bases
from tipo_docker.z_z_z_funtions.funciones import *
from tipo_docker.y_l_resultados_excel_anios.models.funtions.componest.tablas import *

conexion_pos = Databa_bases()
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.descriptors.serialisable import Serialisable
import pandas as pd

class Calculo_Spoa:
    def __init__(self):
            self.archivo = "RESULTADO SIN SPOA" +".txt"              
            self.f = open("doc_sin_spoa/"+self.archivo, "w")
            self.f.write(str("RESULTADO")+", " + str("HR")+", " +str("FECHA")+", "+str("SPOA")+", "+str("DIV")+"\n")

    def validar_spoa_unidad(self, filtro, res_calculo, numero, validar):
        spoa=[]
        no_spoa=[]
        
        if validar == "SI":
            
            startdate = pd.to_datetime("2024-06-04").date()
            # startdate = pd.to_datetime("2024-01-01").date()
            
            if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
                for x in res_calculo[numero]:
            
                    if x[0] >= startdate:

                        # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                        if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                            no_spoa.append(x)
                            # print(x[29])
                        else: 
                            numero_dep = x[29][0:5]
                                
                            # if numero_dep == "00000":
                            #     no_spoa.append(x)
                            # else:
                            #     spoa.append(x)
                            spoa.append(x)

                    else: 
                        spoa.append(x)      
            else:
                spoa =  res_calculo[numero]

                

            spoa = spoa
        else:
            spoa =  res_calculo[numero]

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        return[spoa, no_spoa]

    def validar_spoa(self, filtro, res_calculo, numero, nombre, validar):
        spoa=[]
        no_spoa=[]
        
        startdate = pd.to_datetime("2024-06-04").date()
        # startdate = pd.to_datetime("2024-01-01").date()
        
        if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
            for x in res_calculo[numero]:
        
                if x[0] >= startdate:

                    # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                    if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                        no_spoa.append(x)
                        # print(x[29])
                    else: 
                        numero_dep = x[29][0:5]
                            
                        # if numero_dep == "00000":
                        #     no_spoa.append(x)
                        # else:
                        #     spoa.append(x)
                        spoa.append(x)

                else: 
                    spoa.append(x)      
        else:
            spoa =  res_calculo[numero]

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        
        # f.write(str(nombre)+"\n")
        # f.write("-----------------------------------------------------"+"\n")
        
        numero_id = 1
        for x in no_spoa:
            self.f.write(str(nombre)+", " + str(x[26])+", " +str(x[0])+", "+str(x[29])+", "+str(x[1])+"\n")
            numero_id = numero_id + 1

        # f.write("-----------------------------------------------------"+"\n")
        # f.write(str(numero_id-1)+" Resultados sin SPOA"+"\n")

        if validar == "SI":
            spoa = spoa
        else:
            spoa =  res_calculo[numero]
            

        return[spoa, no_spoa]

    # cuadro de resultados resaltantes tamaño oficio
    def resultados_resaltantes_pdf_oficio(self, fecha_inicial_u_l, fecha_final_u_l, filtro,  dato, wb):


        dato =""
        if  filtro[4] == "lugar":
            if filtro[6]!="" and filtro[6]!="---" :#filtro por municipio
                nueva=filtro[6].split(",")
                dato = ""
                dato_res=""
                if len(nueva) > 1:
                        ids = tuple(nueva)
                        dato = "and dpto = '{}' and mpio in {}".format(filtro[5], ids)
                        dato_res= "and hop_depto = '{}' and hop_mpio in {}".format(filtro[5], ids)

                else:
                    mpio = nueva[0]  
                    dato = "and {} = '{}' and {} = '{}'".format("dpto",filtro[5], "mpio",mpio)
                    dato_res = "and {} = '{}' and {} = '{}'".format("hop_depto", filtro[5], "hop_mpio",mpio)
                filtros =[dato_res, dato, ""]
            else:
                if dato == "":
                    filtros = selecion_filtro(filtro)
                else:
                    filtros = dato
        else:
            if dato == "":
                filtros = selecion_filtro(filtro)
            else:
                filtros = dato
        # print(filtros)

        query = parametros(fecha_inicial_u_l, fecha_final_u_l, filtros, filtro)
        # print(filtro)

        # agr_div
        # division
        # brigada
        # unidad
        # dpto
        # mpio

                
        BUILTIN_FORMATS = {
                                0: 'General',
                                1: '0',
                                2: '0.00',
                                3: '#,##0',
                                4: '#,##0.00',
                                5: '"$"#,##0_);("$"#,##0)',
                                6: '"$"#,##0_);[Red]("$"#,##0)',
                                7: '"$"#,##0.00_);("$"#,##0.00)',
                                8: '"$"#,##0.00_);[Red]("$"#,##0.00)',
                                9: '0%',
                                10: '0.00%',
                                11: '0.00E+00',
                                12: '# ?/?',
                                13: '# ??/??',
                                14: 'mm-dd-yy',
                                15: 'd-mmm-yy',
                                16: 'd-mmm',
                                17: 'mmm-yy',
                                18: 'h:mm AM/PM',
                                19: 'h:mm:ss AM/PM',
                                20: 'h:mm',
                                21: 'h:mm:ss',
                                22: 'm/d/yy h:mm',

                                37: '#,##0_);(#,##0)',
                                38: '#,##0_);[Red](#,##0)',
                                39: '#,##0.00_);(#,##0.00)',
                                40: '#,##0.00_);[Red](#,##0.00)',

                                41: r'_(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)',
                                42: r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)',
                                43: r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)',

                                44: r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)',
                                45: 'mm:ss',
                                46: '[h]:mm:ss',
                                47: 'mmss.0',
                                48: '##0.0E+0',
                                49: '@', }
            
        frmt = BUILTIN_FORMATS[3] 
        frmt_d = BUILTIN_FORMATS[4]   
            # for row in hoja["B:U"]:
            #     cells = row[0], row[-1]
            #     for c in cells:
            #         c.number_format = frmt

        # print(filtro)
        hechos_filtro = ""
        nombre_hoja = "---"
        if filtro[4] == 'unidad':
            if filtro[3]!="" and filtro[3]!="---":
                hechos_filtro = " and {} = '{}'".format("unidad",filtro[3])
                nombre_hoja = "TOTAL " + filtro[3]
            elif filtro[2]!="" and filtro[2]!="---":
                hechos_filtro = " and {} = '{}'".format("brigada",filtro[2])
                nombre_hoja = "TOTAL " + filtro[2]
            elif filtro[1]!="" and filtro[1]!="---":
                hechos_filtro = " and {} = '{}'".format("division",filtro[1])
                nombre_hoja = "TOTAL " + filtro[1]
            elif filtro[0]!="" and filtro[0]!="---":
                hechos_filtro = " and {} = '{}'".format("agr_div",filtro[0])
                nombre_hoja = "TOTAL " + filtro[0]

        elif filtro[4] == 'lugar':

            if filtro[6]!="" and filtro[6]!="---" :#filtro por municipio
                nueva=filtro[6].split(",")
                dato = ""
                dato_res=""
                if len(nueva) > 1:
                        ids = tuple(nueva)
                        hechos_filtro = "and mpio in {}".format(ids)

                else:
                    mpio = nueva[0]  
                    hechos_filtro = " and {} = '{}'".format("mpio",mpio)

                mpio = nueva[0]  
                # hechos_filtro = " and {} = '{}'".format("mpio",mpio)
                nombre_hoja = "TOTAL " + filtro[6]

            elif filtro[5]!= "---" :#filtro por departamento
                hechos_filtro = " and {} = '{}'".format("dpto",filtro[5])
                nombre_hoja = "TOTAL " + filtro[5]

        if nombre_hoja =="---":
            nombre_hoja = "TOTAL EJC"


        if filtro[19] == "gaulas":
            # print(filtro[19])
            hechos_filtro = hechos_filtro + str(" and unidad like 'GG%'")
            nombre_hoja = "TOTAL GAULAS"
                

        querys_amenaza = "SELECT DISTINCT extract(year from fecha_hecho) as anio FROM view_hechos_materializados WHERE fecha_hecho >= '{}'  AND fecha_hecho <= '{}' {} ORDER BY anio ASC".format(fecha_inicial_u_l, fecha_final_u_l, hechos_filtro)
        listado_amenaza = conexion_pos.comando_query(querys_amenaza)

        querys_bris = "SELECT DISTINCT agr_div, extract(year from fecha_hecho) as anio  FROM view_hechos_materializados WHERE fecha_hecho >= '{}'  AND fecha_hecho <= '{}' {} ORDER BY anio ASC".format(fecha_inicial_u_l, fecha_final_u_l, hechos_filtro)
        listado_brig = conexion_pos.comando_query(querys_bris)
    
        querys_div = "SELECT DISTINCT agr_div  FROM view_hechos_materializados WHERE fecha_hecho >= '{}'  AND fecha_hecho <= '{}' {} ORDER BY agr_div ASC".format(fecha_inicial_u_l, fecha_final_u_l, hechos_filtro)
        listado_div = conexion_pos.comando_query(querys_div)

        
        resultados = conexion_pos.comando_query(query[0])
        hechos = conexion_pos.comando_query(query[1])
        erradicacion = conexion_pos.comando_query(query[2])
        # print(query[2])
        res_calculo = estadistica_resultados(resultados, hechos, filtro)


        hojas=0


        hoja = wb.active       
        hoja.title = nombre_hoja


        #RESULTADOS AFECTACION AL ENEMIGO
        hoja.cell(row=1, column=1, value="RESUTADOS")
        hoja.cell(row=2, column=1, value="MENORES RECUPERADOS")
        hoja.cell(row=3, column=1, value="PRESENTACION VOLUNTARIA")
        hoja.cell(row=4, column=1, value="SOMETIMIENTOS A LA JUSTICIA")
        hoja.cell(row=5, column=1, value="CAPTURAS")
        hoja.cell(row=6, column=1, value="MDOM")

        #RESULTADOS AFECTACION PROPIAS TROPAS

        hoja.cell(row=7, column=1, value="ASESINADOS")
        hoja.cell(row=8, column=1, value="HERIDOS")

        #RESULTADOS ARMAMENTO

        hoja.cell(row=9, column=1, value="ARMAS DE LARGO ALCANCE")
        hoja.cell(row=10, column=1, value="ARMAS DE CORTO ALCANCE")
        hoja.cell(row=11, column=1, value="ARMAS DE ACOMPAÑAMIENTO")
        hoja.cell(row=12, column=1, value="MUNICIONES")

        #RESULTADOS COMBATES

        hoja.cell(row=13, column=1, value="COMBATES POSITIVOS")
        hoja.cell(row=14, column=1, value="COMBATES NEGATIVOS")
        hoja.cell(row=15, column=1, value="COMBATES SIN RESULTADOS")
        hoja.cell(row=16, column=1, value="TOTAL DE COMBATES")

        hoja.cell(row=17, column=1, value="HOSTIGAMIENTOS")
        hoja.cell(row=18, column=1, value="ASONADAS")

        #RESULTADOS NARCOTRAFICO
        hoja.cell(row=19, column=1, value="COCAÍNA Kg")
        hoja.cell(row=20, column=1, value="MARIHUANA Kg")
        hoja.cell(row=21, column=1, value="PBC Kg")
        hoja.cell(row=22, column=1, value="HEROHINA Kg")
        hoja.cell(row=23, column=1, value="BASUCO Kg")
        hoja.cell(row=24, column=1, value="DROGAS SINTETICAS und")
        hoja.cell(row=25, column=1, value="LAB. CLORHIDRATO DE COCAINA")
        hoja.cell(row=26, column=1, value="LAB. PASTA O BASE DE COCA")
        hoja.cell(row=27, column=1, value="LAB. HEROHINA")
        hoja.cell(row=28, column=1, value="SEMILLEROS")
        hoja.cell(row=29, column=1, value="MATAS DE COCA EN SEMILLEROS")
        hoja.cell(row=30, column=1, value="INSUMOS SOLIDOS Kg")
        hoja.cell(row=31, column=1, value="INSUMOS LIQUIDOS Gal")
        hoja.cell(row=32, column=1, value="COMBUSTIBLES NARCOTRAFICO Gal")

        #RESULTADOS EXPLOSIVOS
        hoja.cell(row=33, column=1, value="NEUTRALIZACIÓN TERRORISTA")
        hoja.cell(row=34, column=1, value="ARTEFACTOS EXPLOSIVOS Und")
        hoja.cell(row=35, column=1, value="A.E.I")
        hoja.cell(row=36, column=1, value="MAP (Mina Anti Persona)")
        hoja.cell(row=37, column=1, value="EXPLOSIVOS kg")
        hoja.cell(row=38, column=1, value="EXPLOSIVOS m")
        hoja.cell(row=39, column=1, value="EXPLOSIVOS Und")

        hoja.cell(row=40, column=1, value="CORDÓN DETONANTE m")
        hoja.cell(row=41, column=1, value="MECHA LENTA m")
        hoja.cell(row=42, column=1, value="MEDIOS DE LANZAMIENTOS")
        hoja.cell(row=43, column=1, value="DETONADORES ELÉCTRICOS")
        hoja.cell(row=44, column=1, value="DETONADORES ANELÉCTRICOS")

        hoja.cell(row=45, column=1, value="LIBERADOS")
        hoja.cell(row=46, column=1, value="RESCATADOS")
        
        #RESULTADOS HIDROCARBUROS
        hoja.cell(row=47, column=1, value="VÁLVULAS")
        hoja.cell(row=48, column=1, value="REFINERIAS")
        hoja.cell(row=49, column=1, value="AFECTACIONES AL OLEODUCTO")
        hoja.cell(row=50, column=1, value="PETROLEO")
        hoja.cell(row=51, column=1, value="PISCINAS")

        hoja.cell(row=52, column=1, value="INCAUTACION DE PESOS COLOMBIANOS")
        hoja.cell(row=53, column=1, value="INCAUTACION DE DOLARES")
        hoja.cell(row=54, column=1, value="INCAUTACION DE EUROS")

        #RESULTADOS AMAZONIA

        hoja.cell(row=55, column=1, value="CAPTURAS LOE AMAZONÍA")
        hoja.cell(row=56, column=1, value="PLÁNTULAS SEMBRADAS")
        hoja.cell(row=57, column=1, value="INCAUTACIÓN MADERA M3")
        hoja.cell(row=58, column=1, value="ESPECIES ANIMALES INCAUTADOS")

        #RESULTADOS MINERIA

        hoja.cell(row=59, column=1, value="CAPTURAS MINERIA")
        hoja.cell(row=60, column=1, value="YACIMIENTOS MINEROS")
        hoja.cell(row=61, column=1, value="UNIDAD MINERA PRODUCCIÓN")
        hoja.cell(row=62, column=1, value="MAQUINARIA PESADA")
        hoja.cell(row=63, column=1, value="TRACTOR CON URUGA")
        hoja.cell(row=64, column=1, value="EXCAVADORAS")
        hoja.cell(row=65, column=1, value="RETROEXCAVADORAS")
        hoja.cell(row=66, column=1, value="DRAGAS")
        hoja.cell(row=67, column=1, value="MAQUINARIA AMARILLA")
        hoja.cell(row=68, column=1, value="MOTORES")
        hoja.cell(row=69, column=1, value="MEDIOS DE TRANSPORTE")
        hoja.cell(row=70, column=1, value="COMBUSTIBLES EN MINERIA")
        hoja.cell(row=71, column=1, value="EXPLOSIVOS MINERIA kg")
        hoja.cell(row=72, column=1, value="EXPLOSIVOS MINERIA m")
        hoja.cell(row=73, column=1, value="EXPLOSIVOS MINERIA Und")
        hoja.cell(row=74, column=1, value="COLTAN")

        #RESULTADOS EXTINCION DE DOMINIO 

        hoja.cell(row=75, column=1, value="INCAUTACIÓN DE MUEBLES")
        hoja.cell(row=76, column=1, value="INCAUTACIÓN DE INMUEBLES")
        hoja.cell(row=77, column=1, value="INCAUTACIÓN DE VEHICULOS")

        #RESULTADOS CONTRABANDO 

        hoja.cell(row=78, column=1, value="CAPTURAS CONTRABANDO")
        hoja.cell(row=79, column=1, value="COMBUSTIBLES CONTRABANDO")
        hoja.cell(row=80, column=1, value="GASOLINA CONTRABANDO")
        hoja.cell(row=81, column=1, value="ACPM CONTRABANDO")
        hoja.cell(row=82, column=1, value="VEHICULOS DE CONTRABANDO")
        
        hoja.cell(row=83, column=1, value="AFECTACION OLEODUCTO")
        hoja.cell(row=84, column=1, value="APIQUE")
        hoja.cell(row=85, column=1, value="ABOYADURA")
        hoja.cell(row=86, column=1, value="ROTURA")
        hoja.cell(row=87, column=1, value="SABOTAJE AL OLEODUCTO")
        hoja.cell(row=88, column=1, value="TOTAL DE RESULTADOS")

        # for cell in hoja["2:2"]: 
        #         cell.font = Font(bold=True, size=12)
        #         cell.fill  = PatternFill('solid', start_color="ABA2A2")
        #         # cell.alignment = Alignment(wrap_text=True)
        #         cell.alignment=Alignment(
        #                     horizontal='center',
        #                     vertical='top',
        #                     text_rotation=0,
        #                     wrap_text=False,
        #                     shrink_to_fit=False,
        #                     indent=0)
        listados=[]
        col = 2 
        col_merge_inicio = 2
        for y in listado_amenaza:
                col_merge_final = col_merge_inicio
                columna = col_merge_inicio
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 0, "NO") #funcion para filtar el spoa
                numero_res =calculo(valor_calculado[0], y[0], 0, 18)
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 1, "NO") #funcion para filtar el spoa
                numero_presentacion =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 2, "NO") #funcion para filtar el spoa
                sometimientos =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 3, "SI") #funcion para filtar el spoa

                 # print(filtro[17])
                if filtro[17] == "sin_delco":
                    valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 61, "CAPTURAS", "SI" ) #funcion para filtar el spoa
                else:
                    valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 3, "CAPTURAS", "SI" ) #funcion para filtar el spoa

                capturas =calculo(valor_calculado[0], y[0], 0, 18)

                # print(valor_calculado[0][1])
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 4, "SI") #funcion para filtar el spoa
                mdom =calculo(valor_calculado[0], y[0], 0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 5, "SI") #funcion para filtar el spoa
                asesinados =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 6, "SI") #funcion para filtar el spoa
                heridos =calculo(valor_calculado[0], y[0], 0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 7, "SI") #funcion para filtar el spoa
                largas =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 8, "SI") #funcion para filtar el spoa
                cortas =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 9, "SI") #funcion para filtar el spoa
                acompaniamiento =calculo(valor_calculado[0], y[0], 0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 10, "SI") #funcion para filtar el spoa
                municiones =calculo(valor_calculado[0], y[0], 0, 18)

                combates_pos = calculo(res_calculo[11],y[0],1, 14)
                combates_neg = calculo(res_calculo[12],y[0],1, 14)
                combates_sin_res = calculo(res_calculo[13],y[0],1, 14)
                combates  = calculo(res_calculo[14],y[0],1, 14)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 66, "SI") #funcion para filtar el spoa
                hostigamientos   = calculo(valor_calculado[0],y[0],0, 18)
                asonadas   = calculo(res_calculo[67],y[0],0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 21, "SI") #funcion para filtar el spoa
                cocaina = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 22, "SI") #funcion para filtar el spoa
                marihuna = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 23, "SI") #funcion para filtar el spoa
                pbc = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 69, "SI") #funcion para filtar el spoa
                heronia = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 68, "SI") #funcion para filtar el spoa
                basuco = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 70, "SI") #funcion para filtar el spoa
                drogas_sinteticas = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 24, "SI") #funcion para filtar el spoa
                lab_cocaina = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 25, "SI") #funcion para filtar el spoa
                lab_pbc = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 71, "SI") #funcion para filtar el spoa
                lab_heronina = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 26, "SI") #funcion para filtar el spoa
                semilleros = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 27, "SI") #funcion para filtar el spoa
                matas_semilleros = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 29, "SI") #funcion para filtar el spoa
                insumos_solidos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 28, "SI") #funcion para filtar el spoa
                nsumos_liquidos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 89, "SI") #funcion para filtar el spoa
                combustible_narcotrafico = calculo(valor_calculado[0],y[0],0, 18)
  
                neutralizacion_t = calculo(res_calculo[15],y[0],1, 14)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 80, "NO") #funcion para filtar el spoa
                artefacto_explosivo = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 16, "NO") #funcion para filtar el spoa
                ae = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 17, "NO") #funcion para filtar el spoa
                map = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 18, "NO") #funcion para filtar el spoa
                explosivos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 78, "NO") #funcion para filtar el spoa
                explosivos_m = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 79, "NO") #funcion para filtar el spoa
                explosivos_und = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 19, "NO") #funcion para filtar el spoa
                cordon = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 20, "NO") #funcion para filtar el spoa
                mecha = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 72, "NO") #funcion para filtar el spoa
                medios_lanzamientos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 73, "NO") #funcion para filtar el spoa
                detonadores_electricos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 74, "NO") #funcion para filtar el spoa
                detonadores_anaelectricos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 42, "SI") #funcion para filtar el spoa
                liberados = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 43, "SI") #funcion para filtar el spoa
                rescatados = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 44, "NO") #funcion para filtar el spoa
                valvulas = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 45, "NO") #funcion para filtar el spoa
                refinerias = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 86, "NO") #funcion para filtar el spoa
                oleoducto = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 87, "NO") #funcion para filtar el spoa
                petroleo = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 88, "NO") #funcion para filtar el spoa
                piscina = calculo(valor_calculado[0],y[0],0, 18)

 
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 76, "NO") #funcion para filtar el spoa
                pesos = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 75, "NO") #funcion para filtar el spoa
                dolares = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 77, "NO") #funcion para filtar el spoa
                euros = calculo(valor_calculado[0],y[0],0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 30, "SI") #funcion para filtar el spoa          
                cap_amazonia = calculo(valor_calculado[0],y[0],0, 18)
                plantulas = calculo(res_calculo[31],y[0],0, 18)
                madera = calculo(res_calculo[32],y[0],0, 18)
                especies = calculo(res_calculo[63],y[0],0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 33, "SI") #funcion para filtar el spoa                 
                cap_menria = calculo(valor_calculado[0],y[0],0, 18)

                yacimiento_mineros = calculo(res_calculo[35],y[0],1, 14)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 40, "NO") #funcion para filtar el spoa 
                upm = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 38, "NO") #funcion para filtar el spoa 
                maquinaria_pesada = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 39, "NO") #funcion para filtar el spoa 
                tractor_con_oruga = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 36, "NO") #funcion para filtar el spoa 
                excavadoras = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 37, "NO") #funcion para filtar el spoa 
                retroexcavadoras = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 41, "NO") #funcion para filtar el spoa 
                dragas = calculo(valor_calculado[0],y[0],0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 81, "NO") #funcion para filtar el spoa 
                maquinaria_amarilla = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 59, "NO") #funcion para filtar el spoa 
                motores = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 85, "NO") #funcion para filtar el spoa 
                medios_trasnporte = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 60, "SI") #funcion para filtar el spoa 
                combustible_mineria = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 82, "NO") #funcion para filtar el spoa 
                explosivos_mineria_kg = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 83, "NO") #funcion para filtar el spoa 
                explosivos_mineria_m = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 84, "NO") #funcion para filtar el spoa 
                explosivos_mineria_und = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 65, "NO") #funcion para filtar el spoa 
                coltan = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 91, "NO") #funcion para filtar el spoa 
                incautacion_muebles = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 90, "NO") #funcion para filtar el spoa 
                incautacion_inmuebles = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 92, "NO") #funcion para filtar el spoa 
                incautacion_vehiculos = calculo(valor_calculado[0],y[0],0, 18)


                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 49, "SI") #funcion para filtar el spoa 
                capturas_contrabando = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 62, "SI") #funcion para filtar el spoa 
                combustible_contrabando = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 50, "NO") #funcion para filtar el spoa 
                gasolina_contrabando = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 51, "NO") #funcion para filtar el spoa 
                acpm_contrabando = calculo(valor_calculado[0],y[0],0, 18)

                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 53, "NO") #funcion para filtar el spoa 
                vehiculos_contrando = calculo(valor_calculado[0],y[0],0, 18)

                
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 93, "NO") #funcion para filtar el spoa 
                afectacion_oleoducto = calculo(valor_calculado[0],y[0],0, 18)
                                
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 94, "NO") #funcion para filtar el spoa 
                apique = calculo(valor_calculado[0],y[0],0, 18)
                                
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 96, "NO") #funcion para filtar el spoa 
                aboyadura = calculo(valor_calculado[0],y[0],0, 18)
                                        
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 95, "NO") #funcion para filtar el spoa 
                destruido = calculo(valor_calculado[0],y[0],0, 18)
                                
                valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 97, "NO") #funcion para filtar el spoa 
                saboteado = calculo(valor_calculado[0],y[0],0, 18)
                                        
                # valor_calculado = Calculo_Spoa.validar_spoa_unidad(self, filtro, res_calculo, 98, "NO") #funcion para filtar el spoa 
                total_resultados = calculo(res_calculo[98],y[0],1, 14)


                listados.append(y[0])
                col_merge_final = col_merge_final +1
                # print(col_merge_final)
                hoja.cell(row=1, column=col, value=y[0])
                hoja.cell(row=2, column=col, value=numero_res).number_format = frmt
                hoja.cell(row=3, column=col, value=numero_presentacion).number_format = frmt
                hoja.cell(row=4, column=col, value=sometimientos).number_format = frmt
                hoja.cell(row=5, column=col, value=capturas).number_format = frmt
                hoja.cell(row=6, column=col, value=mdom).number_format = frmt

                hoja.cell(row=7, column=col, value=asesinados).number_format = frmt
                hoja.cell(row=8, column=col, value=heridos).number_format = frmt

                hoja.cell(row=9, column=col, value=largas).number_format = frmt
                hoja.cell(row=10, column=col, value=cortas).number_format = frmt
                hoja.cell(row=11, column=col, value=acompaniamiento).number_format = frmt
                hoja.cell(row=12, column=col, value=municiones).number_format = frmt

                hoja.cell(row=13, column=col, value=combates_pos).number_format = frmt
                hoja.cell(row=14, column=col, value=combates_neg).number_format = frmt
                hoja.cell(row=15, column=col, value=combates_sin_res).number_format = frmt
                hoja.cell(row=16, column=col, value=combates).number_format = frmt

                hoja.cell(row=17, column=col, value=hostigamientos).number_format = frmt
                hoja.cell(row=18, column=col, value=asonadas).number_format = frmt

                hoja.cell(row=19, column=col, value=cocaina).number_format = frmt_d
                hoja.cell(row=20, column=col, value=marihuna).number_format = frmt_d
                hoja.cell(row=21, column=col, value=pbc).number_format = frmt_d
                hoja.cell(row=22, column=col, value=heronia).number_format = frmt_d
                hoja.cell(row=23, column=col, value=basuco).number_format = frmt_d
                hoja.cell(row=24, column=col, value=drogas_sinteticas).number_format = frmt
                hoja.cell(row=25, column=col, value=lab_cocaina).number_format = frmt
                hoja.cell(row=26, column=col, value=lab_pbc).number_format = frmt
                hoja.cell(row=27, column=col, value=lab_heronina).number_format = frmt
                hoja.cell(row=28, column=col, value=semilleros).number_format = frmt
                hoja.cell(row=29, column=col, value=matas_semilleros).number_format = frmt
                hoja.cell(row=30, column=col, value=insumos_solidos).number_format = frmt_d
                hoja.cell(row=31, column=col, value=nsumos_liquidos).number_format = frmt_d
                hoja.cell(row=32, column=col, value=combustible_narcotrafico).number_format = frmt_d

                hoja.cell(row=33, column=col, value=neutralizacion_t).number_format = frmt
                hoja.cell(row=34, column=col, value=artefacto_explosivo).number_format = frmt
                hoja.cell(row=35, column=col, value=ae).number_format = frmt
                hoja.cell(row=36, column=col, value=map).number_format = frmt
                hoja.cell(row=37, column=col, value=explosivos).number_format = frmt_d
                hoja.cell(row=38, column=col, value=explosivos_m).number_format = frmt_d
                hoja.cell(row=39, column=col, value=explosivos_und).number_format = frmt_d
                hoja.cell(row=40, column=col, value=cordon).number_format = frmt_d
                hoja.cell(row=41, column=col, value=mecha).number_format = frmt_d
                hoja.cell(row=42, column=col, value=medios_lanzamientos).number_format = frmt
                hoja.cell(row=43, column=col, value=detonadores_electricos).number_format = frmt
                hoja.cell(row=44, column=col, value=detonadores_anaelectricos).number_format = frmt
                
                hoja.cell(row=45, column=col, value=liberados).number_format = frmt
                hoja.cell(row=46, column=col, value=rescatados).number_format = frmt
                hoja.cell(row=47, column=col, value=valvulas).number_format = frmt
                hoja.cell(row=48, column=col, value=refinerias).number_format = frmt

                hoja.cell(row=49, column=col, value=oleoducto).number_format = frmt
                hoja.cell(row=50, column=col, value=petroleo).number_format = frmt
                hoja.cell(row=51, column=col, value=piscina).number_format = frmt

                hoja.cell(row=52, column=col, value=pesos).number_format = frmt
                hoja.cell(row=53, column=col, value=dolares).number_format = frmt
                hoja.cell(row=54, column=col, value=euros).number_format = frmt

                hoja.cell(row=55, column=col, value=cap_amazonia).number_format = frmt
                hoja.cell(row=56, column=col, value=plantulas).number_format = frmt
                hoja.cell(row=57, column=col, value=madera).number_format = frmt_d
                hoja.cell(row=58, column=col, value=especies).number_format = frmt

                hoja.cell(row=59, column=col, value=cap_menria).number_format = frmt
                hoja.cell(row=60, column=col, value=yacimiento_mineros).number_format = frmt
                hoja.cell(row=61, column=col, value=upm).number_format = frmt
                hoja.cell(row=62, column=col, value=maquinaria_pesada).number_format = frmt
                hoja.cell(row=63, column=col, value=tractor_con_oruga).number_format = frmt
                hoja.cell(row=64, column=col, value=excavadoras).number_format = frmt
                hoja.cell(row=65, column=col, value=retroexcavadoras).number_format = frmt
                hoja.cell(row=66, column=col, value=dragas).number_format = frmt
                hoja.cell(row=67, column=col, value=maquinaria_amarilla).number_format = frmt
                hoja.cell(row=68, column=col, value=motores).number_format = frmt
                hoja.cell(row=69, column=col, value=medios_trasnporte).number_format = frmt
                hoja.cell(row=70, column=col, value=combustible_mineria).number_format = frmt
                hoja.cell(row=71, column=col, value=explosivos_mineria_kg).number_format = frmt
                hoja.cell(row=72, column=col, value=explosivos_mineria_m).number_format = frmt
                hoja.cell(row=73, column=col, value=explosivos_mineria_und).number_format = frmt
                hoja.cell(row=74, column=col, value=coltan).number_format = frmt

                hoja.cell(row=75, column=col, value=incautacion_muebles).number_format = frmt
                hoja.cell(row=76, column=col, value=incautacion_inmuebles).number_format = frmt
                hoja.cell(row=77, column=col, value=incautacion_vehiculos).number_format = frmt

                hoja.cell(row=78, column=col, value=capturas_contrabando).number_format = frmt
                hoja.cell(row=79, column=col, value=combustible_contrabando).number_format = frmt
                hoja.cell(row=80, column=col, value=gasolina_contrabando).number_format = frmt
                hoja.cell(row=81, column=col, value=acpm_contrabando).number_format = frmt
                hoja.cell(row=82, column=col, value=vehiculos_contrando).number_format = frmt
                             
                hoja.cell(row=83, column=col, value=afectacion_oleoducto).number_format = frmt
                hoja.cell(row=84, column=col, value=apique).number_format = frmt
                hoja.cell(row=85, column=col, value=aboyadura).number_format = frmt
                hoja.cell(row=86, column=col, value=destruido).number_format = frmt
                hoja.cell(row=87, column=col, value=saboteado).number_format = frmt
                hoja.cell(row=88, column=col, value=total_resultados).number_format = frmt

                col = col + 1 

   
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 0, "RME", "NO") #funcion para filtar el spoa
        total_rme = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 1, "PRESENTACIONES", "NO" ) #funcion para filtar el spoa
        total_presentacion = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 2, "SOMETIMIENTOS", "NO" ) #funcion para filtar el spoa
        total_sometimientos = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        # print(filtro[17])
        if filtro[17] == "sin_delco":
            valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 61, "CAPTURAS", "SI" ) #funcion para filtar el spoa
        else:
            valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 3, "CAPTURAS", "SI" ) #funcion para filtar el spoa

        # valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 61) #funcion para filtar el spoa
        total_capturas= calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 4,  "MDOM", "SI" ) #funcion para filtar el spoa
        total_mdom = calculo_total_sin_filtro(valor_calculado[0], 18)



        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 5, "HERIDOS", "SI" ) #funcion para filtar el spoa
        total_asesinados = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 6, "ASESINADOS", "SI" ) #funcion para filtar el spoa
        total_heridos = calculo_total_sin_filtro(valor_calculado[0], 18)


        # print(valor_calculado[1])
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 7, "ARMAS LARGAS", "SI" ) #funcion para filtar el spoa
        total_largas = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 8, "ARMAS CORTAS", "SI" ) #funcion para filtar el spoa
        total_cortas = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 9, "ARMAS ACOMPAÑAMIENTO", "SI" ) #funcion para filtar el spoa
        total_acompaniamientos = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 10, "MUNICIONES", "SI" ) #funcion para filtar el spoa
        total_municinoes = calculo_total_sin_filtro(valor_calculado[0], 18)


        total_combates_pos = calculo_total_sin_filtro(res_calculo[11], 14)
        total_combates_neg = calculo_total_sin_filtro(res_calculo[12], 14)
        total_combates_sin_res = calculo_total_sin_filtro(res_calculo[13], 14)
        total_combates = calculo_total_sin_filtro(res_calculo[14], 14)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 66, "HOSTIGAMIENTOS", "SI" ) #funcion para filtar el spoa
        total_hostigamientos   = calculo_total_sin_filtro(valor_calculado[0], 18)
        total_asonadas   = calculo_total_sin_filtro(res_calculo[67], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 21, "COCAINA", "SI" ) #funcion para filtar el spoa
        total_cocaina = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 22, "MARIHUANA", "SI" ) #funcion para filtar el spoa
        total_marihuna = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 23, "PBC", "SI" ) #funcion para filtar el spoa
        total_pbc = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 69, "HEROHINA", "SI" ) #funcion para filtar el spoa
        total_heronia = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 68, "BASUCO", "SI" ) #funcion para filtar el spoa
        total_basuco = calculo_total_sin_filtro(valor_calculado[0], 18)


        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 70, "DROGAS SINTETICAS", "SI" ) #funcion para filtar el spoa
        total_drogas_sinteticas = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 24, "LAB COCIANA", "SI" ) #funcion para filtar el spoa
        total_lab_cocaina = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 25, "LAB PBC", "SI" ) #funcion para filtar el spoa
        total_lab_pbc = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 71 , "LAB HEROHINA", "SI" ) #funcion para filtar el spoa
        total_lab_heronina = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 26 , "SEMILLEROS", "SI" ) #funcion para filtar el spoa
        total_semilleros = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 27 , "MASTA DE COCA EN SEMILLEROS", "SI" ) #funcion para filtar el spoa
        total_matas_semilleros = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 29 , "INSUMOS SOLIDOS", "SI" ) #funcion para filtar el spoa
        total_insumos_solidos = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 28 , "INSUMOS LIQUIDOS", "SI" ) #funcion para filtar el spoa
        total_insumos_liquidos = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 89 , "COMBUSTIBLES NARCOTRAFICO Gal", "SI" ) #funcion para filtar el spoa
        total_combustible_narcotrafico = calculo_total_sin_filtro(valor_calculado[0], 18)


        total_neutralizacion_t = calculo_total_sin_filtro(res_calculo[15], 14)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 80, "ARTEFACTOS EXPLOSIVOS", "NO" ) #funcion para filtar el spoa
        total_artefactos_explosivos = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 16, "AE", "NO" ) #funcion para filtar el spoa
        total_ae = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 17, "MAP", "NO" ) #funcion para filtar el spoa
        total_map = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 18, "EXPLOSIVOS kg", "NO" ) #funcion para filtar el spoa
        total_explosivos = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 78, "EXPLOSIVOS m", "NO" ) #funcion para filtar el spoa
        total_explosivos_m = calculo_total_sin_filtro(valor_calculado[0], 18)
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 79, "EXPLOSIVOS Und", "NO" ) #funcion para filtar el spoa
        total_explosivos_und = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 19, "CORDON DECTONANTE", "NO" ) #funcion para filtar el spoa
        total_cordon = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 20, "MECHA LENTA", "NO" ) #funcion para filtar el spoa
        total_mecha = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 72, "MEDIOS DE LANZAMIENTOS", "NO" ) #funcion para filtar el spoa
        total_medios_lanzamientos = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 73, "DECTONADORES ELECTRICOS", "NO" ) #funcion para filtar el spoa
        total_detonadores_electricos = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 74, "DECTONADORES ANAELECRICOS", "NO" ) #funcion para filtar el spoa
        total_detonadores_anaelectricos = calculo_total_sin_filtro(valor_calculado[0], 18)


        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 42, "LIBERADOS", "SI" ) #funcion para filtar el spoa
        total_liberados = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 43, "RESCATADOS", "SI" ) #funcion para filtar el spoa
        total_rescatados = calculo_total_sin_filtro(valor_calculado[0], 18)


        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 44, "VALVULAS", "NO" ) #funcion para filtar el spoa
        total_valvulas = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 45, "REFINERIAS", "NO" ) #funcion para filtar el spoa
        total_refinerias = calculo_total_sin_filtro(valor_calculado[0], 18)


        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 86, "AFECTACIONES AL OLEODUCTO", "NO" ) #funcion para filtar el spoa
        total_afectaciones_oleoducto = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 87, "PETROLEO", "NO" ) #funcion para filtar el spoa
        total_petroleo = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 88, "PISCINAS", "NO" ) #funcion para filtar el spoa
        total_piscinas = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 76, "PESOS COP", "NO" ) #funcion para filtar el spoa
        total_pesos = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 75, "DOLARES", "NO" ) #funcion para filtar el spoa
        total_dolares = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 77, "EUROS", "NO" ) #funcion para filtar el spoa
        total_euros = calculo_total_sin_filtro(valor_calculado[0], 18)


        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 30, "CAPTURAS AMAZONIA", "SI" ) #funcion para filtar el spoa
        total_cap_amazonia = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 31, "PLANTULAS", "NO" ) #funcion para filtar el spoa
        total_plantulas = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 32, "MADERA INCAUTADA", "NO" ) #funcion para filtar el spoa
        total_madera = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 63, "ANIMALES INCAUTADOS", "NO" ) #funcion para filtar el spoa
        total_especies = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 33, "CAPTURAS MINERIA", "SI" ) #funcion para filtar el spoa
        total_cap_menria = calculo_total_sin_filtro(valor_calculado[0], 18)

        total_yacimiento_mineros = calculo_total_sin_filtro(res_calculo[35], 14)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 40, "UPM", "NO" ) #funcion para filtar el spoa
        total_upm = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 38, "MAQUINARIA PESADA", "NO" ) #funcion para filtar el spoa
        total_maquinaria_pesada = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 39, "BULDOCER", "NO" ) #funcion para filtar el spoa
        total_tractor_con_oruga = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 36, "EXCAVADORAS", "NO" ) #funcion para filtar el spoa
        total_excavadoras = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 37, "RETROEXCAVADORAS", "NO" ) #funcion para filtar el spoa
        total_retroexcavadoras = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 41, "DRAGAS", "NO" ) #funcion para filtar el spoa
        total_dragas = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 59, "MOTORES", "NO" ) #funcion para filtar el spoa
        total_motores = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 85, "MEDIOS DE TRANSPORTE", "NO" ) #funcion para filtar el spoa
        total_material_transporte = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 81, "MAQUINARIA AMARILLA", "NO" ) #funcion para filtar el spoa
        total_maquinaria_amarilla = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 60, "COMBUSTIBLE", "SI" ) #funcion para filtar el spoa
        total_combustible_mineria = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 82, "EXPLOSIVOS MINERIA Kg", "NO" ) #funcion para filtar el spoa
        total_explosivos_mineria_kg = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 83, "EXPLOSIVOS MINERIA m", "NO" ) #funcion para filtar el spoa
        total_explosivos_mineria_m = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 84, "EXPLOSIVOS MINERIA Und", "NO" ) #funcion para filtar el spoa
        total_explosivos_mineria_und = calculo_total_sin_filtro(valor_calculado[0], 18)

        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 65, "COLTAN", "NO" ) #funcion para filtar el spoa
        total_coltan = calculo_total_sin_filtro(valor_calculado[0], 18)

        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 91, "INCAUTACIÓN DE MUEBLES", "NO" ) #funcion para filtar el spoa
        muebles = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 90, "INCAUTACIÓN DE INMUEBLES", "NO" ) #funcion para filtar el spoa
        inmuebles = calculo_total_sin_filtro(valor_calculado[0], 18)
        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 92, "COLINCAUTACIÓN DE VEHICULOS", "NO" ) #funcion para filtar el spoa
        vehiculos = calculo_total_sin_filtro(valor_calculado[0], 18)

                 
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 49, "CAPTURAS CONTRABANDO", "SI" ) #funcion para filtar el spoa
        cap_contrabando = calculo_total_sin_filtro(valor_calculado[0], 18)
                
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 62, "COMBUSTIBLES CONTRABANDO", "SI" ) #funcion para filtar el spoa
        combustible_contrando = calculo_total_sin_filtro(valor_calculado[0], 18)
                
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 50, "GASOLINA CONTRABANDO", "NO" ) #funcion para filtar el spoa
        gasolina_contrabando = calculo_total_sin_filtro(valor_calculado[0], 18)
                
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 51, "ACPM CONTRABANDO", "NO" ) #funcion para filtar el spoa
        acpm_contrabando = calculo_total_sin_filtro(valor_calculado[0], 18)
                
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 53, "VEHICULOS DE CONTRABANDO", "NO" ) #funcion para filtar el spoa
        vehiculos_contrando = calculo_total_sin_filtro(valor_calculado[0], 18)
                 
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 93, "AFECTACION OLEODUCTO", "NO" ) #funcion para filtar el spoa
        afectacion_oleoducto = calculo_total_sin_filtro(valor_calculado[0], 18)
                        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 94, "APIQUE", "NO" ) #funcion para filtar el spoa
        apique = calculo_total_sin_filtro(valor_calculado[0], 18)
                        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 96, "ABOYADURA", "NO" ) #funcion para filtar el spoa
        aboyadura = calculo_total_sin_filtro(valor_calculado[0], 18)
                                
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 95, "ROTURA", "NO" ) #funcion para filtar el spoa
        destruido = calculo_total_sin_filtro(valor_calculado[0], 18)
                        
        valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 97, "SABOTAJE AL OLEODUCTO", "NO" ) #funcion para filtar el spoa
        saboteado = calculo_total_sin_filtro(valor_calculado[0], 18)
                                
        # valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo, 98, "TOTAL DE RESULTADOS", "NO" ) #funcion para filtar el spoa
        total_resultados = calculo_total_sin_filtro(res_calculo[98], 14)

        # col= 2
        hoja.cell(row=1, column=col, value="TOTAL")
        hoja.cell(row=2, column=col, value=total_rme).number_format = frmt
        hoja.cell(row=3, column=col, value=total_presentacion).number_format = frmt
        hoja.cell(row=4, column=col, value=total_sometimientos).number_format = frmt
        hoja.cell(row=5, column=col, value=total_capturas).number_format = frmt
        hoja.cell(row=6, column=col, value=total_mdom).number_format = frmt

        hoja.cell(row=7, column=col, value=total_asesinados).number_format = frmt
        hoja.cell(row=8, column=col, value=total_heridos).number_format = frmt

        hoja.cell(row=9, column=col, value=total_largas).number_format = frmt
        hoja.cell(row=10, column=col, value=total_cortas).number_format = frmt
        hoja.cell(row=11, column=col, value=total_acompaniamientos).number_format = frmt
        hoja.cell(row=12, column=col, value=total_municinoes).number_format = frmt

        hoja.cell(row=13, column=col, value=total_combates_pos).number_format = frmt
        hoja.cell(row=14, column=col, value=total_combates_neg).number_format = frmt
        hoja.cell(row=15, column=col, value=total_combates_sin_res).number_format = frmt
        hoja.cell(row=16, column=col, value=total_combates).number_format = frmt

        hoja.cell(row=17, column=col, value=total_hostigamientos).number_format = frmt
        hoja.cell(row=18, column=col, value=total_asonadas).number_format = frmt

        hoja.cell(row=19, column=col, value=total_cocaina).number_format = frmt_d
        hoja.cell(row=20, column=col, value=total_marihuna).number_format = frmt_d
        hoja.cell(row=21, column=col, value=total_pbc).number_format = frmt_d
        hoja.cell(row=22, column=col, value=total_heronia).number_format = frmt_d
        hoja.cell(row=23, column=col, value=total_basuco).number_format = frmt_d
        hoja.cell(row=24, column=col, value=total_drogas_sinteticas).number_format = frmt
        hoja.cell(row=25, column=col, value=total_lab_cocaina).number_format = frmt
        hoja.cell(row=26, column=col, value=total_lab_pbc).number_format = frmt
        hoja.cell(row=27, column=col, value=total_lab_heronina).number_format = frmt
        hoja.cell(row=28, column=col, value=total_semilleros).number_format = frmt
        hoja.cell(row=29, column=col, value=total_matas_semilleros).number_format = frmt
        hoja.cell(row=30, column=col, value=total_insumos_solidos).number_format = frmt_d
        hoja.cell(row=31, column=col, value=total_insumos_liquidos).number_format = frmt_d
        hoja.cell(row=32, column=col, value=total_combustible_narcotrafico).number_format = frmt_d
       
        hoja.cell(row=33, column=col, value=total_neutralizacion_t).number_format = frmt
        hoja.cell(row=34, column=col, value=total_artefactos_explosivos).number_format = frmt
        hoja.cell(row=35, column=col, value=total_ae).number_format = frmt
        hoja.cell(row=36, column=col, value=total_map).number_format = frmt
        hoja.cell(row=37, column=col, value=total_explosivos).number_format = frmt_d
        hoja.cell(row=38, column=col, value=total_explosivos_m).number_format = frmt_d
        hoja.cell(row=39, column=col, value=total_explosivos_und).number_format = frmt_d

        hoja.cell(row=40, column=col, value=total_cordon).number_format = frmt_d
        hoja.cell(row=41, column=col, value=total_mecha).number_format = frmt_d
        hoja.cell(row=42, column=col, value=total_medios_lanzamientos).number_format = frmt
        hoja.cell(row=43, column=col, value=total_detonadores_electricos).number_format = frmt
        hoja.cell(row=44, column=col, value=total_detonadores_anaelectricos).number_format = frmt

        hoja.cell(row=45, column=col, value=total_liberados).number_format = frmt
        hoja.cell(row=46, column=col, value=total_rescatados).number_format = frmt
        hoja.cell(row=47, column=col, value=total_valvulas).number_format = frmt
        hoja.cell(row=48, column=col, value=total_refinerias).number_format = frmt

        hoja.cell(row=49, column=col, value=total_afectaciones_oleoducto).number_format = frmt
        hoja.cell(row=50, column=col, value=total_petroleo).number_format = frmt
        hoja.cell(row=51, column=col, value=total_piscinas).number_format = frmt

        hoja.cell(row=52, column=col, value=total_pesos).number_format = frmt
        hoja.cell(row=53, column=col, value=total_dolares).number_format = frmt
        hoja.cell(row=54, column=col, value=total_euros).number_format = frmt
    
        hoja.cell(row=55, column=col, value=total_cap_amazonia).number_format = frmt
        hoja.cell(row=56, column=col, value=total_plantulas).number_format = frmt
        hoja.cell(row=57, column=col, value=total_madera).number_format = frmt_d
        hoja.cell(row=58, column=col, value=total_especies).number_format = frmt

        hoja.cell(row=59, column=col, value=total_cap_menria).number_format = frmt
        hoja.cell(row=60, column=col, value=total_yacimiento_mineros).number_format = frmt
        hoja.cell(row=61, column=col, value=total_upm).number_format = frmt
        hoja.cell(row=62, column=col, value=total_maquinaria_pesada).number_format = frmt
        hoja.cell(row=63, column=col, value=total_tractor_con_oruga).number_format = frmt
        hoja.cell(row=64, column=col, value=total_excavadoras).number_format = frmt
        hoja.cell(row=65, column=col, value=total_retroexcavadoras).number_format = frmt
        hoja.cell(row=66, column=col, value=total_dragas).number_format = frmt
        hoja.cell(row=67, column=col, value=total_maquinaria_amarilla).number_format = frmt
        hoja.cell(row=68, column=col, value=total_motores).number_format = frmt
        hoja.cell(row=69, column=col, value=total_material_transporte).number_format = frmt
        hoja.cell(row=70, column=col, value=total_combustible_mineria).number_format = frmt
        hoja.cell(row=71, column=col, value=total_explosivos_mineria_kg).number_format = frmt
        hoja.cell(row=72, column=col, value=total_explosivos_mineria_m).number_format = frmt
        hoja.cell(row=73, column=col, value=total_explosivos_mineria_und).number_format = frmt
        hoja.cell(row=74, column=col, value=total_coltan).number_format = frmt

        hoja.cell(row=75, column=col, value=muebles).number_format = frmt
        hoja.cell(row=76, column=col, value=inmuebles).number_format = frmt
        hoja.cell(row=77, column=col, value=vehiculos).number_format = frmt

        hoja.cell(row=78, column=col, value=cap_contrabando).number_format = frmt
        hoja.cell(row=79, column=col, value=combustible_contrando).number_format = frmt
        hoja.cell(row=80, column=col, value=gasolina_contrabando).number_format = frmt
        hoja.cell(row=81, column=col, value=acpm_contrabando).number_format = frmt
        hoja.cell(row=82, column=col, value=vehiculos_contrando).number_format = frmt
  
        hoja.cell(row=83, column=col, value=afectacion_oleoducto).number_format = frmt
        hoja.cell(row=84, column=col, value=apique).number_format = frmt
        hoja.cell(row=85, column=col, value=aboyadura).number_format = frmt
        hoja.cell(row=86, column=col, value=destruido).number_format = frmt
        hoja.cell(row=87, column=col, value=saboteado).number_format = frmt
        hoja.cell(row=88, column=col, value=total_resultados).number_format = frmt

        for cell in hoja["1:1"]: 
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill  = PatternFill('solid', start_color="800000")
                # cell.alignment = Alignment(wrap_text=True)
                cell.alignment=Alignment(
                            horizontal='center',
                            vertical='top',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)
                

    
