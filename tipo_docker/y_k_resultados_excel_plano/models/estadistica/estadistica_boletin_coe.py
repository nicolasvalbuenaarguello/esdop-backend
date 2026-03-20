# El presente modulo es para el manejo de la estadística de la Dirección de operaciones del Ejército Nacional 
from tipo_docker.y_k_resultados_excel_plano.models.conexion_pos import Databa_bases
from tipo_docker.z_z_z_funtions.funciones import *
from tipo_docker.y_k_resultados_excel_plano.models.funtions.componest.tablas import *

conexion_pos = Databa_bases()
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.descriptors.serialisable import Serialisable
import pandas as pd
from openpyxl.styles import Font, colors, fills, Alignment, PatternFill, NamedStyle

import os
from dotenv import load_dotenv
load_dotenv()  

import zipfile

class Calculo_Spoa:
    def __init__(self):
            self.LINK = os.getenv('DIRECION_3_B')
            self.archivo = "RESULTADO_SIN_SPOA" +".txt"              
            self.f = open(self.LINK+self.archivo, "w")
            self.f.write(str("RESULTADO")+", " + str("HR")+", " +str("FECHA")+", "+str("SPOA")+", "+str("DIV")+"\n")

    def validar_spoa_unidad(self, filtro, res_calculo, validar, fecha):
        spoa=[]
        no_spoa=[]
        
        if validar == "SI":
            
            startdate = pd.to_datetime(fecha).date()
            # startdate = pd.to_datetime("2024-01-01").date()
            
            if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
                for x in res_calculo:
            
                    if x[0] >= startdate:

                        # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                        if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                            no_spoa.append(x)
                            # print(x[29])
                        else: 
                            numero_dep = x[29][0:5]
                                
                            if numero_dep == "00000":
                                no_spoa.append(x)
                            else:
                                spoa.append(x)
                            # spoa.append(x)

                    else: 
                        spoa.append(x)      
            else:
                spoa =  res_calculo

            spoa = spoa
        else:
            spoa =  res_calculo

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        return[spoa, no_spoa]

    def validar_spoa(self, filtro, res_calculo,  validar, fecha):
        spoa=[]
        no_spoa=[]

        
        startdate = pd.to_datetime(str(fecha)).date()
        # startdate = pd.to_datetime("2024-01-01").date()
        y = []
        print(filtro[16])
        if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
            for x in res_calculo:
                # print(str(len(x))+" - "+str(x[29]))
                numero_dep = x[29][0:5]
                if x[0] >= startdate:

                    # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                    if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                        y= list(x)
                        y.append("CON ERROR")
                        f = tuple(y)
                        no_spoa.append(f)
                        # print(x[29])       
                    elif numero_dep == "00000":
                                y= list(x)
                                y.append("CON ERROR")
                                f = tuple(y)
                                no_spoa.append(f)
                    else:
                                y= list(x)
                                y.append("OK SPOA")
                                f = tuple(y)
                                no_spoa.append(f)
                        # spoa.append(x)

                else: 
                    y= list(x)
                    y.append("OK SPOA")
                    f = tuple(y)
                    no_spoa.append(f)
                        # spoa.append(x)      
        else:
            spoa =  res_calculo

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        
        # f.write(str(nombre)+"\n")
        # f.write("-----------------------------------------------------"+"\n")
        
        numero_id = 1
        # print(no_spoa)
        # for x in no_spoa:
            
        #     no_spoa[29]
        #     self.f.write(str(x[26])+", "+str(x[28])+", "+ str(x[0])+", " +str(x[1])+", "+str(x[2])+", "+str(x[3])+", "+str(x[4])+", "+str(x[5])+", "+str(x[6])+", "+str(x[7])+", "+str(x[8])+", "+str(x[9])+", "+str(x[20])+", "+str(x[21])+", "+str(x[27])+", "+    str(x[29])+", "+str(x[30])+", "+str(x[31])+", "+str(x[32])+", "+str(x[33])+"\n")


        # f.write("-----------------------------------------------------"+"\n")
        # f.write(str(numero_id-1)+" Resultados sin SPOA"+"\n")

        if validar == "SI":
            spoa = spoa
        else:
            spoa =  res_calculo
            

        return[spoa, no_spoa]

    # cuadro de resultados resaltantes tamaño oficio
    def resultados_resaltantes_pdf_oficio(self, fecha_inicial_u_l, fecha_final_u_l, filtro,  dato, wb):

        dato =""
        if  filtro[4] == "lugar":
            if filtro[6]!="" and filtro[6]!="---" :#filtro por municipio
                nueva=filtro[6].split(",")
                dato = ""
                dato_res=""
                if len(nueva) > 1:
                        ids = tuple(nueva)
                        dato = "and dpto = '{}' and mpio in {}".format(filtro[5], ids)
                        dato_res= "and hop_depto = '{}' and hop_mpio in {}".format(filtro[5], ids)

                else:
                    mpio = nueva[0]  
                    dato = "and {} = '{}' and {} = '{}'".format("dpto",filtro[5], "mpio",mpio)
                    dato_res = "and {} = '{}' and {} = '{}'".format("hop_depto", filtro[5], "hop_mpio",mpio)
                filtros =[dato_res, dato, ""]
            else:
                if dato == "":
                    filtros = selecion_filtro(filtro)
                else:
                    filtros = dato
        else:
            if dato == "":
                filtros = selecion_filtro(filtro)
            else:
                filtros = dato

        query = parametros(fecha_inicial_u_l, fecha_final_u_l, filtros, filtro)

        resultados = conexion_pos.comando_query(query[0])
        hechos = conexion_pos.comando_query(query[1])
        erradicacion = conexion_pos.comando_query(query[2])

        res_calculo = estadistica_resultados(resultados, hechos, filtro)
        
        # print(filtro)

        # agr_div
        # division
        # brigada
        # unidad
        # dpto
        # mpio

                
        BUILTIN_FORMATS = {
                                0: 'General',
                                1: '0',
                                2: '0.00',
                                3: '#,##0',
                                4: '#,##0.00',
                                5: '"$"#,##0_);("$"#,##0)',
                                6: '"$"#,##0_);[Red]("$"#,##0)',
                                7: '"$"#,##0.00_);("$"#,##0.00)',
                                8: '"$"#,##0.00_);[Red]("$"#,##0.00)',
                                9: '0%',
                                10: '0.00%',
                                11: '0.00E+00',
                                12: '# ?/?',
                                13: '# ??/??',
                                14: 'mm-dd-yy',
                                15: 'd-mmm-yy',
                                16: 'd-mmm',
                                17: 'mmm-yy',
                                18: 'h:mm AM/PM',
                                19: 'h:mm:ss AM/PM',
                                20: 'h:mm',
                                21: 'h:mm:ss',
                                22: 'm/d/yy h:mm',

                                37: '#,##0_);(#,##0)',
                                38: '#,##0_);[Red](#,##0)',
                                39: '#,##0.00_);(#,##0.00)',
                                40: '#,##0.00_);[Red](#,##0.00)',

                                41: r'_(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)',
                                42: r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)',
                                43: r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)',

                                44: r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)',
                                45: 'mm:ss',
                                46: '[h]:mm:ss',
                                47: 'mmss.0',
                                48: '##0.0E+0',
                                49: '@', }
            
        frmt = BUILTIN_FORMATS[3] 
        frmt_d = BUILTIN_FORMATS[4]   
            # for row in hoja["B:U"]:
            #     cells = row[0], row[-1]
            #     for c in cells:
            #         c.number_format = frmt

        # print(filtro)
        hechos_filtro = ""
        nombre_hoja = "---"


        
        if filtro[4] == 'unidad':
            if filtro[3]!="" and filtro[3]!="---":
                hechos_filtro = " and {} = '{}'".format("unidad",filtro[3])
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[3]
            elif filtro[2]!="" and filtro[2]!="---":
                hechos_filtro = " and {} = '{}'".format("brigada",filtro[2])
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[2]
            elif filtro[1]!="" and filtro[1]!="---":
                hechos_filtro = " and {} = '{}'".format("division",filtro[1])
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[1]
            elif filtro[0]!="" and filtro[0]!="---":
                hechos_filtro = " and {} = '{}'".format("agr_div",filtro[0])
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[0]

        elif filtro[4] == 'lugar':
            if filtro[6]!="" and filtro[6]!="---" :#filtro por municipio
                nueva=filtro[6].split(",")
                dato = ""
                dato_res=""
                if len(nueva) > 1:
                        ids = tuple(nueva)
                        hechos_filtro = "and mpio in {}".format(ids)

                else:
                    mpio = nueva[0]  
                    hechos_filtro = " and {} = '{}'".format("mpio",mpio)

                mpio = nueva[0]  
                # hechos_filtro = " and {} = '{}'".format("mpio",mpio)
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[6]

            elif filtro[5]!= "---" :#filtro por departamento
                hechos_filtro = " and {} = '{}'".format("dpto",filtro[5])
                nombre_hoja = "ARCHIVO REVISTA DE " + filtro[5]

        if nombre_hoja =="---":
            nombre_hoja = "ARCHIVO REVISTA DE EJC"


        if filtro[19] == "gaulas":
            print(filtro[19])
            hechos_filtro = hechos_filtro + str(" and unidad like 'GG%'")
            nombre_hoja = "ARCHIVO REVISTA DE GAULAS"
            


            
        if filtro[17] == "sin_delco":
            valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo[61], "SI", fecha_inicial_u_l) #funcion para filtar el spoa
        else:
            valor_calculado = Calculo_Spoa.validar_spoa(self, filtro, res_calculo[3], "SI", fecha_inicial_u_l) #funcion para filtar el spoa

        hojas=0



        # font = Font(name='Arial',
        #             size=8,
        #             bold=True,
        #             color="FFFFFF")

        fill = PatternFill(fill_type=fills.FILL_SOLID,
                        start_color="800000")
        alignment = Alignment(wrap_text=True)


        hoja = wb.active       
        hoja.title = nombre_hoja
        numero  =  1
        
        hoja.cell(row=numero, column=1, value="CAPTURAS REPORTADAS") 
        hoja.merge_cells(start_row=1, end_row=1, start_column=1, end_column=21)



        hoja.cell(row=numero, column=22, value="CON MEDIDA DE ASEGURAMIENTO") 
        hoja.merge_cells(start_row=1, end_row=1, start_column=22, end_column=23)

        hoja.cell(row=numero, column=24, value="PRIVATIVO DE LA LIBERTAD") 
        hoja.merge_cells(start_row=1, end_row=1, start_column=24, end_column=25)

        
        hoja.cell(row=numero, column=26, value="INFORMACION") 
        hoja.merge_cells(start_row=1, end_row=1, start_column=26, end_column=26)
                
        hoja.cell(row=numero, column=27, value="OBSERVACIONES") 
        hoja.merge_cells(start_row=1, end_row=2, start_column=27, end_column=27)

        numero = numero +1

        hoja.cell(row=numero, column=1, value=str("HR"))
        hoja.cell(row=numero, column=2, value=str("FECHA"))
        hoja.cell(row=numero, column=3, value=str("DIV"))
        hoja.cell(row=numero, column=4, value=str('DIV_FT'))
        hoja.cell(row=numero, column=5, value=str('BR'))
        hoja.cell(row=numero, column=6, value=str('UT'))
        hoja.cell(row=numero, column=7, value=str('DEPARTAMENTO'))
        hoja.cell(row=numero, column=8, value=str('MUNICIPIO'))
        hoja.cell(row=numero, column=9, value=str('AMENAZA'))
        hoja.cell(row=numero, column=10, value=str('LOE - LOO'))
        hoja.cell(row=numero, column=11, value=str('ORDOP'))
        hoja.cell(row=numero, column=12, value=str('RESULTADOS'))
        hoja.cell(row=numero, column=13, value=str('NOMBRE'))
        hoja.cell(row=numero, column=14, value=str('CEDULA'))
        hoja.cell(row=numero, column=15, value=str('CANTIDAD'))
        
        hoja.cell(row=numero, column=16, value=str('SPOA'))
        hoja.cell(row=numero, column=17, value=str('COORDINADA'))
        hoja.cell(row=numero, column=18, value=str('CONJUNTA'))
        
        hoja.cell(row=numero, column=19, value=str('RESUMEN'))
        hoja.cell(row=numero, column=20, value=str('ALERTA'))

        hoja.cell(row=numero, column=21, value=str('ESTADO DEL PROCESO'))

        hoja.cell(row=numero, column=22, value=str('SI'))
        hoja.cell(row=numero, column=23, value=str('NO'))
        
        hoja.cell(row=numero, column=24, value=str('SI'))
        hoja.cell(row=numero, column=25, value=str('NO'))
                
        hoja.cell(row=numero, column=26, value=str('CENTRO DE RECLUSION'))
        
        


        numero = numero +1
        for x in  valor_calculado[1]:
             
            hoja.cell(row=numero, column=1, value=str(x[26]))
            hoja.cell(row=numero, column=2, value=str(x[0]))
            hoja.cell(row=numero, column=3, value=str(x[1]))
            hoja.cell(row=numero, column=4, value=str(x[2]))
            hoja.cell(row=numero, column=5, value=str(x[3]))
            hoja.cell(row=numero, column=6, value=str(x[4]))
            hoja.cell(row=numero, column=7, value=str(x[5]))
            hoja.cell(row=numero, column=8, value=str(x[6]))
            hoja.cell(row=numero, column=9, value=str(x[7]))
            hoja.cell(row=numero, column=10, value=str(x[8]))
            hoja.cell(row=numero, column=11, value=str(x[33]))
            hoja.cell(row=numero, column=12, value=str(x[9]))
            hoja.cell(row=numero, column=13, value=str(x[20]))
            hoja.cell(row=numero, column=14, value=str(x[21]))
            hoja.cell(row=numero, column=15, value=x[18]).number_format = frmt
            
            spoa = "'"+str(x[29])
            hoja.cell(row=numero, column=16, value=spoa)

            hoja.cell(row=numero, column=17, value=str(x[30]))
            hoja.cell(row=numero, column=18, value=str(x[31]))
            hoja.cell(row=numero, column=19, value=str(x[27]))
            cell= hoja.cell(row=numero, column=20, value=str(x[34]))

            if str(x[34]) == "CON ERROR":
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill  = PatternFill('solid', start_color="800000")


            numero = numero + 1
        #RESULTADOS AFECTACION AL ENEMIGO
        
        # hoja.cell(row=3, column=1, value="MENORES RECUPERADOS")
        # hoja.cell(row=4, column=1, value="PRESENTACION VOLUNTARIA")
        # hoja.cell(row=5, column=1, value="SOMETIMIENTOS A LA JUSTICIA")
        # hoja.cell(row=6, column=1, value="CAPTURAS")
        # hoja.cell(row=7, column=1, value="MDOM")
        # print(valor_calculado[0])


        # print(filtro[17])


        for cell in hoja["1:1"]: 
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill  = PatternFill('solid', start_color="800000")
                # cell.alignment = Alignment(wrap_text=True)
            cell.alignment=Alignment(
                horizontal='center',
                vertical='top',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
            
        for cell in hoja["2:2"]: 
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill  = PatternFill('solid', start_color="5A616B")
                # cell.alignment = Alignment(wrap_text=True)
            cell.alignment=Alignment(
                horizontal='center',
                vertical='top',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
      

                


    
