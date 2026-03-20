# El presente modulo es para el manejo de la estadística de la Dirección de operaciones del Ejército Nacional 
from tipo_docker.y_m_resultados_excel_resultados.models.conexion_pos import Databa_bases
from tipo_docker.z_z_z_funtions.funciones import *
from tipo_docker.y_m_resultados_excel_resultados.models.funtions.componest.tablas import *

conexion_pos = Databa_bases()
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.descriptors.serialisable import Serialisable
import pandas as pd
from openpyxl.styles import Font, colors, fills, Alignment, PatternFill, NamedStyle
import numpy as np

import os
from dotenv import load_dotenv
load_dotenv()  
import time

import zipfile

class Calculo_Spoa:
    def __init__(self):
            self.LINK = os.getenv('DIRECION_3_B')
            self.archivo = "RESULTADO_SIN_SPOA" +".txt"              
            self.f = open(self.LINK+self.archivo, "w")
            self.f.write(str("RESULTADO")+", " + str("HR")+", " +str("FECHA")+", "+str("SPOA")+", "+str("DIV")+"\n")

    def validar_spoa_unidad(self, filtro, res_calculo, validar, fecha):
        spoa=[]
        no_spoa=[]
        
        if validar == "SI":
            
            startdate = pd.to_datetime(fecha).date()
            # startdate = pd.to_datetime("2024-01-01").date()
            
            if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
                for x in res_calculo:
            
                    if x[0] >= startdate:

                        # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                        if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                            no_spoa.append(x)
                            # print(x[29])
                        else: 
                            numero_dep = x[29][0:5]
                                
                            if numero_dep == "00000":
                                no_spoa.append(x)
                            else:
                                spoa.append(x)
                            # spoa.append(x)

                    else: 
                        spoa.append(x)      
            else:
                spoa =  res_calculo

            spoa = spoa
        else:
            spoa =  res_calculo

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        return[spoa, no_spoa]

    def validar_spoa(self, filtro, res_calculo,  validar, fecha):
        spoa=[]
        no_spoa=[]

        
        startdate = pd.to_datetime(str(fecha)).date()
        # startdate = pd.to_datetime("2024-01-01").date()
        y = []
        print(filtro[16])
        if filtro[16] == "sin_spoa" or filtro[16] == "res_sin_spoa" :
            for x in res_calculo:
                # print(str(len(x))+" - "+str(x[29]))
                numero_dep = x[29][0:5]
                if x[0] >= startdate:

                    # if x[29] =="-" or  x[29] =="" or  x[29] =="0":
                    if x[29] =="-" or  x[29] =="" or len(x[29]) != 21 or  x[29] =="0":
                        y= list(x)
                        y.append("CON ERROR")
                        f = tuple(y)
                        no_spoa.append(f)
                        # print(x[29])       
                    elif numero_dep == "00000":
                                y= list(x)
                                y.append("CON ERROR")
                                f = tuple(y)
                                no_spoa.append(f)
                    else:
                                y= list(x)
                                y.append("OK SPOA")
                                f = tuple(y)
                                no_spoa.append(f)
                        # spoa.append(x)

                else: 
                    y= list(x)
                    y.append("OK SPOA")
                    f = tuple(y)
                    no_spoa.append(f)
                        # spoa.append(x)      
        else:
            spoa =  res_calculo

        if filtro[16]== "res_sin_spoa":
                    # print(len(no_spoa))
                    spoa = no_spoa
        #DOCUMENTOS SIN SPOA

        
        # f.write(str(nombre)+"\n")
        # f.write("-----------------------------------------------------"+"\n")
        
        numero_id = 1
        # print(no_spoa)
        # for x in no_spoa:
            
        #     no_spoa[29]
        #     self.f.write(str(x[26])+", "+str(x[28])+", "+ str(x[0])+", " +str(x[1])+", "+str(x[2])+", "+str(x[3])+", "+str(x[4])+", "+str(x[5])+", "+str(x[6])+", "+str(x[7])+", "+str(x[8])+", "+str(x[9])+", "+str(x[20])+", "+str(x[21])+", "+str(x[27])+", "+    str(x[29])+", "+str(x[30])+", "+str(x[31])+", "+str(x[32])+", "+str(x[33])+"\n")


        # f.write("-----------------------------------------------------"+"\n")
        # f.write(str(numero_id-1)+" Resultados sin SPOA"+"\n")

        if validar == "SI":
            spoa = spoa
        else:
            spoa =  res_calculo
            

        return[spoa, no_spoa]

    # cuadro de resultados resaltantes tamaño oficio
    def resultados_resaltantes_pdf_oficio(self, fecha_inicial_u_l, fecha_final_u_l, filtro,  dato, wb):

        dato =""
        filtros = selecion_filtro(filtro)
        query = parametros(fecha_inicial_u_l, fecha_final_u_l, filtros, filtro)

        resultados = conexion_pos.comando_query(query[0])
        hechos = conexion_pos.comando_query(query[1])
        erradicacion = conexion_pos.comando_query(query[2])

        
        # print(filtro)

        # agr_div
        # division
        # brigada
        # unidad
        # dpto
        # mpio

                
        BUILTIN_FORMATS = {
                                0: 'General',
                                1: '0',
                                2: '0.00',
                                3: '#,##0',
                                4: '#,##0.00',
                                5: '"$"#,##0_);("$"#,##0)',
                                6: '"$"#,##0_);[Red]("$"#,##0)',
                                7: '"$"#,##0.00_);("$"#,##0.00)',
                                8: '"$"#,##0.00_);[Red]("$"#,##0.00)',
                                9: '0%',
                                10: '0.00%',
                                11: '0.00E+00',
                                12: '# ?/?',
                                13: '# ??/??',
                                14: 'mm-dd-yy',
                                15: 'd-mmm-yy',
                                16: 'd-mmm',
                                17: 'mmm-yy',
                                18: 'h:mm AM/PM',
                                19: 'h:mm:ss AM/PM',
                                20: 'h:mm',
                                21: 'h:mm:ss',
                                22: 'm/d/yy h:mm',

                                37: '#,##0_);(#,##0)',
                                38: '#,##0_);[Red](#,##0)',
                                39: '#,##0.00_);(#,##0.00)',
                                40: '#,##0.00_);[Red](#,##0.00)',

                                41: r'_(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)',
                                42: r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)',
                                43: r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)',

                                44: r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)',
                                45: 'mm:ss',
                                46: '[h]:mm:ss',
                                47: 'mmss.0',
                                48: '##0.0E+0',
                                49: '@', }
            
        frmt = BUILTIN_FORMATS[3] 
        frmt_d = BUILTIN_FORMATS[4]   
            # for row in hoja["B:U"]:
            #     cells = row[0], row[-1]
            #     for c in cells:
            #         c.number_format = frmt

        # print(filtro)
        hechos_filtro = ""
        nombre_hoja = "---"


        nombre_hoja = "HECHOS"

        hojas=0


        # font = Font(name='Arial',
        #             size=8,
        #             bold=True,
        #             color="FFFFFF")

        fill = PatternFill(fill_type=fills.FILL_SOLID,
                        start_color="800000")
        alignment = Alignment(wrap_text=True)


        hoja = wb.active       
        hoja.title = nombre_hoja
        numero  =  1

        hoja.cell(row=numero, column=1, value=str("HR"))
        hoja.cell(row=numero, column=2, value=str("BOLETIN"))
        hoja.cell(row=numero, column=3, value=str("HECHOS"))
        hoja.cell(row=numero, column=4, value=str("FECHA"))
        hoja.cell(row=numero, column=5, value=str("HORA"))
        hoja.cell(row=numero, column=6, value=str("DIV"))
        hoja.cell(row=numero, column=7, value=str('DIV_FT'))
        hoja.cell(row=numero, column=8, value=str('BR'))
        hoja.cell(row=numero, column=9, value=str('UT'))
        hoja.cell(row=numero, column=10, value=str('DEPARTAMENTO'))
        hoja.cell(row=numero, column=11, value=str('MUNICIPIO'))
        hoja.cell(row=numero, column=12, value=str('LUGAR'))
        hoja.cell(row=numero, column=13, value=str('SITIO'))
        hoja.cell(row=numero, column=14, value=str('AMENAZA'))
        hoja.cell(row=numero, column=15, value=str('OPERACION'))
        hoja.cell(row=numero, column=16, value=str('ORDOP'))
        hoja.cell(row=numero, column=17, value=str('RESUMEN'))
        hoja.cell(row=numero, column=18, value=str('LATITUD'))
        hoja.cell(row=numero, column=19, value=str('LONGITUD'))

        hoja.cell(row=numero, column=20, value=str('ESTRATEGIA AFECTA'))
        hoja.cell(row=numero, column=21, value=str('APOYO DAAVA'))
        hoja.cell(row=numero, column=22, value=str('APOYO BLICA'))
        hoja.cell(row=numero, column=23, value=str('APOYO CONAT'))
        hoja.cell(row=numero, column=24, value=str('HECHO POS'))

        hoja.cell(row=numero, column=25, value=str('ACCION ENEMIGA'))
        hoja.cell(row=numero, column=26, value=str('COMPAÑIA'))
        hoja.cell(row=numero, column=27, value=str('PELOTON'))
        hoja.cell(row=numero, column=28, value=str('COMANDANTE'))
        hoja.cell(row=numero, column=29, value=str('OPE COORDINADA'))
        hoja.cell(row=numero, column=30, value=str('OPE. CONJUNTA'))
        hoja.cell(row=numero, column=31, value=str('TIPO OPERACION'))

        hoja.cell(row=numero, column=32, value=str('INICIATIVA'))
        hoja.cell(row=numero, column=33, value=str('CLASE SOLDADO'))
        hoja.cell(row=numero, column=34, value=str('PROCE. INFO'))
        hoja.cell(row=numero, column=35, value=str('HOP EXITO'))
        hoja.cell(row=numero, column=36, value=str('CANTIDAD'))

        numero = numero +1
        #hechos =np.array(hechos)

        for x in hechos:
             
            hoja.cell(row=numero, column=1, value=str(x[18]))
            hoja.cell(row=numero, column=2, value=str(x[29]))
            hoja.cell(row=numero, column=3, value=str(x[0]))
            hoja.cell(row=numero, column=4, value=str(x[1]))
            hoja.cell(row=numero, column=5, value=str(x[35]))
            hoja.cell(row=numero, column=6, value=str(x[2]))
            hoja.cell(row=numero, column=7, value=str(x[3]))
            hoja.cell(row=numero, column=8, value=str(x[4]))
            hoja.cell(row=numero, column=9, value=str(x[5]))
            hoja.cell(row=numero, column=10, value=str(x[6]))
            hoja.cell(row=numero, column=11, value=str(x[7]))
            hoja.cell(row=numero, column=12, value=str(x[30]))
            hoja.cell(row=numero, column=13, value=str(x[24]))
            hoja.cell(row=numero, column=14, value=str(x[8]))
            hoja.cell(row=numero, column=15, value=str(x[15]))
            hoja.cell(row=numero, column=16, value=str(x[28]))
            hoja.cell(row=numero, column=17, value=str(x[19]))
            hoja.cell(row=numero, column=18, value=str(x[16]))
            hoja.cell(row=numero, column=19, value=str(x[17]))

            hoja.cell(row=numero, column=20, value=str(x[9]))
            hoja.cell(row=numero, column=21, value=str(x[10]))
            hoja.cell(row=numero, column=22, value=str(x[11]))
            hoja.cell(row=numero, column=23, value=str(x[12]))
            hoja.cell(row=numero, column=24, value=str(x[13]))

            hoja.cell(row=numero, column=25, value=str(x[20]))
            hoja.cell(row=numero, column=26, value=str(x[21]))
            hoja.cell(row=numero, column=27, value=str(x[22]))
            hoja.cell(row=numero, column=28, value=str(x[23]))
            hoja.cell(row=numero, column=29, value=str(x[25]))
            hoja.cell(row=numero, column=30, value=str(x[26]))
            hoja.cell(row=numero, column=31, value=str(x[27]))

            hoja.cell(row=numero, column=32, value=str(x[31]))
            hoja.cell(row=numero, column=33, value=str(x[32]))
            hoja.cell(row=numero, column=34, value=str(x[33]))
            hoja.cell(row=numero, column=35, value=str(x[34]))
            hoja.cell(row=numero, column=36, value=x[14]).number_format = frmt


            numero = numero + 1


        for cell in hoja["1:1"]: 
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill  = PatternFill('solid', start_color="800000")
                # cell.alignment = Alignment(wrap_text=True)
            cell.alignment=Alignment(
                horizontal='center',
                vertical='top',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
            
        hoja = wb.create_sheet("RESULTADOS")
        numero  =  1

        hoja.cell(row=numero, column=1, value=str("HR"))
        hoja.cell(row=numero, column=2, value=str("BOLETIN"))
        hoja.cell(row=numero, column=3, value=str("HECHO"))
        hoja.cell(row=numero, column=4, value=str("FECHA"))
        hoja.cell(row=numero, column=5, value=str("HORA"))
        hoja.cell(row=numero, column=6, value=str("DIV"))
        hoja.cell(row=numero, column=7, value=str('DIV_FT'))
        hoja.cell(row=numero, column=8, value=str('BR'))
        hoja.cell(row=numero, column=9, value=str('UT'))
        hoja.cell(row=numero, column=10, value=str('DEPARTAMENTO'))
        hoja.cell(row=numero, column=11, value=str('MUNICIPIO'))
        hoja.cell(row=numero, column=12, value=str('LUGAR'))
        hoja.cell(row=numero, column=13, value=str('SITIO'))
        hoja.cell(row=numero, column=14, value=str('AMENAZA'))
        hoja.cell(row=numero, column=15, value=str('ESTRUTURA'))
        hoja.cell(row=numero, column=16, value=str('OPERACION'))
        hoja.cell(row=numero, column=17, value=str('ORDOP'))
        hoja.cell(row=numero, column=18, value=str('RES ACCION'))
        hoja.cell(row=numero, column=19, value=str('RES TIPO'))
        hoja.cell(row=numero, column=20, value=str('RES SUBTIPO'))
        hoja.cell(row=numero, column=21, value=str('RES CLASE'))
        hoja.cell(row=numero, column=22, value=str('CANTIDAD'))
        hoja.cell(row=numero, column=23, value=str('UNIDAD MEDIDA'))
        hoja.cell(row=numero, column=24, value=str('APOYO DAAVA'))
        hoja.cell(row=numero, column=25, value=str('APOYO BLICA'))
        hoja.cell(row=numero, column=26, value=str('APOYO CONAT'))
        hoja.cell(row=numero, column=27, value=str('ESTRATEGIA AFECTA'))

        hoja.cell(row=numero, column=28, value=str('GRADO'))
        hoja.cell(row=numero, column=29, value=str('NOMBRE'))
        hoja.cell(row=numero, column=30, value=str('DOCUMENTO'))
        hoja.cell(row=numero, column=31, value=str('EDAD'))
        hoja.cell(row=numero, column=32, value=str('SEXO'))
        hoja.cell(row=numero, column=33, value=str('ARAMAMENTO INCAUTADO'))
        hoja.cell(row=numero, column=34, value=str('LATITUD'))
        hoja.cell(row=numero, column=35, value=str('LOGITUD'))
        hoja.cell(row=numero, column=36, value=str('RESUMEN'))

        hoja.cell(row=numero, column=37, value=str('SPOA'))
        hoja.cell(row=numero, column=38, value=str('OPE COORDINADA'))
        hoja.cell(row=numero, column=39, value=str('OPE CONJUNTA'))
        hoja.cell(row=numero, column=40, value=str('TIPO OPERACION'))
        hoja.cell(row=numero, column=41, value=str('COMANDANTE'))
        
        hoja.cell(row=numero, column=42, value=str('ESPECIALIDAD'))
        hoja.cell(row=numero, column=43, value=str('NIVEL JERARQUIA'))
        hoja.cell(row=numero, column=44, value=str('ARMA'))
        hoja.cell(row=numero, column=45, value=str('JUDICIALIZACION'))
        hoja.cell(row=numero, column=46, value=str('FECHA JUDICIALIZACION'))
        hoja.cell(row=numero, column=47, value=str('SIGAH'))
        hoja.cell(row=numero, column=48, value=str('TIPO OPERACION'))
        hoja.cell(row=numero, column=49, value=str('HECHO POSITIVO'))
     
        numero = numero +1
        #resultados =np.array(resultados)
        #hechos =np.array(hechos)

        
        for x in resultados:
             
            hoja.cell(row=numero, column=1, value=str(x[26]))
            hoja.cell(row=numero, column=2, value=str(x[35]))
            hoja.cell(row=numero, column=3, value=str(x[28]))
            hoja.cell(row=numero, column=4, value=str(x[0]))
            hoja.cell(row=numero, column=5, value=str(x[36]))
            hoja.cell(row=numero, column=6, value=str(x[1]))
            hoja.cell(row=numero, column=7, value=str(x[2]))
            hoja.cell(row=numero, column=8, value=str(x[3]))
            hoja.cell(row=numero, column=9, value=str(x[4]))
            hoja.cell(row=numero, column=10, value=str(x[5]))
            hoja.cell(row=numero, column=11, value=str(x[6]))
            hoja.cell(row=numero, column=12, value=str(x[37]))
            hoja.cell(row=numero, column=13, value=str(x[38]))
            hoja.cell(row=numero, column=14, value=str(x[7]))
            hoja.cell(row=numero, column=15, value=str(x[39]))
            hoja.cell(row=numero, column=16, value=str(x[8]))
            hoja.cell(row=numero, column=17, value=str(x[33]))

            hoja.cell(row=numero, column=18, value=str(x[9]))
            hoja.cell(row=numero, column=19, value=str(x[10]))
            hoja.cell(row=numero, column=20, value=str(x[11]))
            hoja.cell(row=numero, column=21, value=str(x[12]))
            hoja.cell(row=numero, column=22, value=x[18]).number_format = frmt_d
            hoja.cell(row=numero, column=23, value=str(x[13]))

            hoja.cell(row=numero, column=24, value=str(x[14]))
            hoja.cell(row=numero, column=25, value=str(x[15]))
            hoja.cell(row=numero, column=26, value=str(x[16]))
            hoja.cell(row=numero, column=27, value=str(x[17]))

            hoja.cell(row=numero, column=28, value=str(x[19]))
            hoja.cell(row=numero, column=29, value=str(x[20]))
            hoja.cell(row=numero, column=30, value=str(x[21]))
            hoja.cell(row=numero, column=31, value=str(x[22]))
            hoja.cell(row=numero, column=32, value=str(x[40]))
            hoja.cell(row=numero, column=33, value=str(x[23]))
            hoja.cell(row=numero, column=34, value=str(x[24]))
            hoja.cell(row=numero, column=35, value=str(x[25]))
            hoja.cell(row=numero, column=36, value=str(x[27]))

            
            spoa = "'"+str(x[29])
            hoja.cell(row=numero, column=37, value=spoa)

            hoja.cell(row=numero, column=38, value=str(x[30]))
            hoja.cell(row=numero, column=39, value=str(x[31]))
            hoja.cell(row=numero, column=40, value=str(x[32]))
            hoja.cell(row=numero, column=41, value=str(x[34]))

            hoja.cell(row=numero, column=42, value=str(x[41]))
            hoja.cell(row=numero, column=43, value=str(x[42]))
            hoja.cell(row=numero, column=44, value=str(x[43]))
            hoja.cell(row=numero, column=45, value=str(x[44]))
            hoja.cell(row=numero, column=46, value=str(x[45]))
            hoja.cell(row=numero, column=47, value=str(x[46]))
            hoja.cell(row=numero, column=48, value=str(x[47]))
            hoja.cell(row=numero, column=49, value=str(x[48]))


            numero = numero + 1

        for cell in hoja["1:1"]: 
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill  = PatternFill('solid', start_color="800000")
                # cell.alignment = Alignment(wrap_text=True)
            cell.alignment=Alignment(
                horizontal='center',
                vertical='top',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)

            
        #for cell in hoja["2:2"]: 
            #cell.font = Font(bold=True, size=12, color='FFFFFF')
            #cell.fill  = PatternFill('solid', start_color="5A616B")
            #    # cell.alignment = Alignment(wrap_text=True)
            #cell.alignment=Alignment(
                #horizontal='center',
                #vertical='top',
                #text_rotation=0,
                #wrap_text=False,
                #shrink_to_fit=False,
                #indent=0)
      