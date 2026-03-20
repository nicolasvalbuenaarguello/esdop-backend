import openpyxl
from models.base_datos_posgrest import *
#from funciones.configuracion import *
import sys
def transformar(dato):
    dato = str(dato)
    dato = dato.replace("datetime.time(",  " '")
    dato = dato.replace("?",  ' ')
    dato = dato.replace("#",  ' ')
    # dato = dato.replace("/",  ' ')
    dato = dato.replace("*",  ' ')
    dato = dato.replace(",",  '-')
    dato = dato.replace("datetime.datetime(", " '")
    # dato = dato.replace("(",  ' ')
    dato = dato.replace("),",  "',")
    # dato = dato.replace(")",  ' ')
    dato = dato.replace("$",  ' ')
    dato = dato.replace("%",  ' ')
    dato = dato.replace("{",  ' ')
    dato = dato.replace("}",  ' ')
    dato = dato.replace("[",  ' ')
    dato = dato.replace("]",  ' ')
    dato = dato.replace("<",  ' ')
    dato = dato.replace(">",  ' ')
    dato = dato.replace("¨",  ' ')
    dato = dato.replace("^",  ' ')
    dato = dato.replace("~",  ' ')
    dato = dato.replace("`",  ' ')
    dato = dato.replace("'",  '')
    dato = dato.replace('"',  ' ')
    
    
    return dato
    
def reemplazar(dato_1,dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32, dato_33, dato_34, dato_35, dato_36, dato_37, dato_38, dato_39, dato_40, dato_41, dato_42, dato_43, dato_44, dato_45, dato_46, dato_47, dato_48	, dato_49	, dato_50	, dato_51	, dato_52	, dato_53	, dato_54	, dato_55	, dato_56	, dato_57	, dato_58	, dato_59	, dato_60	, dato_61	, dato_62	, dato_63	, dato_64	, dato_65	, dato_66	, dato_67	, dato_68	, dato_69	, dato_70	, dato_71	, dato_72	, dato_73	, dato_74	, dato_75	, dato_76	, dato_77	, dato_78	, dato_79	, dato_80	, dato_81	, dato_82	, dato_83	, dato_84	, dato_85	, dato_86	, dato_87	, dato_88	, dato_89	, dato_90	, dato_91	, dato_92	, dato_93	, dato_94	, dato_95	, dato_96	, dato_97	, dato_98	, dato_99	, dato_100	, dato_101	, dato_102	, dato_103	):
    
    dato_1=transformar(str( dato_1.value))
    dato_2=transformar(str( dato_2.value))
    dato_3=transformar(str( dato_3.value))
    dato_4=transformar(str( dato_4.value))
    dato_5=transformar(str( dato_5.value))
    dato_6=transformar(str( dato_6.value))
    dato_7=transformar(str( dato_7.value))
    dato_8=transformar(str( dato_8.value))
    dato_9=transformar(str( dato_9.value))
    dato_10=transformar(str(dato_10.value))
    dato_11=transformar(str(dato_11.value))
    dato_12=transformar(str(dato_12.value))
    dato_13=transformar(str(dato_13.value))
    dato_14=transformar(str(dato_14.value))
    dato_15=transformar(str(dato_15.value))
    dato_16=transformar(str(dato_16.value))
    dato_17=transformar(str(dato_17.value))
    dato_18=transformar(str(dato_18.value))
    dato_19=transformar(str(dato_19.value))
    dato_20=transformar(str(dato_20.value))
    dato_21=transformar(str(dato_21.value))
    dato_22=transformar(str(dato_22.value))
    dato_23=transformar(str(dato_23.value))
    dato_24=transformar(str(dato_24.value))
    dato_25=transformar(str(dato_25.value))
    dato_26=transformar(str(dato_26.value))
    dato_27=transformar(str(dato_27.value))
    dato_28=transformar(str(dato_28.value))
    dato_29=transformar(str(dato_29.value))
    dato_30=transformar(str(dato_30.value))
    dato_31=transformar(str(dato_31.value))
    dato_32=transformar(str(dato_32.value))
    dato_33=transformar(str(dato_33.value))
    dato_34=transformar(str(dato_34.value))
    dato_35=transformar(str(dato_35.value))
    dato_36=transformar(str(dato_36.value))
    dato_37=transformar(str(dato_37.value))
    dato_38=transformar(str(dato_38.value))
    dato_39=transformar(str(dato_39.value))
    dato_40=transformar(str(dato_40.value))
    dato_41=transformar(str(dato_41.value))
    dato_42=transformar(str(dato_42.value))
    dato_43=transformar(str(dato_43.value))
    dato_44=transformar(str(dato_44.value))
    dato_45=transformar(str(dato_45.value))
    dato_46=transformar(str(dato_46.value))
    dato_47=transformar(str(dato_47.value))
    dato_48=transformar(str(dato_48.value))
    dato_49=transformar(str(dato_49.value))
    dato_50=transformar(str(dato_50.value))
    dato_51=transformar(str(dato_51.value))
    dato_52=transformar(str(dato_52.value))
    dato_53=transformar(str(dato_53.value))
    dato_54=transformar(str(dato_54.value))
    dato_55=transformar(str(dato_55.value))
    dato_56=transformar(str(dato_56.value))
    dato_57=transformar(str(dato_57.value))
    dato_58=transformar(str(dato_58.value))
    dato_59=transformar(str(dato_59.value))
    dato_60=transformar(str(dato_60.value))
    dato_61=transformar(str(dato_61.value))
    dato_62=transformar(str(dato_62.value))
    dato_63=transformar(str(dato_63.value))
    dato_64=transformar(str(dato_64.value))
    dato_65=transformar(str(dato_65.value))
    dato_66=transformar(str(dato_66.value))
    dato_67=transformar(str(dato_67.value))
    dato_68=transformar(str(dato_68.value))
    dato_69=transformar(str(dato_69.value))
    dato_70=transformar(str(dato_70.value))
    dato_71=transformar(str(dato_71.value))
    dato_72=transformar(str(dato_72.value))
    dato_73=transformar(str(dato_73.value))
    dato_74=transformar(str(dato_74.value))
    dato_75=transformar(str(dato_75.value))
    dato_76=transformar(str(dato_76.value))
    dato_77=transformar(str(dato_77.value))
    dato_78=transformar(str(dato_78.value))
    dato_79=transformar(str(dato_79.value))
    dato_80=transformar(str(dato_80.value))
    dato_81=transformar(str(dato_81.value))
    dato_82=transformar(str(dato_82.value))
    dato_83=transformar(str(dato_83.value))
    dato_84=transformar(str(dato_84.value))
    dato_85=transformar(str(dato_85.value))
    dato_86=transformar(str(dato_86.value))
    dato_87=transformar(str(dato_87.value))
    dato_88=transformar(str(dato_88.value))
    dato_89=transformar(str(dato_89.value))
    dato_90=transformar(str(dato_90.value))
    dato_91=transformar(str(dato_91.value))
    dato_92=transformar(str(dato_92.value))
    dato_93=transformar(str(dato_93.value))
    dato_94=transformar(str(dato_94.value))
    dato_95=transformar(str(dato_95.value))
    dato_96=transformar(str(dato_96.value))
    dato_97=transformar(str(dato_97.value))
    dato_98=transformar(str(dato_98.value))
    dato_99=transformar(str(dato_99.value))
    dato_100=transformar(str(dato_100.value))
    dato_101=transformar(str(dato_101.value))
    dato_102=transformar(str(dato_102.value))
    dato_103=transformar(str(dato_103.value))

    dato = [dato_1,dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32, dato_33, dato_34, dato_35, dato_36, dato_37, dato_38, dato_39, dato_40, dato_41, dato_42, dato_43, dato_44, dato_45, dato_46, dato_47, dato_48	, dato_49	, dato_50	, dato_51	, dato_52	, dato_53	, dato_54	, dato_55	, dato_56	, dato_57	, dato_58	, dato_59	, dato_60	, dato_61	, dato_62	, dato_63	, dato_64	, dato_65	, dato_66	, dato_67	, dato_68	, dato_69	, dato_70	, dato_71	, dato_72	, dato_73	, dato_74	, dato_75	, dato_76	, dato_77	, dato_78	, dato_79	, dato_80	, dato_81	, dato_82	, dato_83	, dato_84	, dato_85	, dato_86	, dato_87	, dato_88	, dato_89	, dato_90	, dato_91	, dato_92	, dato_93	, dato_94	, dato_95	, dato_96	, dato_97	, dato_98	, dato_99	, dato_100	, dato_101	, dato_102	, dato_103	]

    return dato
  

#coding='utf8'
class conversion ():
    def __init__(self):
        pass
           
    def leer_excel(fichero):
        global hojas
        global wb
        wb = openpyxl.load_workbook(fichero)
        hojas = wb.get_sheet_names()
        
                
            
    def convertir(libro, hoja, resultado):
        global a_1
        global dato_1
        global dato_2
        global fila
        
        print("dos")
        



        # wb = openpyxl.load_workbook(libro)
        sheet = wb[hoja] # Obtenga el libro de trabajo actualmente activo
        fila = sheet.max_row # Número máximo de líneas
        fila = fila +1
        columna = sheet.max_column # Número máximo de columnas

        
        
        
        a_1 = ""
        dato_1 = []
        dato_2 = ""
        
        # if resultado == "Hoja1":
            # ri =2
        a_1 =()
        for ri in range(2, fila ):
               
                a_1 = reemplazar(sheet.cell(row = ri, column = 1),sheet.cell(row = ri, column = 2),sheet.cell(row = ri, column = 3),sheet.cell(row = ri, column = 4),sheet.cell(row = ri, column = 5),sheet.cell(row = ri, column = 6),sheet.cell(row = ri, column = 7),sheet.cell(row = ri, column = 8),sheet.cell(row = ri, column = 9),sheet.cell(row = ri, column = 10),sheet.cell(row = ri, column = 11),sheet.cell(row = ri, column = 12),sheet.cell(row = ri, column = 13),sheet.cell(row = ri, column = 14),sheet.cell(row = ri, column = 15),sheet.cell(row = ri, column = 16),sheet.cell(row = ri, column = 17),sheet.cell(row = ri, column = 18),sheet.cell(row = ri, column = 19),sheet.cell(row = ri, column = 20), sheet.cell(row = ri, column = 21),sheet.cell(row = ri, column = 22),sheet.cell(row = ri, column = 23),sheet.cell(row = ri, column = 24),sheet.cell(row = ri, column = 25),sheet.cell(row = ri, column = 26),sheet.cell(row = ri, column = 27),sheet.cell(row = ri, column = 28),sheet.cell(row = ri, column = 29),sheet.cell(row = ri, column = 30), sheet.cell(row = ri, column = 31),sheet.cell(row = ri, column = 32),sheet.cell(row = ri, column = 33),sheet.cell(row = ri, column = 34),sheet.cell(row = ri, column = 35),sheet.cell(row = ri, column = 36),sheet.cell(row = ri, column = 37),sheet.cell(row = ri, column = 38),sheet.cell(row = ri, column = 39),sheet.cell(row = ri, column = 40),sheet.cell(row = ri, column = 41),sheet.cell(row = ri, column = 42),sheet.cell(row = ri, column = 43),sheet.cell(row = ri, column = 44),sheet.cell(row = ri, column = 45),sheet.cell(row = ri, column = 46),sheet.cell(row = ri, column = 47),sheet.cell(row = ri, column = 48)	,sheet.cell(row = ri, column = 49)	,sheet.cell(row = ri, column = 50)	,sheet.cell(row = ri, column = 51)	,sheet.cell(row = ri, column = 52)	,sheet.cell(row = ri, column = 53)	,sheet.cell(row = ri, column = 54)	,sheet.cell(row = ri, column = 55)	,sheet.cell(row = ri, column = 56)	,sheet.cell(row = ri, column = 57)	,sheet.cell(row = ri, column = 58)	,sheet.cell(row = ri, column = 59)	,sheet.cell(row = ri, column = 60)	,sheet.cell(row = ri, column = 61)	,sheet.cell(row = ri, column = 62)	,sheet.cell(row = ri, column = 63)	,sheet.cell(row = ri, column = 64)	,sheet.cell(row = ri, column = 65)	,sheet.cell(row = ri, column = 66)	,sheet.cell(row = ri, column = 67)	,sheet.cell(row = ri, column = 68)	,sheet.cell(row = ri, column = 69)	,sheet.cell(row = ri, column = 70)	,sheet.cell(row = ri, column = 71)	,sheet.cell(row = ri, column = 72)	,sheet.cell(row = ri, column = 73)	,sheet.cell(row = ri, column = 74)	,sheet.cell(row = ri, column = 75)	,sheet.cell(row = ri, column = 76)	,sheet.cell(row = ri, column = 77)	,sheet.cell(row = ri, column = 78)	,sheet.cell(row = ri, column = 79)	,sheet.cell(row = ri, column = 80)	,sheet.cell(row = ri, column = 81)	,sheet.cell(row = ri, column = 82)	,sheet.cell(row = ri, column = 83)	,sheet.cell(row = ri, column = 84)	,sheet.cell(row = ri, column = 85)	,sheet.cell(row = ri, column = 86)	,sheet.cell(row = ri, column = 87)	,sheet.cell(row = ri, column = 88)	,sheet.cell(row = ri, column = 89)	,sheet.cell(row = ri, column = 90)	,sheet.cell(row = ri, column = 91)	,sheet.cell(row = ri, column = 92)	,sheet.cell(row = ri, column = 93)	,sheet.cell(row = ri, column = 94)	,sheet.cell(row = ri, column = 95)	,sheet.cell(row = ri, column = 96)	,sheet.cell(row = ri, column = 97)	,sheet.cell(row = ri, column = 98)	,sheet.cell(row = ri, column = 99)	,sheet.cell(row = ri, column = 100)	,sheet.cell(row = ri, column = 101)	,sheet.cell(row = ri, column = 102)	,sheet.cell(row = ri, column = 103))  


                dato_1.append(a_1)
 
                      
                # print( str(ri) + " completado de " + str(fila) , end='\r')
                # sys.stdout.flush()
                
        dato_1 = str(dato_1)
        dato_1 = dato_1.replace(']',  '),')
        dato_1 = dato_1.replace('[',  '(')
        dato_1 = dato_1.replace('("',  '(')
        dato_1 = dato_1.replace('")',  ')')
        dato_1 = dato_1.replace(',,',  ', ') 
        dato_1 = dato_1.replace('", "',  '), (') 

        dato_1 =dato_1[:-3]
        dato_1 =dato_1[1:]


 
        save_data(dato_1, resultado)

            
            # print(dato_2)
 


 
 
 