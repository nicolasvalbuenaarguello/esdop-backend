from datetime import datetime
import json
# coding: utf-8
import psycopg2
import base64 

from os import getcwd
import os
import shutil
from dotenv import load_dotenv
load_dotenv()

from z_z_configuarcion.caligrafia import *
from z_z_configuarcion.fechas import *

from flask import make_response
from datetime import date, time, datetime

from fpdf import FPDF
from math import sqrt, pi, sin, cos
from fpdf.php import sprintf
#Funcion para la creacion del reporte de resultados 
from __init__ import *

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class PDF(FPDF):
    
    def __init__(self, orientation = 'P', unit = 'mm', format = 'A4'):
        self.ext_gstates = [ ]
        
        super(PDF, self).__init__(orientation, unit, format)

    def parametros(self, **kwargs):
        self.titulo = ""
        self.tamanio = "oficio"
        self.logo = ""
        self.fecha_titulo = ""
        self.permiso = "EJC"
        self.nivel = ""
        self.usuario = ""
        self.pie_pagina = "NO"
        self.fondo_pagina = "SI"
        self.seguridad = "JEMOP"
        self.precidencial =  "OK"
        self.ruta = ""
        self.imagen = ""
        self.img_qr = ""
        self.total_pel=""

        for k in kwargs.items():
            if "titulo" == k[0]:
                self.titulo = k[1]
            if "tamanio" == k[0]:
                self.tamanio = k[1]
            if "logo" == k[0]:
                self.logo = k[1]
            if "fecha_titulo" == k[0]:
                self.fecha_titulo = k[1]
            if "permiso" == k[0]:
                self.permiso = k[1]
            if "nivel" == k[0]:
                self.nivel = k[1]
            if "usuario" == k[0]:
                self.usuario = k[1]
            if "pie_pagina" == k[0]:
                self.pie_pagina = k[1]
                
            if "fondo_pagina" == k[0]:
                self.fondo_pagina = k[1]
            if "seguridad" == k[0]:
                self.seguridad = k[1]
            if "precidencial" == k[0]:
                self.precidencial = k[1]
            if "ruta" == k[0]:
                self.ruta = k[1]
            if "imagen" == k[0]:
                self.imagen = k[1]
            if "img_qr" == k[0]:
                self.img_qr = k[1]
            if "total_pel" == k[0]:
                self.total_pel = k[1]
        
    def header(self):
        #imagen
        
        if self.fondo_pagina == "SI":
            self.image(self.imagen,0,0,279,216)
            self.set_font('Arial Black', '', 16)
            self.set_text_color(56,87,35)
            self.text(50, 20, self.titulo)
            # fecha = strftime(self.fecha, '%y')
            self.image(self.img_qr+"/QR.png",253,3,23,23)

            self.set_text_color(90,90,90)
            self.set_font('Arial Narrow', 'B', 30)
            self.text(10,50,str(self.total_pel[0]))
            self.set_font('BebasNeue', '', 20)
            self.set_text_color(56,87,35)
            self.text(10,56,str("TOTAL"))
            self.text(10,62,str("PELOTONES"))

            self.set_draw_color(140,140,140)
            self.set_line_width(0.7)
            self.line(10, 63, 35, 63)
            self.set_font('BebasNeue', '', 18)
            self.set_text_color(80,80,80)

            self.text(10,68,"DISPOSITIVO")
            self.text(10,74, str(self.total_pel[1]))
            self.text(10,80, str(self.total_pel[2]))
            self.text(10,86, str(self.total_pel[3]))

        
        self.ln(25)

    def footer(self):
        
        if self.pie_pagina =="SI":
            self.set_text_color(80,80,80)
            # Go to 1.5 cm from bottom
            self.set_y(-15)
            # Select Arial italic 8
            self.set_font('Arial', 'B', 10)
            # Print centered page number
            # self.cell(0, 10, 'Page %s' % self.page_no(), 0, 0, 'C')

            dias = datetime.now().day
            meses = datetime.now().month
            anios = datetime.now().year
            # print(datetime.now().month)
            meses = mes(meses)
            meses =str(meses)
            meses = meses.upper()
            # print(meses)   
            fecha_actual = str(dias) +" "+ str(meses) +" "+ str(anios)

            if self.precidencial == "OK":
                self.text(55, 212, "SECRETO          " + str(fecha_actual))
                self.text(15, 212, 'Fuente: SICOE')
            else:
                self.text(95, 212, "SECRETO ")
                self.text(15, 212, 'Fuente: SICOE')


            if self.seguridad == "JEMOP": 
                self.text(140, 212, '"Se deben verificar los datos con CEDE3 - DISEO"')
            

            if self.permiso != "EJC":
                self.set_text_color(128,0,0)
                self.text(230, 212, "Usuario: '"+str(self.usuario)+" unidad (" + str(self.nivel)+")'")



    def rounded_rect(self, x, y, w, h, r, style = '', corners = '1234'):
    
        k = self.k
        hp = self.h
        if(style=='F'):
            op='f'
        elif(style=='FD' or style=='DF'):
            op='B'
        else:
            op='S'
        myArc = 4/3 * (sqrt(2) - 1)
        self._out('%.2F %.2F m' % ((x+r)*k,(hp-y)*k))

        xc = x+w-r
        yc = y+r
        self._out('%.2F %.2F l' % (xc*k,(hp-y)*k))
        if '2' not in corners:
            self._out('%.2F %.2F l' % ((x+w)*k,(hp-y)*k))
        else:
            self._arc(xc + r*myArc, yc - r, xc + r, yc - r*myArc, xc + r, yc)

        xc = x+w-r
        yc = y+h-r
        self._out('%.2F %.2F l' % ((x+w)*k,(hp-yc)*k))
        if '3' not in corners:
            self._out('%.2F %.2F l' % ((x+w)*k,(hp-(y+h))*k))
        else:
            self._arc(xc + r, yc + r*myArc, xc + r*myArc, yc + r, xc, yc + r)

        xc = x+r
        yc = y+h-r
        self._out('%.2F %.2F l' % (xc*k,(hp-(y+h))*k))
        if '4' not in corners:
            self._out('%.2F %.2F l' % (x*k,(hp-(y+h))*k))
        else:
            self._arc(xc - r*myArc, yc + r, xc - r, yc + r*myArc, xc - r, yc)

        xc = x+r 
        yc = y+r
        self._out('%.2F %.2F l' % (x*k,(hp-yc)*k))
        if '1' not in corners:
            self._out('%.2F %.2F l' % (x*k,(hp-y)*k))
            self._out('%.2F %.2F l' % ((x+r)*k,(hp-y)*k))
        else:
            self._arc(xc - r, yc - r*myArc, xc - r*myArc, yc - r, xc, yc - r)
        self._out(op)
    
    def _arc(self, x1, y1, x2, y2, x3, y3):
    
        h = self.h
        self._out('%.2F %.2F %.2F %.2F %.2F %.2F c ' % (x1*self.k, (h-y1)*self.k,
            x2*self.k, (h-y2)*self.k, x3*self.k, (h-y3)*self.k))
        
    def sector(self, xc, yc, r, a, b, style='FD', cw=True, o=90):
    
        d0 = a - b
        if cw:
            d = b
            b = o - a
            a = o - d
        else:
            b += o
            a += o
        
        while a<0:
            a += 360
        while a>360:
            a -= 360
        while b<0:
            b += 360
        while b>360:
            b -= 360
        if a > b:
            b += 360
        b = b/360*2*pi
        a = a/360*2*pi
        d = b - a
        if d == 0 and d0 != 0:
            d = 2*pi
        k = self.k
        hp = self.h
        if sin(d/2):
            myArc = 4/3*(1-cos(d/2))/sin(d/2)*r
        else:
            myArc = 0
        #first put the center
        self._out('%.2F %.2F m' % ((xc)*k,(hp-yc)*k))
        #put the first point
        self._out('%.2F %.2F l' % ((xc+r*cos(a))*k,((hp-(yc-r*sin(a)))*k)))
        #draw the arc
        if d < pi/2:
            self.sector_arc(xc+r*cos(a)+myArc*cos(pi/2+a),
                        yc-r*sin(a)-myArc*sin(pi/2+a),
                        xc+r*cos(b)+myArc*cos(b-pi/2),
                        yc-r*sin(b)-myArc*sin(b-pi/2),
                        xc+r*cos(b),
                        yc-r*sin(b)
                        )
        else:
            b = a + d/4
            myArc = 4/3*(1-cos(d/8))/sin(d/8)*r
            self.sector_arc(xc+r*cos(a)+myArc*cos(pi/2+a),
                        yc-r*sin(a)-myArc*sin(pi/2+a),
                        xc+r*cos(b)+myArc*cos(b-pi/2),
                        yc-r*sin(b)-myArc*sin(b-pi/2),
                        xc+r*cos(b),
                        yc-r*sin(b)
                        )
            a = b
            b = a + d/4
            self.sector_arc(xc+r*cos(a)+myArc*cos(pi/2+a),
                        yc-r*sin(a)-myArc*sin(pi/2+a),
                        xc+r*cos(b)+myArc*cos(b-pi/2),
                        yc-r*sin(b)-myArc*sin(b-pi/2),
                        xc+r*cos(b),
                        yc-r*sin(b)
                        )
           
            a = b
            b = a + d/4
            self.sector_arc(xc+r*cos(a)+myArc*cos(pi/2+a),
                        yc-r*sin(a)-myArc*sin(pi/2+a),
                        xc+r*cos(b)+myArc*cos(b-pi/2),
                        yc-r*sin(b)-myArc*sin(b-pi/2),
                        xc+r*cos(b),
                        yc-r*sin(b)
                        )
            a = b
            b = a + d/4
            self.sector_arc(xc+r*cos(a)+myArc*cos(pi/2+a),
                        yc-r*sin(a)-myArc*sin(pi/2+a),
                        xc+r*cos(b)+myArc*cos(b-pi/2),
                        yc-r*sin(b)-myArc*sin(b-pi/2),
                        xc+r*cos(b),
                        yc-r*sin(b)
                        )
        
        #terminate drawing
        if style=='F':
            op='f'
        elif style=='FD' or style=='DF':
            op='b'
        else:
            op='s'
        self._out(op)
    
    def sector_arc(self, x1, y1, x2, y2, x3, y3 ):
    
        h = self.h
        self._out('%.2F %.2F %.2F %.2F %.2F %.2F c' %
            (x1*self.k,
            (h-y1)*self.k,
            x2*self.k,
            (h-y2)*self.k,
            x3*self.k,
            (h-y3)*self.k))
 
    def set_alpha(self, alpha, bm='Normal'):
        state = {
            'ca': alpha,
            'CA': alpha,
            'BM': '/' + bm
        }
        self.ext_gstates.append(state)
        self._set_ext_gstate(len(self.ext_gstates))

    def _set_ext_gstate(self, gstate_index):
        self._out(sprintf('/GS%d gs', gstate_index))

    def _enddoc(self):
        if len(self.ext_gstates) > 0 and self.pdf_version < '1.4':
            self.pdf_version = '1.4'
        super(PDF, self)._enddoc()

    def _putextgstates(self):
        for gstate in self.ext_gstates:
            self._newobj()
            gstate['n'] = self.n
            self._out('<</Type /ExtGState')
            self._out(sprintf('/ca %.3F', gstate['ca']))
            self._out(sprintf('/CA %.3F', gstate['CA']))
            self._out('/BM ' + gstate['BM'])
            self._out('>>')
            self._out('endobj')

    def _putresourcedict(self):
        super(PDF, self)._putresourcedict()
        self._out('/ExtGState <<')
        for index, gstate in enumerate(self.ext_gstates):
            self._out('/GS' + str(index+1) + ' ' + str(gstate['n']) +' 0 R')
        self._out('>>')

    def _putresources(self):
        self._putextgstates()
        super(PDF, self)._putresources()



def llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, unidad ):
        CANTIDAD_2.append(int(total_pel[4]))
        CANTIDAD_2.append(int(total_pel[5]))
        CANTIDAD_2.append(int(total_pel[6]))
        CANTIDAD_2.append(int(total_pel[7]))

        columns.append(unidad)
        columns.append(unidad)
        columns.append(unidad)
        columns.append(unidad)

        ANIO.append("PEL.")
        ANIO.append("OPE.")
        ANIO.append("ENT.")
        ANIO.append("DES.")
        return[CANTIDAD_2, columns, ANIO]

        #unsad unidad de serviciona adminitrativos
        #segyar 
def validacion_cantidad_tupla(tupla):
                  
    if len(tupla) == 1:
        valor = " = '"+tupla[0]+"'"
        
    else:
        tupla = tuple(tupla)
        valor = " in {}" .format(tupla)
    return valor 

def sumar_personal(x):
    return [int(x[i]) for i in range(16, 21)]
EXCLUIR_ESTADOS = {"NOV", "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM"  }
UNIDADES = {
        "DIV01": ["DIV01"],
        "DIV02": ["DIV02", "CENOR", "FUDRA3"],
        "DIV03": ["DIV03", "FTCEC"],
        "DIV04": ["DIV04"],
        "DIV05": ["DIV05"],
        "DIV06": ["DIV06"],
        "DIV07": ["DIV07", "FTCTI"],
        "DIV08": ["DIV08", "FURON"],
        "FUTOM": ["FUTCO", "FUDRA1", "FUTOM"],
        "DAVAA": ["DAVAA"],
        "TREJC": ["COLOG", "CAFUE", "COING", "JEMOP", "JEMGF", "SECEJ", "TEJC", "TREJC", "CEDOC", "CAIMI"],
        "CAAID": ["CAAID"],
        "CAOCC": ["CAOCC"],
        "DIVFE": ["DIVFE"],
        "CONAT": ["CONAT"]
    }
def estado_valido(estado):
    return estado not in EXCLUIR_ESTADOS

def calcular_CODE_2(pdf, dato, division, pos_x, pos_y):

    def sumar_lista(l1, l2):
        return [a + b for a, b in zip(l1, l2)]

    # Diccionario por tipo
    categorias = {
        "Descanso": [0, 0, 0, 0, 0],
        "Entrenamiento": [0, 0, 0, 0, 0],
        "Novedades": [0, 0, 0, 0, 0],
        "Operaciones": [0, 0, 0, 0, 0],
        "Emb": [0, 0, 0, 0, 0],
    }

    for x in dato:
        estado = x[12]
        actividad = x[25]
        valores = sumar_personal(x)

        if actividad == "Descanso" and estado_valido(estado):
            categorias["Descanso"] = sumar_lista(categorias["Descanso"], valores)

        elif actividad == "Entrenamiento" and estado_valido(estado):
            categorias["Entrenamiento"] = sumar_lista(categorias["Entrenamiento"], valores)

        elif actividad == "Operaciones" and estado_valido(estado):
            categorias["Operaciones"] = sumar_lista(categorias["Operaciones"], valores)

        if actividad == "Novedades" or estado == "NOV":
            categorias["Novedades"] = sumar_lista(categorias["Novedades"], valores)

        if estado in { "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM", "NOV"}:
            categorias["Emb"] = sumar_lista(categorias["Emb"], valores)

    # Totales
    def total_categoria(cat):
        return sum(categorias[cat])

    total_personal = sum(total_categoria(c) for c in categorias)

    # Formatear para mostrar
    def fmt(n):
        return f"{n:,}".replace(",", ".")

    pdf.set_text_color(125, 0, 0)
    pdf.set_font('BebasNeue', '', 12)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 10)
    pos_y += 4
    pdf.text(pos_x, pos_y, "TOTAL: " + fmt(total_personal))
    pos_y += 4
    pdf.text(pos_x, pos_y, "OPE: " + fmt(total_categoria("Operaciones")))
    pos_y += 4
    pdf.text(pos_x, pos_y, "DES: " + fmt(total_categoria("Descanso")))
    pos_y += 4
    pdf.text(pos_x, pos_y, "ENT: " + fmt(total_categoria("Entrenamiento")))
    pos_y += 4
    pdf.text(pos_x, pos_y, "EM: " + fmt(total_categoria("Emb")))
    pos_y += 4
    #pdf.text(pos_x, pos_y, "NOV: " + fmt(total_categoria("Novedades")))



    #print(str(len(str(total)))+ " - " + str(len(str(division))))


def calcular_CODE(pdf, dato, division, pos_x, pos_y, div_escudo):
    from os.path import exists

    # Mostrar escudo
    imagen = f"C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/{div_escudo}.png"
    if exists(imagen):
        pdf.image(imagen, pos_x - 6, pos_y - 5, 6, 6)



    def sumar_personal(x):
        return [int(x[i]) for i in range(16, 21)]

    def sumar_lista(l1, l2):
        return [a + b for a, b in zip(l1, l2)]

    categorias = {
        "Descanso": [0, 0, 0, 0, 0],
        "Entrenamiento": [0, 0, 0, 0, 0],
        "Operaciones": [0, 0, 0, 0, 0],
        "Emb": [0, 0, 0, 0, 0],
    }

    total_general = [0, 0, 0, 0, 0]  # OFI, SUB, SLP, SL18, SL12

    for x in dato:
        estado = x[12]
        actividad = x[25]
        valores = sumar_personal(x)

        total_general = sumar_lista(total_general, valores)

        if actividad == "Descanso" and estado not in EXCLUIR_ESTADOS:
            categorias["Descanso"] = sumar_lista(categorias["Descanso"], valores)

        elif actividad == "Entrenamiento" and estado not in EXCLUIR_ESTADOS:
            categorias["Entrenamiento"] = sumar_lista(categorias["Entrenamiento"], valores)

        elif actividad == "Operaciones" and estado not in EXCLUIR_ESTADOS:
            categorias["Operaciones"] = sumar_lista(categorias["Operaciones"], valores)

        if estado in { "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM", "NOV"}:
            categorias["Emb"] = sumar_lista(categorias["Emb"], valores)

    # Calcular totales
    def total_categoria(cat):
        return sum(categorias[cat])

    total_personal = sum(total_general)
    total_descanso = total_categoria("Descanso")
    total_entrenamiento = total_categoria("Entrenamiento")
    total_operaciones = total_categoria("Operaciones")
    total_emb = total_categoria("Emb")

    total_novedades = total_personal - (
        total_descanso + total_entrenamiento + total_operaciones + total_emb
    )

    # Formato
    def fmt(n):
        return f"{n:,}".replace(",", ".")

    pdf.set_text_color(125, 0, 0)
    pdf.set_font("BebasNeue", "", 14)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font("BebasNeue", "", 12)

    pos_y += 4
    pdf.text(pos_x, pos_y, "TOTAL: " + fmt(total_personal))
    pos_y += 4
    pdf.text(pos_x, pos_y, "OPE: " + fmt(total_operaciones))
    pos_y += 4
    pdf.text(pos_x, pos_y, "DES: " + fmt(total_descanso))
    pos_y += 4
    pdf.text(pos_x, pos_y, "ENT: " + fmt(total_entrenamiento))
    pos_y += 4
    pdf.text(pos_x, pos_y, "EM: " + fmt(total_emb))
    pos_y += 4
    #pdf.text(pos_x, pos_y, "NOV: " + fmt(total_novedades))


def cuadros_CODE(pdf, CUNDINAMARCA, lugar, pos_x, pos_y, r,g,b):

    pos_x_x = pos_x + 2
    pos_y_y = pos_y + 1

    pdf.set_text_color(10,10,10)
    pdf.set_font('BebasNeue', '', 10)

    pdf.set_fill_color(r, g, b)
    pdf.set_draw_color(r,g,b)
    pdf.rounded_rect(pos_x, pos_y, 20, 27, 1,'DF', '1234')



    calcular_CODE_2(pdf, CUNDINAMARCA, lugar,  pos_x_x, pos_y_y)


def creaacon(pdf, dato, division, pos_x, pos_y, div_escudo):
    # Cargar escudo
    imagen = f"C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/{div_escudo}.png"
    pdf.image(imagen, pos_x - 6, pos_y - 5, 6, 6)

    # Calcular totales
    totales = calcular_totales_insitop(dato)

    # Formatear textos
    total_txt = formato_colombiano(totales["total_general"])
    resultados_cuadros = f"{formato_colombiano(totales['ofi'])} - {formato_colombiano(totales['sub'])}"
    

    resultados_cuadros_slp = formato_colombiano(totales["slp"])
    if totales["sl18"] > 0 or totales["sl12"] > 0:
        resultados_cuadros_slp += f" - {formato_colombiano(totales['sl18'])} - {formato_colombiano(totales['sl12'])}"

    # Escribir en PDF
    pdf.set_text_color(125, 0, 0)
    pdf.set_font('BebasNeue', '', 14)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)

    pos_y += 4
    pdf.text(pos_x, pos_y, total_txt)
    pos_y += 4
    pdf.text(pos_x, pos_y, resultados_cuadros)
    pos_y += 4

    if totales["total_sol"] > 0:
        pdf.text(pos_x, pos_y, resultados_cuadros_slp)

def pelotones_total(dato, dato_2):


    def es_estado_valido(estado):
        return estado not in EXCLUIR_ESTADOS

    total_pelotones = sum(1 for x in dato if es_estado_valido(x[3]))

    operaciones = 0
    entrenamiento = 0

    for x in dato_2:
        if not es_estado_valido(x[3]):
            continue
        actividad = x[5]
        if actividad == "Operaciones":
            operaciones += 1
        elif actividad == "Entrenamiento":
            entrenamiento += 1

    descanso = total_pelotones - (operaciones + entrenamiento)

    def formatear(numero):
        return f"{numero:,}".replace(",", ".")

    return [
        formatear(total_pelotones),
        formatear(operaciones),
        formatear(entrenamiento),
        formatear(descanso)
    ]


def calcular_totales_insitop(dato):
    ofi = sum(int(x[16]) for x in dato)
    sub = sum(int(x[17]) for x in dato)
    slp = sum(int(x[18]) for x in dato)
    sl18 = sum(int(x[19]) for x in dato)
    sl12 = sum(int(x[20]) for x in dato)

    total_sol = slp + sl18 + sl12
    total_general = ofi + sub + total_sol

    return {
        "ofi": ofi,
        "sub": sub,
        "slp": slp,
        "sl18": sl18,
        "sl12": sl12,
        "total_sol": total_sol,
        "total_general": total_general
    }

def formato_colombiano(numero):
    return format(numero, ',').replace(',', '.')

def pelotones_TOTAL_2(dato, dato_2):

    peloton = sum(1 for x in dato if x[3] not in EXCLUIR_ESTADOS)

    operaciones = sum(1 for x in dato_2 if x[3] not in EXCLUIR_ESTADOS and x[5] == "Operaciones")
    entremaniento = sum(1 for x in dato_2 if x[3] not in EXCLUIR_ESTADOS and x[5] == "Entrenamiento")
    descanso = peloton - operaciones - entremaniento

    def formatear(valor, etiqueta):
        return f"{etiqueta} {format(valor, ',').replace(',', '.')}"

    return [
        format(peloton, ',').replace(',', '.'),
        formatear(operaciones, "OPE:"),
        formatear(entremaniento, "ENTR:"),
        formatear(descanso, "DESC:"),
        peloton, operaciones, entremaniento, descanso
    ]


def pelotones_TOTAL_2_gaula(dato, dato_2):
    peloton = 0
    operaciones = 0
    descanso = 0
    entrenamiento = 0

    # Filtrado por estados no válidos


    # Conteo de pelotones válidos
    for x in dato:
        if x[3] not in EXCLUIR_ESTADOS:
            peloton += 1

    # Conteo por actividad en pelotones válidos
    for x in dato_2:
        if x[3] not in EXCLUIR_ESTADOS:
            actividad = x[5]
            if actividad == "Descanso":
                descanso += 1
            elif actividad == "Entrenamiento":
                entrenamiento += 1
            elif actividad == "Operaciones":
                operaciones += 1

    # Ajuste de descanso si hay desfase
    descanso = peloton - operaciones - entrenamiento

    # Formateo de resultados
    peloton_a = f"{peloton:,}".replace(",", ".")
    operaciones_a = f"OPE: {operaciones:,}".replace(",", ".")
    entrenamiento_a = f"ENTR: {entrenamiento:,}".replace(",", ".")
    descanso_a = f"DESC: {descanso:,}".replace(",", ".")

    return [
        peloton_a,
        operaciones_a,
        entrenamiento_a,
        descanso_a,
        peloton,
        operaciones,
        entrenamiento,
        descanso,
    ]

def pelotones_2(pdf, dato, dato_2, division, pos_x, pos_y):
    # Inicializar contadores
    total_pelotones = 0
    operaciones = 0
    descanso = 0
    entrenamiento = 0

    # Filtrar pelotones válidos en dato

    for x in dato:
        if x[3] not in EXCLUIR_ESTADOS:
            total_pelotones += 1

    # Clasificar actividades en dato_2
    for x in dato_2:
        if x[3] not in EXCLUIR_ESTADOS:
            actividad = x[5]
            if actividad == "Descanso":
                descanso += 1
            elif actividad == "Entrenamiento":
                entrenamiento += 1
            elif actividad == "Operaciones":
                operaciones += 1

    # Recalcular descanso restante
    descanso = total_pelotones - operaciones - entrenamiento

    # Formatear textos
    texto_peloton = f"PEL: {total_pelotones:,}".replace(',', '.')
    texto_operaciones = f"OPE: {operaciones:,}".replace(',', '.')
    texto_entrenamiento = f"ENTR: {entrenamiento:,}".replace(',', '.')
    texto_descanso = f"DESC: {descanso:,}".replace(',', '.')

    # Escribir en PDF
    pdf.set_text_color(125, 0, 0)
    pdf.set_font('BebasNeue', '', 14)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)

    pos_y += 4
    pdf.text(pos_x, pos_y, texto_peloton)

    pos_y += 1
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)
    pdf.line(pos_x, pos_y, pos_x + 15, pos_y)

    pos_y += 5
    pdf.text(pos_x, pos_y, "DISPOSITIVO")

    pos_y += 4
    pdf.text(pos_x, pos_y, texto_operaciones)

    pos_y += 4
    pdf.text(pos_x, pos_y, texto_descanso)
    
    pos_y += 4
    pdf.text(pos_x, pos_y, texto_entrenamiento)
    
  
def pelotones_2_gaula(pdf, dato, dato_2, division, pos_x, pos_y):
    # Inicialización de contadores
    peloton = 0
    operaciones = 0
    descanso = 0
    entrenamiento = 0

    # Conteo de pelotones válidos
    for x in dato:
        if x[3] not in EXCLUIR_ESTADOS:
            peloton += 1

    # Clasificación por actividad
    for x in dato_2:
        if x[3] not in EXCLUIR_ESTADOS:
            if x[5] == "Descanso":
                descanso += 1
            elif x[5] == "Entrenamiento":
                entrenamiento += 1
            elif x[5] == "Operaciones":
                operaciones += 1

    # Reajuste de descanso
    descanso = peloton - operaciones - entrenamiento

    # Formateo de valores para mostrar
    peloton_txt = f"PEL: {peloton:,}".replace(",", ".")
    operaciones_txt = f"OPE: {operaciones:,}".replace(",", ".")
    entrenamiento_txt = f"ENTR: {entrenamiento:,}".replace(",", ".")
    descanso_txt = f"DESC: {descanso:,}".replace(",", ".")

    # Escritura en PDF
    pdf.set_text_color(125, 0, 0)
    pdf.set_font('BebasNeue', '', 14)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)

    pos_y += 4
    pdf.text(pos_x, pos_y, peloton_txt)

    pos_y += 1
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)
    pdf.line(pos_x, pos_y, pos_x + 15, pos_y)

    pos_y += 5
    pdf.text(pos_x, pos_y, "DISPOSITIVO")
    pos_y += 4
    pdf.text(pos_x, pos_y, operaciones_txt)
    pos_y += 4
    pdf.text(pos_x, pos_y, entrenamiento_txt)
    pos_y += 4
    pdf.text(pos_x, pos_y, descanso_txt)
 
def pelotones(pdf, dato, dato_2, division, pos_x, pos_y, div_escudo):
    """Agrega información de pelotones a un PDF con estadísticas por estado y división."""

    # Cargar imagen de escudo de la división
    imagen = f"C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/{div_escudo}.png"
    pdf.image(imagen, pos_x - 6, pos_y - 5, 6, 6)

    # Inicializar contadores
    total_pelotones = 0
    operaciones = 0
    descanso = 0
    entrenamiento = 0


    # Contar pelotones válidos
    total_pelotones = sum(1 for x in dato if x[3] not in EXCLUIR_ESTADOS)

    # Clasificar estados de pelotones válidos
    for x in dato_2:
        if x[3] not in EXCLUIR_ESTADOS:
            estado = x[5]
            if estado == "Operaciones":
                operaciones += 1
            elif estado == "Entrenamiento":
                entrenamiento += 1

    # Calcular pelotones en descanso como residuo
    descanso = total_pelotones - operaciones - entrenamiento

    # Formato de salida con separador de miles
    def format_label(label, value):
        return f"{label} {format(value, ',').replace(',', '.')}"

    peloton_label = format_label("PEL:", total_pelotones)
    operaciones_label = format_label("OPER:", operaciones)
    entrenamiento_label = format_label("ENTR:", entrenamiento)
    descanso_label = format_label("DESC:", descanso)

    # Dibujar en PDF
    pdf.set_text_color(125, 0, 0)
    pdf.set_font('BebasNeue', '', 14)
    pdf.text(pos_x, pos_y, str(division))

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)

    pos_y += 4
    pdf.text(pos_x, pos_y, peloton_label)

    pos_y += 1
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)
    pdf.line(pos_x, pos_y, pos_x + 15, pos_y)

    pos_y += 5
    pdf.text(pos_x, pos_y, "DISPOSITIVO")
    pos_y += 4
    pdf.text(pos_x, pos_y, operaciones_label)
    pos_y += 4
    pdf.text(pos_x, pos_y, descanso_label)
    pos_y += 4
    pdf.text(pos_x, pos_y, entrenamiento_label)
   

def cuadros_divisiones_pelotones_gaula(pdf, dato, dato_2, pos_x, pos_y, salto):
    """Genera cuadros por brigada y unidad con resumen de pelotones y dispositivos."""


    unidades = sorted(set(x[2] for x in dato))
    brigadas = sorted(set(x[1] for x in dato))

    pdf.ln(salto)

    for brigada in brigadas:
        # Filtrar solo pelotones válidos de la brigada actual
        tamanio = [x for x in dato if x[1] == brigada and x[3] not in EXCLUIR_ESTADOS]

        if len(tamanio) == 0:
            continue

        # Encabezado de brigada
        pdf.set_draw_color(40, 40, 40)
        pdf.set_fill_color(125, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('BebasNeue', '', 14)

        pdf.cell(pos_x)
        pdf.cell(120, 5, brigada, 1, 0, "C", fill=True)

        pdf.set_fill_color(80, 80, 80)
        pdf.cell(10)
        pdf.cell(90, 5, "DISPOSITIVOS", 1, 0, "C", fill=True)
        pdf.ln()

        # Encabezado de columnas
        pdf.set_fill_color(125, 0, 0)
        pdf.cell(pos_x)
        headers = ["NUM", "ARMA", "UNIDAD", "PELOTÓN"]
        for h in headers:
            pdf.cell(30, 5, h, 1, 0, "C", fill=True)

        pdf.set_fill_color(80, 80, 80)
        pdf.cell(10)
        dispositivos = ["OPERACIONES", "ENTRENAMIENTO", "DESCANSO"]
        for d in dispositivos:
            pdf.cell(30, 5, d, 1, 0, "C", fill=True)
        pdf.ln()

        # Datos por unidad
        pdf.set_text_color(40, 40, 40)
        pdf.set_font('BebasNeue', '', 12)

        total_brigada = total_operaciones = total_entrenamiento = total_descanso = 0
        num = 0

        for unidad in unidades:
            peloton = arma = brig = ""
            peloton_count = operaciones = entrenamiento = descanso = 0

            for x in dato:
                if x[3] in EXCLUIR_ESTADOS:
                    continue
                if x[1] == brigada and x[2] == unidad:
                    peloton_count += 1
                    brig = x[1]
                    arma = x[4]

            for x in dato_2:
                if x[3] in EXCLUIR_ESTADOS:
                    continue
                if x[1] == brigada and x[2] == unidad:
                    estado = x[5]
                    if estado == "Operaciones":
                        operaciones += 1
                    elif estado == "Entrenamiento":
                        entrenamiento += 1
                    elif estado == "Descanso":
                        descanso += 1

            if peloton_count == 0:
                continue

            num += 1
            total_brigada += peloton_count
            total_operaciones += operaciones
            total_entrenamiento += entrenamiento
            total_descanso += descanso

            # Fila de unidad
            pdf.cell(pos_x)
            pdf.cell(30, 5, str(num), 1, 0, "C")
            pdf.cell(30, 5, arma, 1, 0, "C")
            pdf.cell(30, 5, unidad, 1, 0, "C")
            pdf.cell(30, 5, f"{peloton_count:,}".replace(",", "."), 1, 0, "C")

            pdf.cell(10)
            for valor in [operaciones, entrenamiento, descanso]:
                txt = f"{valor:,}".replace(",", ".") if valor > 0 else ""
                pdf.cell(30, 5, txt, 1, 0, "C")
            pdf.ln()

        # Fila total por brigada
        if total_brigada > 0:
            pdf.set_fill_color(125, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('BebasNeue', '', 14)

            pdf.cell(pos_x)
            pdf.cell(90, 5, "TOTAL", 1, 0, "C", fill=True)
            pdf.cell(30, 5, f"{total_brigada:,}".replace(",", "."), 1, 0, "C", fill=True)

            pdf.set_fill_color(80, 80, 80)
            pdf.cell(10)
            for valor in [total_operaciones, total_entrenamiento, total_descanso]:
                txt = f"{valor:,}".replace(",", ".") if valor > 0 else ""
                pdf.cell(30, 5, txt, 1, 0, "C", fill=True)
            pdf.ln(5)
        pdf.ln(5)

    # Restaurar estilo por defecto
    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)
   
def cuadros_divisiones_pelotones(pdf, dato, dato_2, pos_x, pos_y, salto):
    unidad = sorted(set(x[2] for x in dato))
    brigadas = sorted(set(x[1] for x in dato))

    pdf.ln(salto)
    for brigada in brigadas:
        datos_brigada = [x for x in dato if x[1] == brigada and x[3] not in EXCLUIR_ESTADOS]

        if not datos_brigada:
            continue

        # Encabezado de la brigada
        pdf.set_draw_color(40, 40, 40)
        pdf.set_fill_color(125, 0, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('BebasNeue', '', 14)

        pdf.cell(pos_x)
        pdf.cell(120, 5, brigada, 1, 0, "C", fill=True)

        pdf.set_fill_color(80, 80, 80)
        pdf.cell(10)
        pdf.cell(90, 5, "DISPOSITIVOS", 1, 0, "C", fill=True)
        pdf.ln()

        # Subencabezados
        pdf.set_fill_color(125, 0, 0)
        pdf.cell(pos_x)
        for header in ["NUM", "ARMA", "UNIDAD", "PELOTÓN"]:
            pdf.cell(30, 5, header, 1, 0, "C", fill=True)

        pdf.set_fill_color(80, 80, 80)
        pdf.cell(10)
        for header in ["OPERACIONES", "ENTRENAMIENTO", "DESCANSO"]:
            pdf.cell(30, 5, header, 1, 0, "C", fill=True)
        pdf.ln()

        pdf.set_text_color(40, 40, 40)
        pdf.set_font('BebasNeue', '', 12)

        total_brigada = total_operaciones = total_descanso = total_entrenamiento = 0
        contador = 0

        for unidad_actual in unidad:
            datos_unidad = [x for x in dato if x[1] == brigada and x[2] == unidad_actual and x[3] not in EXCLUIR_ESTADOS]
            if not datos_unidad:
                continue

            peloton = len(datos_unidad)
            arma = datos_unidad[0][4] if datos_unidad else ""

            descanso = entremaniento = operaciones = 0
            for x in dato_2:
                if x[1] == brigada and x[2] == unidad_actual and x[3] not in EXCLUIR_ESTADOS:
                    if x[5] == "Descanso":
                        descanso += 1
                    elif x[5] == "Entrenamiento":
                        entremaniento += 1
                    elif x[5] == "Operaciones":
                        operaciones += 1

            descanso = peloton - operaciones - entremaniento
            contador += 1

            total_brigada += peloton
            total_operaciones += operaciones
            total_entrenamiento += entremaniento
            total_descanso += descanso

            pdf.cell(pos_x)
            pdf.cell(30, 5, str(contador), 1, 0, "C")
            pdf.cell(30, 5, arma, 1, 0, "C")
            pdf.cell(30, 5, unidad_actual, 1, 0, "C")
            pdf.cell(30, 5, format(peloton, ",").replace(",", "."), 1, 0, "C")

            pdf.cell(10)
            for val in [operaciones, entremaniento, descanso]:
                val_str = format(val, ",").replace(",", ".") if val > 0 else ""
                pdf.cell(30, 5, val_str, 1, 0, "C")
            pdf.ln()

        if total_brigada > 0:
            pdf.set_fill_color(125, 0, 0)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('BebasNeue', '', 14)
            pdf.cell(pos_x)
            pdf.cell(90, 5, "TOTAL", 1, 0, "C", fill=True)
            pdf.cell(30, 5, format(total_brigada, ",").replace(",", "."), 1, 0, "C", fill=True)

            pdf.set_fill_color(80, 80, 80)
            pdf.cell(10)
            for val in [total_operaciones, total_entrenamiento, total_descanso]:
                val_str = format(val, ",").replace(",", ".") if val > 0 else ""
                pdf.cell(30, 5, val_str, 1, 0, "C", fill=True)
            pdf.ln(5)
        pdf.ln(5)

    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)

def cuadros_divisiones_pelotones_pelotones(pdf, dato, dato_2, pos_x, pos_y, salto):
    
    
    unidades = sorted(set(x[2] for x in dato))
    brigadas = sorted(set(x[1] for x in dato))

    pdf.ln(salto)

    # Encabezado principal
    pdf.set_draw_color(40, 40, 40)
    pdf.set_fill_color(125, 0, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('BebasNeue', '', 14)
    pdf.cell(pos_x)
    pdf.cell(120, 5, "BRIGADA", 1, 0, "C", fill=True)

    pdf.set_fill_color(80, 80, 80)
    pdf.cell(10)
    pdf.cell(90, 5, "DISPOSITIVOS", 1, 0, "C", fill=True)
    pdf.ln()

    # Sub-encabezado
    pdf.set_fill_color(125, 0, 0)
    pdf.cell(pos_x)
    pdf.cell(30, 5, "NUM", 1, 0, "C", fill=True)
    pdf.cell(50, 5, "UNIDAD", 1, 0, "C", fill=True)
    pdf.cell(40, 5, "PELOTÓN", 1, 0, "C", fill=True)

    pdf.set_fill_color(80, 80, 80)
    pdf.cell(10)
    pdf.cell(30, 5, "OPERACIONES", 1, 0, "C", fill=True)
    pdf.cell(30, 5, "ENTRENAMIENTO", 1, 0, "C", fill=True)
    pdf.cell(30, 5, "DESCANSO", 1, 0, "C", fill=True)
    pdf.ln()

    pdf.set_text_color(40, 40, 40)
    pdf.set_font('BebasNeue', '', 12)

    num = 0

    for brigada in brigadas:
        # Filtrar registros válidos para esta brigada
        registros = [x for x in dato if x[1] == brigada and x[3] not in EXCLUIR_ESTADOS]
        if not registros:
            continue

        total_pelotones = 0
        total_operaciones = 0
        total_entrenamiento = 0

        for unidad in unidades:
            pelotones = [x for x in dato if x[1] == brigada and x[2] == unidad and x[3] not in EXCLUIR_ESTADOS]
            if not pelotones:
                continue

            cantidad_pelotones = len(pelotones)
            total_pelotones += cantidad_pelotones

            # Operaciones, entrenamiento y descanso
            dispositivos = [x for x in dato_2 if x[1] == brigada and x[2] == unidad and x[3] not in EXCLUIR_ESTADOS]
            operaciones = sum(1 for x in dispositivos if x[5] == "Operaciones")
            entrenamiento = sum(1 for x in dispositivos if x[5] == "Entrenamiento")
            descanso = cantidad_pelotones - operaciones - entrenamiento

            total_operaciones += operaciones
            total_entrenamiento += entrenamiento

        # Solo mostrar brigadas con pelotones
        if total_pelotones > 0:
            total_descanso = total_pelotones - total_operaciones - total_entrenamiento
            num += 1

            def fmt(value):
                return "" if value == 0 else format(value, ',').replace(',', '.')

            pdf.set_fill_color(125, 0, 0)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font('BebasNeue', '', 14)
            pdf.cell(pos_x)
            pdf.cell(30, 5, str(num), 1, 0, "C")
            pdf.cell(50, 5, brigada, 1, 0, "C")
            pdf.cell(40, 5, fmt(total_pelotones), 1, 0, "C")

            pdf.set_fill_color(80, 80, 80)
            pdf.cell(10)
            pdf.cell(30, 5, fmt(total_operaciones), 1, 0, "C")
            pdf.cell(30, 5, fmt(total_entrenamiento), 1, 0, "C")
            pdf.cell(30, 5, fmt(total_descanso), 1, 0, "C")
            pdf.ln()

    # Reset estilo
    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 12)
    pdf.set_draw_color(140, 140, 140)
    pdf.set_line_width(0.7)

def DIVI_PEL(pdf, dato, dato_2, pos_x, pos_y):
    """
    Inserta en el PDF un resumen de pelotones y su estado (operaciones, entrenamiento, descanso).
    - `dato` y `dato_2`: listas de listas con información de pelotones.
    - `pdf`: objeto PDF (por ejemplo, de FPDF).
    - `pos_x`, `pos_y`: posición inicial para insertar texto en el PDF.
    """

    # Inicialización de contadores
    peloton_total = 0
    operaciones = 0
    entrenamiento = 0

    # Estados que se excluyen del conteo


    # Conteo total de pelotones válidos
    for x in dato:
        if x[3] not in EXCLUIR_ESTADOS:
            peloton_total += 1

    # Conteo por actividad de los pelotones
    for x in dato_2:
        if x[3] not in EXCLUIR_ESTADOS:
            if x[5] == "Descanso":
                continue  # El descanso lo calcularemos al final
            elif x[5] == "Entrenamiento":
                entrenamiento += 1
            elif x[5] == "Operaciones":
                operaciones += 1

    # Calcular descanso como los restantes
    descanso = peloton_total - operaciones - entrenamiento

    # Formatear valores para presentación
    peloton_txt = f"PEL. {peloton_total:,}".replace(",", ".")
    operaciones_txt = f"OPER. {operaciones:,}".replace(",", ".")
    entrenamiento_txt = f"ENTR. {entrenamiento:,}".replace(",", ".")
    descanso_txt = f"DESC. {descanso:,}".replace(",", ".")

    # Estilo del texto en el PDF
    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 10)

    # Posicionar y escribir texto en el PDF
    pos_y += 4
    pdf.text(pos_x, pos_y, peloton_txt)

    pos_y += 4
    pdf.text(pos_x, pos_y, operaciones_txt)

    pos_y += 4
    pdf.text(pos_x, pos_y, f"{entrenamiento_txt} - {descanso_txt}")
   

def DIVI(pdf, dato, pos_x, pos_y):
    # Inicialización de acumuladores
    ofi = sub = slp = sl18 = sl12 = 0

    # Sumar valores por tipo
    for x in dato:
        ofi += int(x[16])
        sub += int(x[17])
        slp += int(x[18])
        sl18 += int(x[19])
        sl12 += int(x[20])

    # Totales
    total_sol = slp + sl18 + sl12
    total = ofi + sub + total_sol
    total_str = format(total, ',').replace(',', '.')

    # Cadena de resultados oficiales y suboficiales
    if total_sol > 0:
        resultados_cuadros = (
            "(" + 
            format(ofi, ',').replace(',', '.') + " - " + 
            format(sub, ',').replace(',', '.')
        )
    else:
        resultados_cuadros = (
            format(ofi, ',').replace(',', '.') + " - " +
            format(sub, ',').replace(',', '.') + " - " +
            format(total_sol, ',').replace(',', '.') + ")"
        )

    # Cadena de resultados SLP
    if sl18 > 0 or sl12 > 0:
        resultados_cuadros_slp = (
            format(slp, ',').replace(',', '.') + " - " +
            format(sl18, ',').replace(',', '.') + " - " +
            format(sl12, ',').replace(',', '.') + ")"
        )
    else:
        resultados_cuadros_slp = format(slp, ',').replace(',', '.') + ")"

    # Escribir en PDF
    pdf.set_text_color(10, 10, 10)
    pdf.set_font('BebasNeue', '', 10)

    pos_y += 4
    pdf.text(pos_x, pos_y, total_str)

    pos_y += 4
    pdf.text(pos_x, pos_y, resultados_cuadros)

    pos_y += 4
    if total_sol > 0:
        pdf.text(pos_x, pos_y, resultados_cuadros_slp)
 
def cuadros_peloton(pdf, CUNDINAMARCA, DATO_2, lugar, pos_x, pos_y, r,g,b):


    pos_x_x = pos_x + 2
    pos_y_y = pos_y + 1

    pdf.set_text_color(10,10,10)
    pdf.set_font('BebasNeue', '', 10)

    pdf.set_fill_color(r, g, b)
    pdf.set_draw_color(r,g,b)
    pdf.rounded_rect(pos_x, pos_y, 20, 27, 1,'DF', '1234')



    pelotones_2(pdf, CUNDINAMARCA, DATO_2, lugar,  pos_x_x, pos_y_y)
    
def cuadros_peloton_gaula(pdf, CUNDINAMARCA, DATO_2, lugar, pos_x, pos_y, r,g,b):

  
    pos_x_x = pos_x + 2
    pos_y_y = pos_y + 1

    pdf.set_text_color(10,10,10)
    pdf.set_font('BebasNeue', '', 10)

    pdf.set_fill_color(r, g, b)
    pdf.set_draw_color(r,g,b)
    pdf.rounded_rect(pos_x, pos_y, 20, 27, 1,'DF', '1234')



    pelotones_2_gaula(pdf, CUNDINAMARCA, DATO_2, lugar,  pos_x_x, pos_y_y)

def cuadros_PEL(pdf, CUNDINAMARCA, DATO_2, lugar, pos_x, pos_y, r,g,b):

    pos_x_x = pos_x + 26.5
    pos_x_y = pos_x + 3

    pos_y_x = pos_y + 7
    pos_y_y = pos_y - 1

    pdf.set_text_color(10,10,10)
    pdf.set_font('BebasNeue', '', 10)

    pdf.set_fill_color(r, g, b)
    pdf.set_draw_color(r,g,b)
    pdf.rounded_rect(pos_x, pos_y, 25, 11, 1,'DF', '1234')
    pdf.text(pos_x_y,pos_y_x,str(lugar))


    DIVI_PEL(pdf, CUNDINAMARCA, DATO_2, pos_x_x, pos_y_y)


def cuadros(pdf, CUNDINAMARCA, lugar, pos_x, pos_y, r,g,b):

    pos_x_x = pos_x + 26.5
    pos_x_y = pos_x + 3

    pos_y_x = pos_y + 7
    pos_y_y = pos_y - 1

    pdf.set_text_color(10,10,10)
    pdf.set_font('BebasNeue', '', 10)

    pdf.set_fill_color(r, g, b)
    pdf.set_draw_color(r,g,b)
    pdf.rounded_rect(pos_x, pos_y, 25, 11, 1,'DF', '1234')
    pdf.text(pos_x_y,pos_y_x,str(lugar))


    DIVI(pdf, CUNDINAMARCA, pos_x_x, pos_y_y)



#Funcion de resultados nueva ayuda
def connect():
    conn = psycopg2.connect(" \
        dbname=dirop \
        user=postgres \
        password=NICval10**")
    return conn


def graficas_estado_mayor_unidades(dato):

    fecha_insitop = dato["fecha_insitop"]
    tipo_filtro = dato["tipo_filtro"]
    unidad = dato["unidad"]
    label = []
    unidad_f_a = dato["unidad_f"]

   
    numero=0
    filtros = ""
    if tipo_filtro == "SIGLA_DIVISION":
        filtros = "SIGLA_BRIGADA"
        numero=10
    elif tipo_filtro == "SIGLA_BRIGADA":
        filtros = "SIGLA_UNIDD"
        numero=11
    elif tipo_filtro == "SIGLA_UNIDD":
        filtros = "COMPANIA"
        numero=12
    elif tipo_filtro == "COMPANIA":
        filtros = "PELOTON"
        numero=13
    else:
        filtros = "SIGLA_DIVISION"
        numero=9

    opciones =""
    if  unidad == "DIV02":
        opciones = " and {} = 'DIV02'  or {} = 'CENOR' or {} = 'FUDRA3'  ".format(tipo_filtro,tipo_filtro,tipo_filtro)   

    elif unidad == "DIV03":
        opciones = """ and {} = 'DIV03'  or {} = 'FTCEC'  """.format(tipo_filtro,tipo_filtro)  

    elif unidad == "DIV07":
        opciones = """ and {} = 'DIV07'  or {} = 'FTCTI'  """.format(tipo_filtro, tipo_filtro)  

    elif unidad == "DIV08":
        opciones = """ and {} = 'DIV08'  or {} = 'FURON'  """.format(tipo_filtro, tipo_filtro)  
    
    elif unidad == "FUTCO":
        opciones = """ and {} = 'FUTCO'  or {} = 'FUDRA1'  """.format(tipo_filtro, tipo_filtro)  

    elif unidad == "DAVAA":
        opciones = """ and {} = 'DAVAA'  or {} = 'CONAT'  """.format(tipo_filtro, tipo_filtro)  
    
    else:
        if tipo_filtro == "COMPANIA":
            opciones = """ and SIGLA_UNIDD = '{}' and  {} = '{}' """.format(unidad_f_a, tipo_filtro, unidad)  
        else:
            opciones = """ and {} = '{}' """.format(tipo_filtro, unidad)  
 
  
    conn = connect()
    cursor = conn.cursor()

    query_2 = """ select DISTINCT {} from  registro_insitop where FECHA_INSITOP like '%{}%'  {} ORDER BY {} ASC ;""".format(filtros, fecha_insitop, opciones, filtros)
 
    cursor.execute(query_2)
    label = cursor.fetchall()
    

    query = """ select * from  registro_insitop where FECHA_INSITOP like '%{}%' {} ORDER BY FECHA_INSITOP ASC ;""".format(fecha_insitop, opciones)
    cursor.execute(query)
    dat_2 = cursor.fetchall()

    
    valor = []

    total_descanso_data =[]
    total_entrenamiento_data =[]
    total_operaciones_data =[]
    total_emb_data =[]
    total_novedades_data =[]
        
    conn.commit()
    conn.close()
    cursor.close


    for lab in label:


        ofi_descanso = 0
        sub_descanso = 0
        slp_descanso = 0
        sl18_descanso = 0
        sl12_descanso = 0
        total_descanso = 0

        ofi_entrenamiento = 0
        sub_entrenamiento = 0
        slp_entrenamiento = 0
        sl18_entrenamiento = 0
        sl12_entrenamiento = 0
        total_entrenamiento = 0

        ofi_novedades = 0
        sub_novedades = 0
        slp_novedades = 0
        sl18_novedades = 0
        sl12_novedades = 0
        total_novedades = 0

        
        ofi_operaciones = 0
        sub_operaciones = 0
        slp_operaciones = 0
        sl18_operaciones = 0
        sl12_operaciones = 0
        total_operaciones = 0

        
        ofi_emb = 0
        sub_emb = 0
        slp_emb = 0
        sl18_emb = 0
        sl12_emb = 0
        total_emb = 0

        ofi = 0
        sub = 0
        slp = 0
        sl18 = 0
        sl12 = 0

        total= 0
        
        for x in dat_2:
  
            if lab[0]==x[numero]:

                                       
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])

                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                    total_descanso =  ofi_descanso + sub_descanso + slp_descanso + sl18_descanso + sl12_descanso
                    total_entrenamiento =  ofi_entrenamiento + sub_entrenamiento + slp_entrenamiento + sl18_entrenamiento + sl12_entrenamiento
                    total_operaciones =  ofi_operaciones + sub_operaciones + slp_operaciones + sl18_operaciones + sl12_operaciones
                    total_emb =  ofi_emb + sub_emb + slp_emb + sl18_emb + sl12_emb
                        
            
                    total = ofi + sub + slp + sl18 + sl12
                    total_novedades =  total - (total_descanso + total_entrenamiento + total_operaciones + total_emb)

 

        total_descanso_data.append(total_descanso)
        total_entrenamiento_data.append(total_entrenamiento)
        total_operaciones_data.append(total_operaciones)
        total_emb_data.append(total_emb)
        total_novedades_data.append(total_novedades)


    return[total_descanso_data, total_entrenamiento_data, total_operaciones_data, total_emb_data, total_novedades_data, label, filtros, unidad]

def graficas_estado_mayor(dato):

    fecha_insitop = dato["fecha_insitop"]
  
    conn = connect()
    cursor = conn.cursor()

    query = """ select * from  registro_insitop where FECHA_INSITOP like '%{}%' ORDER BY FECHA_INSITOP ASC ;""".format(fecha_insitop)
  

    cursor.execute(query)
    dat_2 = cursor.fetchall()


    label = ["DIV01", "DIV02", "DIV03", "DIV04", "DIV05", "DIV06", "DIV07", "DIV08", "FUTCO", "DAVAA", "BLICA", "CAAID", "CAOCC", "COING", "JEMOP"]
    valor = []

    total_descanso_data =[]
    total_entrenamiento_data =[]
    total_operaciones_data =[]
    total_emb_data =[]
    total_novedades_data =[]
        
    conn.commit()
    conn.close()
    cursor.close

    for lab in label:

        ofi_descanso = 0
        sub_descanso = 0
        slp_descanso = 0
        sl18_descanso = 0
        sl12_descanso = 0
        total_descanso = 0

        ofi_entrenamiento = 0
        sub_entrenamiento = 0
        slp_entrenamiento = 0
        sl18_entrenamiento = 0
        sl12_entrenamiento = 0
        total_entrenamiento = 0

        ofi_novedades = 0
        sub_novedades = 0
        slp_novedades = 0
        sl18_novedades = 0
        sl12_novedades = 0
        total_novedades = 0

        
        ofi_operaciones = 0
        sub_operaciones = 0
        slp_operaciones = 0
        sl18_operaciones = 0
        sl12_operaciones = 0
        total_operaciones = 0

        
        ofi_emb = 0
        sub_emb = 0
        slp_emb = 0
        sl18_emb = 0
        sl12_emb = 0
        total_emb = 0

        ofi = 0
        sub = 0
        slp = 0
        sl18 = 0
        sl12 = 0

        total= 0
        
        for x in dat_2:

            

            if lab==x[9]:

                if  "DIV02" == x[9] or "CENOR" == x[9] or "FUDRA3" == x[9]:

                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])

                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                elif "DIV03" == x[9] or "FTCEC" == x[9]:
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])
                              
                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                elif "DIV07" == x[9] or "FTCTI" == x[9]:
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])
                                    
                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                elif "DIV08" == x[9] or "FURON":
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])
                              
                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                elif "FUTCO" == x[9] or "FUDRA1" == x[9]:
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])
                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                elif "DAVAA" == x[9] or "CONAT" == x[9]:
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])
                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])
                else:
                                        
                    ofi = ofi +int(x[16])
                    sub = sub +int(x[17])
                    slp = slp +int(x[18])
                    sl18 = sl18 +int(x[19])
                    sl12 = sl12 +int(x[20])

                    if x[25] == "Descanso":

                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_descanso = ofi_descanso +int(x[16])
                            sub_descanso = sub_descanso +int(x[17])
                            slp_descanso = slp_descanso +int(x[18])
                            sl18_descanso = sl18_descanso +int(x[19])
                            sl12_descanso = sl12_descanso +int(x[20])

                    if x[25] == "Entrenamiento":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_entrenamiento = ofi_entrenamiento +int(x[16])
                            sub_entrenamiento = sub_entrenamiento +int(x[17])
                            slp_entrenamiento = slp_entrenamiento +int(x[18])
                            sl18_entrenamiento = sl18_entrenamiento +int(x[19])
                            sl12_entrenamiento = sl12_entrenamiento +int(x[20])
                    
                    if x[25] == "Novedades" or x[12] == "NOV"  :

                            ofi_novedades = ofi_novedades +int(x[16])
                            sub_novedades = sub_novedades +int(x[17])
                            slp_novedades = slp_novedades +int(x[18])
                            sl18_novedades = sl18_novedades +int(x[19])
                            sl12_novedades = sl12_novedades +int(x[20])
                    
                    if x[25] == "Operaciones":
                        if  x[12] != "NOV" and x[12] != "CDO" and x[12] != "EM" and x[12] != "EMB" and x[12] != "EMD" and x[12] != "PLM":
                            ofi_operaciones = ofi_operaciones +int(x[16])
                            sub_operaciones = sub_operaciones +int(x[17])
                            slp_operaciones = slp_operaciones +int(x[18])
                            sl18_operaciones = sl18_operaciones +int(x[19])
                            sl12_operaciones = sl12_operaciones +int(x[20])

                    if x[12] == "CDO" or x[12] == "EM" or x[12] == "EMB" or x[12] == "EMD" or x[12] == "PLM":
                        ofi_emb = ofi_emb +int(x[16])
                        sub_emb = sub_emb +int(x[17])
                        slp_emb = slp_emb +int(x[18])
                        sl18_emb = sl18_emb +int(x[19])
                        sl12_emb = sl12_emb +int(x[20])

                total_descanso =  ofi_descanso + sub_descanso + slp_descanso + sl18_descanso + sl12_descanso
                total_entrenamiento =  ofi_entrenamiento + sub_entrenamiento + slp_entrenamiento + sl18_entrenamiento + sl12_entrenamiento
                total_operaciones =  ofi_operaciones + sub_operaciones + slp_operaciones + sl18_operaciones + sl12_operaciones
                total_emb =  ofi_emb + sub_emb + slp_emb + sl18_emb + sl12_emb
                
    
                total = ofi + sub + slp + sl18 + sl12
                total_novedades =  total - (total_descanso + total_entrenamiento + total_operaciones + total_emb)
  

        total_descanso_data.append(total_descanso)
        total_entrenamiento_data.append(total_entrenamiento)
        total_operaciones_data.append(total_operaciones)
        total_emb_data.append(total_emb)
        total_novedades_data.append(total_novedades)
            


    return[total_descanso_data, total_entrenamiento_data, total_operaciones_data, total_emb_data, total_novedades_data, label]

def graficas_listado_dia_unidades(dato):

    tipo_filtro = dato["tipo_filtro"]
    unidad_f = dato["unidad"]
    unidad_f_a = dato["unidad_f"]
    label = []

    numero=0
    filtros = ""
    if tipo_filtro == "SIGLA_DIVISION":
        filtros = "SIGLA_BRIGADA"
        numero=10
    elif tipo_filtro == "SIGLA_BRIGADA":
        filtros = "SIGLA_UNIDD"
        numero=11
    elif tipo_filtro == "SIGLA_UNIDD":
        filtros = "COMPANIA"
        numero=12
    elif tipo_filtro == "COMPANIA":
        filtros = "PELOTON"
        numero=13
    else:
        filtros = "SIGLA_DIVISION"
        numero=9

    opciones =""

    if  unidad_f == "DIV02":
        opciones = "  {} = 'DIV02'  or {} = 'CENOR' or {} = 'FUDRA3'  ".format(tipo_filtro,tipo_filtro,tipo_filtro)   

    elif unidad_f == "DIV03":
        opciones = """  {} = 'DIV03'  or {} = 'FTCEC'  """.format(tipo_filtro,tipo_filtro)  

    elif unidad_f == "DIV07":
        opciones = """  {} = 'DIV07'  or {} = 'FTCTI'  """.format(tipo_filtro, tipo_filtro)  

    elif unidad_f == "DIV08":
        opciones = """ {} = 'DIV08'  or {} = 'FURON'  """.format(tipo_filtro, tipo_filtro)  
    
    elif unidad_f == "FUTCO":
        opciones = """  {} = 'FUTCO'  or {} = 'FUDRA1'  """.format(tipo_filtro, tipo_filtro)  

    elif unidad_f == "DAVAA":
        opciones = """  {} = 'DAVAA'  or {} = 'CONAT'  """.format(tipo_filtro, tipo_filtro)  
    
    else:
        if tipo_filtro == "COMPANIA":
            opciones = """ SIGLA_UNIDD = '{}' and  {} = '{}' """.format(unidad_f_a, tipo_filtro, unidad_f)  
        else:
            opciones = """  {} = '{}' """.format(tipo_filtro, unidad_f) 
 
  
    conn = connect()
    cursor = conn.cursor()

    query_2 = """ select DISTINCT {} from  registro_insitop where  {} ORDER BY {} ASC ;""".format(filtros,  opciones, filtros)
    cursor.execute(query_2)
    label = cursor.fetchall()
    

    query = """ select * from  registro_insitop where {} ORDER BY FECHA_INSITOP ASC ;""".format(opciones)
    cursor.execute(query)
    dat_2 = cursor.fetchall()

    
    query = """ select DISTINCT FECHA_INSITOP from  registro_insitop  where {} ORDER BY FECHA_INSITOP ASC ;""".format(opciones)

    cursor.execute(query)
    data = cursor.fetchall()


    
    

    valor = []

    unidades = []

    conn.commit()
    conn.close()
    cursor.close
    label_2=[]
    for fec in data:
        dia = fec[0].replace("00:00:00","")
        label_2.append(dia)
        
    for lab in label:
       
        valor_div01=[]
        
        for fec in data:



            total_unidad=0
            
            
            div01_total=0
            
            DATO_1_fecha = list(filter(lambda dat_2: fec[0] == dat_2[2], dat_2))

            DATO_1 = list(filter(lambda DATO_1_fecha: lab[0] == DATO_1_fecha[numero], DATO_1_fecha))



            ofi = 0
            sub = 0
            slp = 0
            sl18 = 0
            sl12 = 0
            for x in DATO_1_fecha:
                
                ofi = ofi +int(x[16])
                sub = sub +int(x[17])
                slp = slp +int(x[18])
                sl18 = sl18 +int(x[19])
                sl12 = sl12 +int(x[20])
            total_unidad = ofi + sub + slp + sl18 + sl12
 
            div01=0
            ofi_div01 = 0
            sub_div01 = 0
            slp_div01 = 0
            sl18_div01 = 0
            sl12_div01 = 0 
            for x in DATO_1:
                
                ofi_div01 = ofi_div01 +int(x[16])
                sub_div01 = sub_div01 +int(x[17])
                slp_div01 = slp_div01 +int(x[18])
                sl18_div01 = sl18_div01 +int(x[19])
                sl12_div01 = sl12_div01 +int(x[20])

                div01 = ofi_div01 + sub_div01 + slp_div01 + sl18_div01 + sl12_div01
            div01_total = div01_total + div01


                
            valor_div01.append(div01_total)
            #general    
            
            valor.append(total_unidad)     
            
            #general
        unidades.append(valor_div01)
        
        
    return[label_2, valor, unidades, label, filtros, unidad_f]


def graficas(dato):


    conn = connect()
    cursor = conn.cursor()


    query = """ select DISTINCT FECHA_INSITOP from  registro_insitop ORDER BY FECHA_INSITOP ASC ;"""
    cursor.execute(query)
    data = cursor.fetchall()

    query = """ select * from  registro_insitop ORDER BY FECHA_INSITOP ASC ;"""
    cursor.execute(query)
    dat_2 = cursor.fetchall()

    query_3 = """ select DISTINCT SIGLA_DIVISION from  registro_insitop ORDER BY SIGLA_DIVISION ASC ;"""
    cursor.execute(query_3)
    data_3 = cursor.fetchall()
    
    for unidad in data_3:

        if "DIV01" == unidad[0]:
            DIV01 = list(filter(lambda dat_2: "DIV01" == dat_2[9], dat_2))
            unidad_div01 = "DIV01"

        elif "DIV02" == unidad[0] or "CENOR" == unidad[0] or "FUDRA3" == unidad[0]:
            DIV02 = list(filter(lambda dat_2: "DIV02" == dat_2[9] or "CENOR" == dat_2[9] or "FUDRA3" == dat_2[9], dat_2))
            unidad_div02 = "DIV02"

        elif "DIV03" == unidad[0] or "FTCEC" == unidad[0]:
            DIV03 = list(filter(lambda dat_2: "DIV03" == dat_2[9] or "FTCEC" == dat_2[9], dat_2))
            unidad_div03 = "DIV03"

        elif "DIV04" == unidad[0]:
            DIV04 = list(filter(lambda dat_2: "DIV04" == dat_2[9], dat_2))
            unidad_div04 = "DIV04"

        elif "DIV05" == unidad[0]:
            DIV05 = list(filter(lambda dat_2: "DIV05" == dat_2[9], dat_2))
            unidad_div05 = "DIV05"

        elif "DIV06" == unidad[0]:
            DIV06 = list(filter(lambda dat_2: "DIV06" == dat_2[9], dat_2))
            unidad_div06 = "DIV06"

        elif "DIV07" == unidad[0] or "FTCTI" == unidad[0]:
            DIV07 = list(filter(lambda dat_2: "DIV07" == dat_2[9] or "FTCTI" == dat_2[9], dat_2))
            unidad_div07 = "DIV07"

        elif "DIV08" == unidad[0] or "FURON" == unidad[0]:
            DIV08 = list(filter(lambda dat_2: "DIV08" == dat_2[9] or "FURON" == dat_2[9], dat_2))
            unidad_div08 = "DIV08"

        elif "FUTCO" == unidad[0] or "FUDRA1" == unidad[0]:
            FUTCO = list(filter(lambda dat_2: "FUTCO" == dat_2[9] or "FUDRA1" == dat_2[9], dat_2))
            unidad_FUTCO = "FUTCO"

        elif "DAVAA" == unidad[0] or "CONAT" == unidad[0]:
            DAVAA = list(filter(lambda dat_2: "DAVAA" == dat_2[9] or "CONAT" == dat_2[9], dat_2))
            unidad_DAVAA = "DAVAA"

        elif "BLICA" == unidad[0]:
            BLICA = list(filter(lambda dat_2: "BLICA" == dat_2[9], dat_2))
            unidad_BLICA = "BLICA"

        elif "CAAID" == unidad[0]:
            CAAID = list(filter(lambda dat_2: "CAAID" == dat_2[9], dat_2))
            unidad_CAAID = "CAAID"

        elif "CAOCC" == unidad[0]:
            CAOCC = list(filter(lambda dat_2: "CAOCC" == dat_2[9], dat_2))
            unidad_CAOCC = "CAOCC"

        elif "COING" == unidad[0]:
            COING = list(filter(lambda dat_2: "COING" == dat_2[9], dat_2))
            unidad_COING = "COING"
        elif "JEMOP" == unidad[0]:
            JEMOP = list(filter(lambda dat_2: "JEMOP" == dat_2[9], dat_2))
            unidad_COING = "JEMOP"
        else:
            TREJC = list(filter(lambda dat_2: "TREJC" == dat_2[9], dat_2))
            unidad_COING = "TREJC"

    
    labe_unidades=["DIV01", "DIV02", "DIV03", "DIV04", "DIV05", "DIV06", "DIV07", "DIV08", "FUTCO", "DAVAA", "BLICA", "CAAID", "CAOCC", "COING", "JEMOP", "TREJC" ]


    label = []
    valor = []
    valor_div01 = []
    valor_div02 = []
    valor_div03 = []
    valor_div04 = []
    valor_div05 = []
    valor_div06 = []
    valor_div07 = []
    valor_div08 = []
    valor_DAVAA = []
    valor_FUTCO = []
    valor_BLICA = []
    valor_CAAID = []
    valor_CAOCC = []
    valor_COING = []
    valor_JEMOP = []
    valor_TRECJ = []
    

    for da in data:
        
        ofi = 0
        sub = 0
        slp = 0
        sl18 = 0
        sl12 = 0

        total= 0
        total_div=0
        total_div02=0
        total_div03=0

        ofi_div01=0
        sub_div01=0
        slp_div01=0
        sl18_div01=0
        sl12_div01=0

        ofi_div02 = 0
        sub_div02 = 0
        slp_div02 = 0
        sl18_div02 = 0
        sl12_div02 = 0

        ofi_div03 = 0
        sub_div03 = 0
        slp_div03 = 0
        sl18_div03 = 0
        sl12_div03  = 0
        
        ofi_div04 = 0
        sub_div04 = 0
        slp_div04 = 0
        sl18_div04 = 0
        sl12_div04  = 0

        ofi_div05 = 0
        sub_div05 = 0
        slp_div05 = 0
        sl18_div05 = 0
        sl12_div05  = 0

        ofi_div06 = 0
        sub_div06 = 0
        slp_div06 = 0
        sl18_div06 = 0
        sl12_div06  = 0
        
        ofi_div07 = 0
        sub_div07 = 0
        slp_div07 = 0
        sl18_div07 = 0
        sl12_div07  = 0

        ofi_div08 = 0
        sub_div08 = 0
        slp_div08 = 0
        sl18_div08 = 0
        sl12_div08  = 0

        ofi_FUTCO = 0
        sub_FUTCO = 0
        slp_FUTCO = 0
        sl18_FUTCO = 0
        sl12_FUTCO  = 0

        ofi_DAVAA = 0
        sub_DAVAA = 0
        slp_DAVAA = 0
        sl18_DAVAA = 0
        sl12_DAVAA  = 0

        ofi_BLICA = 0
        sub_BLICA = 0
        slp_BLICA = 0
        sl18_BLICA = 0
        sl12_BLICA  = 0

        ofi_CAAID = 0
        sub_CAAID = 0
        slp_CAAID = 0
        sl18_CAAID = 0
        sl12_CAAID  = 0

        ofi_CAOCC = 0
        sub_CAOCC = 0
        slp_CAOCC = 0
        sl18_CAOCC = 0
        sl12_CAOCC  = 0

        ofi_COING = 0
        sub_COING = 0
        slp_COING = 0
        sl18_COING = 0
        sl12_COING  = 0
        
        ofi_JEMOP = 0
        sub_JEMOP = 0
        slp_JEMOP = 0
        sl18_JEMOP = 0
        sl12_JEMOP  = 0


        div01=0
        div02=0
        div03=0
        div04=0
        div05=0
        div06=0
        div07=0
        div08=0
        futco = 0
        davaa = 0
        blica = 0
        caaid = 0
        caocc = 0
        coing = 0
        jemop = 0
        trecj = 0

        ofi_trecj =0
        sub_trecj =0
        slp_trecj =0
        sl18_trecj =0
        sl12_trecj =0
        #print(str(res))

        DATO_1_fecha = list(filter(lambda dat_2: da[0] == dat_2[2], dat_2))

        DIV01_2= list(filter(lambda DIV01: da[0] == DIV01[2], DIV01))
        DIV02_2= list(filter(lambda DIV02: da[0] == DIV02[2], DIV02))
        DIV03_2= list(filter(lambda DIV03: da[0] == DIV03[2], DIV03))
        DIV04_2= list(filter(lambda DIV04: da[0] == DIV04[2], DIV04))
        DIV05_2= list(filter(lambda DIV05: da[0] == DIV05[2], DIV05))
        DIV06_2= list(filter(lambda DIV06: da[0] == DIV06[2], DIV06))
        DIV07_2= list(filter(lambda DIV07: da[0] == DIV07[2], DIV07))
        DIV08_2= list(filter(lambda DIV08: da[0] == DIV08[2], DIV08))
        FUTCO_2= list(filter(lambda FUTCO: da[0] == FUTCO[2], FUTCO))
        DAVAA_2= list(filter(lambda DAVAA: da[0] == DAVAA[2], DAVAA))
        BLICA_2= list(filter(lambda BLICA: da[0] == BLICA[2], BLICA))
        CAAID_2= list(filter(lambda CAAID: da[0] == CAAID[2], CAAID))
        CAOCC_2= list(filter(lambda CAOCC: da[0] == CAOCC[2], CAOCC))
        COING_2= list(filter(lambda COING: da[0] == COING[2], COING))
        JEMOP_2= list(filter(lambda JEMOP: da[0] == JEMOP[2], JEMOP))
        TREJC_2= list(filter(lambda TREJC: da[0] == TREJC[2], TREJC))

        

        for x in DATO_1_fecha:

            ofi = ofi +int(x[16])
            sub = sub +int(x[17])
            slp = slp +int(x[18])
            sl18 = sl18 +int(x[19])
            sl12 = sl12 +int(x[20])
        total = ofi + sub + slp + sl18 + sl12
    
        for x in DIV01_2:
            ofi_div01 = ofi_div01 +int(x[16])
            sub_div01 = sub_div01 +int(x[17])
            slp_div01 = slp_div01 +int(x[18])
            sl18_div01 = sl18_div01 +int(x[19])
            sl12_div01 = sl12_div01 +int(x[20])

        div01 = ofi_div01 + sub_div01 + slp_div01 + sl18_div01 + sl12_div01
        valor_div01.append(div01) 
                            
        for x in DIV02_2:
            ofi_div02 = ofi_div02 +int(x[16])
            sub_div02 = sub_div02 +int(x[17])
            slp_div02 = slp_div02 +int(x[18])
            sl18_div02 = sl18_div02 +int(x[19])
            sl12_div02 = sl12_div02 +int(x[20])

        div02 = ofi_div02 + sub_div02 + slp_div02 + sl18_div02 + sl12_div02 
        valor_div02.append(div02)


        for x in DIV03_2:
            ofi_div03 = ofi_div03 +int(x[16])
            sub_div03 = sub_div03 +int(x[17])
            slp_div03 = slp_div03 +int(x[18])
            sl18_div03 = sl18_div03 +int(x[19])
            sl12_div03 = sl12_div03 +int(x[20])

        div03 = ofi_div03 + sub_div03 + slp_div03 + sl18_div03 + sl12_div03 
        valor_div03.append(div03)
                
        for x in DIV04_2:
            ofi_div04 = ofi_div04 +int(x[16])
            sub_div04 = sub_div04 +int(x[17])
            slp_div04 = slp_div04 +int(x[18])
            sl18_div04 = sl18_div04 +int(x[19])
            sl12_div04 = sl12_div04 +int(x[20])

        div04 = ofi_div04 + sub_div04 + slp_div04 + sl18_div04 + sl12_div04
        valor_div04.append(div04) 
                               
        for x in DIV05_2:
            ofi_div05 = ofi_div05 +int(x[16])
            sub_div05 = sub_div05 +int(x[17])
            slp_div05 = slp_div05 +int(x[18])
            sl18_div05 = sl18_div05 +int(x[19])
            sl12_div05 = sl12_div05 +int(x[20])

        div05 = ofi_div05 + sub_div05 + slp_div05 + sl18_div05 + sl12_div05
        valor_div05.append(div05) 

        for x in DIV06_2:
            ofi_div06 = ofi_div06 +int(x[16])
            sub_div06 = sub_div06 +int(x[17])
            slp_div06 = slp_div06 +int(x[18])
            sl18_div06 = sl18_div06 +int(x[19])
            sl12_div06 = sl12_div06 +int(x[20])

        div06 = ofi_div06 + sub_div06 + slp_div06 + sl18_div06 + sl12_div06
        valor_div06.append(div06)  
                
        for x in DIV07_2:
            ofi_div07 = ofi_div07 +int(x[16])
            sub_div07 = sub_div07 +int(x[17])
            slp_div07 = slp_div07 +int(x[18])
            sl18_div07 = sl18_div07 +int(x[19])
            sl12_div07 = sl12_div07 +int(x[20])

        div07 = ofi_div07 + sub_div07 + slp_div07 + sl18_div07 + sl12_div07
        valor_div07.append(div07)  
                                
        for x in DIV08_2:
            ofi_div08 = ofi_div08 +int(x[16])
            sub_div08 = sub_div08 +int(x[17])
            slp_div08 = slp_div08 +int(x[18])
            sl18_div08 = sl18_div08 +int(x[19])
            sl12_div08 = sl12_div08 +int(x[20])

        div08 = ofi_div08 + sub_div08 + slp_div08 + sl18_div08 + sl12_div08
        valor_div08.append(div08) 

        for x in FUTCO_2:
            ofi_FUTCO = ofi_FUTCO +int(x[16])
            sub_FUTCO = sub_FUTCO +int(x[17])
            slp_FUTCO = slp_FUTCO +int(x[18])
            sl18_FUTCO = sl18_FUTCO +int(x[19])
            sl12_FUTCO = sl12_FUTCO +int(x[20])

        futco = ofi_FUTCO + sub_FUTCO + slp_FUTCO + sl18_FUTCO + sl12_FUTCO 

        valor_DAVAA.append(futco)


        for x in DAVAA_2:                               
            ofi_DAVAA= ofi_DAVAA+int(x[16])
            sub_DAVAA= sub_DAVAA+int(x[17])
            slp_DAVAA= slp_DAVAA+int(x[18])
            sl18_DAVAA= sl18_DAVAA+int(x[19])
            sl12_DAVAA= sl12_DAVAA+int(x[20])

        davaa= ofi_DAVAA+ sub_DAVAA+ slp_DAVAA+ sl18_DAVAA+ sl12_DAVAA
        valor_FUTCO.append(davaa)

        for x in BLICA_2:   
            ofi_BLICA= ofi_BLICA+int(x[16])
            sub_BLICA= sub_BLICA+int(x[17])
            slp_BLICA= slp_BLICA+int(x[18])
            sl18_BLICA= sl18_BLICA+int(x[19])
            sl12_BLICA= sl12_BLICA+int(x[20])

        blica= ofi_BLICA+ sub_BLICA+ slp_BLICA+ sl18_BLICA+ sl12_BLICA
        valor_BLICA.append(blica)

      
        for x in CAAID_2:                
            ofi_CAAID= ofi_CAAID+int(x[16])
            sub_CAAID= sub_CAAID+int(x[17])
            slp_CAAID= slp_CAAID+int(x[18])
            sl18_CAAID= sl18_CAAID+int(x[19])
            sl12_CAAID= sl12_CAAID+int(x[20])

        caaid= ofi_CAAID+ sub_CAAID+ slp_CAAID+ sl18_CAAID+ sl12_CAAID
        valor_CAAID.append(caaid)

                
        for x in CAOCC_2:
                ofi_CAOCC= ofi_CAOCC+int(x[16])
                sub_CAOCC= sub_CAOCC+int(x[17])
                slp_CAOCC= slp_CAOCC+int(x[18])
                sl18_CAOCC= sl18_CAOCC+int(x[19])
                sl12_CAOCC= sl12_CAOCC+int(x[20])
        caocc= ofi_CAOCC+ sub_CAOCC+ slp_CAOCC+ sl18_CAOCC+ sl12_CAOCC
        valor_CAOCC.append(caocc)


        for x in COING_2:         
                    ofi_COING= ofi_COING+int(x[16])
                    sub_COING= sub_COING+int(x[17])
                    slp_COING= slp_COING+int(x[18])
                    sl18_COING= sl18_COING+int(x[19])
                    sl12_COING= sl12_COING+int(x[20])
        coing= ofi_COING+ sub_COING+ slp_COING+ sl18_COING+ sl12_COING
        valor_COING.append(coing)
        
                                                
        for x in JEMOP_2:
                ofi_JEMOP= ofi_JEMOP+int(x[16])
                sub_JEMOP= sub_JEMOP+int(x[17])
                slp_JEMOP= slp_JEMOP+int(x[18])
                sl18_JEMOP= sl18_JEMOP+int(x[19])
                sl12_JEMOP= sl12_JEMOP+int(x[20])
        jemop= ofi_JEMOP+ sub_JEMOP+ slp_JEMOP+ sl18_JEMOP+ sl12_JEMOP
        valor_JEMOP.append(jemop)
          
        for x in TREJC_2:
                ofi_trecj= ofi_trecj+int(x[16])
                sub_trecj= sub_trecj+int(x[17])
                slp_trecj= slp_trecj+int(x[18])
                sl18_trecj= sl18_trecj+int(x[19])
                sl12_trecj= sl12_trecj+int(x[20])
        trecj= ofi_trecj+ sub_trecj+ slp_trecj+ sl18_trecj+ sl12_trecj
        valor_TRECJ.append(trecj)


        dia = da[0].replace("00:00:00","")
        label.append(dia)
        valor.append(total)

        
    unidades=[ valor_div01, valor_div02, valor_div03, valor_div04, valor_div05, valor_div06, valor_div07, valor_div08, valor_DAVAA, valor_FUTCO, valor_BLICA, valor_CAAID, valor_CAOCC, valor_COING, valor_JEMOP, valor_TRECJ]

    conn.commit()
    conn.close()
    cursor.close
    tipo_filtro = "SIGLA_DIVISION"
    return[label, valor, unidades, labe_unidades, tipo_filtro]

def listado_inistop(dato):


    conn = connect()
    cursor = conn.cursor()


    query = """ select DISTINCT FECHA_INSITOP from  registro_insitop ORDER BY FECHA_INSITOP DESC ;"""

    cursor.execute(query)
    data = cursor.fetchall()

    conn.commit()
    conn.close()
    cursor.close
    return[data]
# Validación de fechas
def validar_fecha(fecha_texto):
    try:
        return datetime.strptime(fecha_texto, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError(f"Fecha inválida: {fecha_texto}")

from collections import defaultdict
from flask import jsonify  # o FastAPI JSONResponse si usas FastAPI
from collections import defaultdict

def obtener_operaciones(dato):
    fechaInicio = validar_fecha(dato["fechaInicio"])
    fechaFin = validar_fecha(dato["fechaFin"])
    
    dpto =  dato["dpto"]
    mpio =  dato["mpio"]
    divisiones_filtrados =  dato["divisiones_filtrados"]

    dpto=dpto.split(",")
    mpio=mpio.split(",")
    divisiones_filtrados=divisiones_filtrados.split(",")
    where_dpto = ""
    where_mpio = ""
    where_unidad = ""

    item = json.loads(dato['subregion'])
    if isinstance(item, dict) and "nombre" in item:
        valor_1 = validacion_cantidad_tupla([d["nombre"] for d in item["departamentos"]])
        valor_2 = validacion_cantidad_tupla([
            m for d in item["departamentos"] for m in d["municipios"]
        ])

        where_dpto = "and DEPARTAMENTO {}".format(valor_1)

        where_mpio = "and MUNICIPIO {}".format(valor_2)

    else:
        valor_1 = validacion_cantidad_tupla(dpto)
        valor_2 = validacion_cantidad_tupla(mpio)
        
        if len(dpto[0])>1:
            where_dpto = "and DEPARTAMENTO {}".format(valor_1)

        if len(mpio[0])>1:
            where_mpio = "and MUNICIPIO {}".format(valor_2)
    

    conn = connect()
    cursor = conn.cursor()
    unidades ={
        "DIV01":["DIV01"],
        "DIV02":["DIV02","CENOR", "FUDRA3" ],
        "DIV03":["DIV03","FTCEC" ],
        "DIV04":["DIV04" ],
        "DIV05":["DIV05" ],
        "DIV06":["DIV06" ],
        "DIV07":["DIV07","FTCTI" ],
        "DIV08":["DIV08","FURON" ],
        "FUTOM":["FUTCO","FUDRA1","FUTOM" ],
        "DAVAA":["DAVAA","CONAT" ],
        "TREJC":["COLOG","CAFUE","COING","JEMOP","JEMGF","SECEJ","TEJC" ,"TREJC","CEDOC" ],
        "CAAID":["CAAID" ],
        "CAOCC":["CAOCC" ]
    }
    unidad_filtrada=[]
    if len(divisiones_filtrados[0])>1:
        for x in unidades:
            #print(x)
            #print(unidades[x])
            for y in divisiones_filtrados:
                if x == y:
                    for z in unidades[x]:
                        unidad_filtrada.append(z)

        valor_3 = validacion_cantidad_tupla(unidad_filtrada)
        where_unidad=  "and SIGLA_DIVISION {}".format(valor_3)

    conn = connect()
    cursor = conn.cursor()

    query_4_pto = """
        SELECT DISTINCT 
            SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, 
            LINEA_DE_OPERACIONES, ORD_ESC_BAT, CODE, 
            OPERACION, TAREAS_ACCION_DECISIVA_TTF, DEPARTAMENTO 
        FROM registro_insitop 
        WHERE TO_DATE(FECHA_INSITOP, 'YYYY-MM-DD') BETWEEN %s AND %s and CODE = 'Operaciones' {} {} {}
    """.format(where_dpto, where_mpio, where_unidad)

    cursor.execute(query_4_pto, (fechaInicio, fechaFin))
    data_4_pto = cursor.fetchall()
    cursor.close()
    conn.close()

    estructura = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for row in data_4_pto:
        (division, brigada, unidad, linea, orden, code, operacion, tarea, departamento) = row

        estructura[division][brigada][unidad].append({
            "linea": linea,
            "orden": orden,
            "code": code,
            "operacion": operacion,
            "tarea": tarea,
            "departamento": departamento
        })

    # Convertir defaultdicts a dict normales para respuesta JSON
    resultado = {div: {brig: dict(unids) for brig, unids in brigadas.items()} for div, brigadas in estructura.items()}

    return resultado  # ✅ ya no usamos jsonify

def calcular_resumen_personal(data):
    def fmt(num):
        return '{:,}'.format(num).replace(',', '.')

    ofi = sub = slp = sl18 = sl12 = 0
    for x in data:
        ofi += int(x[16])
        sub += int(x[17])
        slp += int(x[18])
        sl18 += int(x[19])
        sl12 += int(x[20])
    
    total_sol = slp + sl18 + sl12
    total = total_sol + ofi + sub
    total_p = fmt(total)

    if total_sol > 0:
        resultados_cuadros = f"{fmt(ofi)} - {fmt(sub)}"
    else:
        resultados_cuadros = f"{fmt(ofi)} - {fmt(sub)} - {fmt(total_sol)}"

    if sl18 > 0 or sl12 > 0:
        resultados_cuadros_slp = f"{fmt(slp)} - {fmt(sl18)} - {fmt(sl12)}"
    else:
        resultados_cuadros_slp = f"{fmt(slp)}"

    # Inicializar acumuladores por categoría
    ofi_descanso = sub_descanso = slp_descanso = sl18_descanso = sl12_descanso = 0
    ofi_entrenamiento = sub_entrenamiento = slp_entrenamiento = sl18_entrenamiento = sl12_entrenamiento = 0
    ofi_novedades = sub_novedades = slp_novedades = sl18_novedades = sl12_novedades = 0
    ofi_operaciones = sub_operaciones = slp_operaciones = sl18_operaciones = sl12_operaciones = 0
    ofi_emb = sub_emb = slp_emb = sl18_emb = sl12_emb = 0

    for x in data:
        cat = x[25]
        cod = x[12]

        if cat == "Descanso" and cod not in ["NOV", "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM"]:
            ofi_descanso += int(x[16])
            sub_descanso += int(x[17])
            slp_descanso += int(x[18])
            sl18_descanso += int(x[19])
            sl12_descanso += int(x[20])

        elif cat == "Entrenamiento" and cod not in ["NOV",  "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM"]:
            ofi_entrenamiento += int(x[16])
            sub_entrenamiento += int(x[17])
            slp_entrenamiento += int(x[18])
            sl18_entrenamiento += int(x[19])
            sl12_entrenamiento += int(x[20])

        elif cat == "Novedades" or cod == "NOV":
            ofi_novedades += int(x[16])
            sub_novedades += int(x[17])
            slp_novedades += int(x[18])
            sl18_novedades += int(x[19])
            sl12_novedades += int(x[20])

        elif cat == "Operaciones" and cod not in ["NOV",  "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM"]:
            ofi_operaciones += int(x[16])
            sub_operaciones += int(x[17])
            slp_operaciones += int(x[18])
            sl18_operaciones += int(x[19])
            sl12_operaciones += int(x[20])

        if cod in ["NOV",  "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM",  "P. MAYOR", "STAFF" ,  "PLM"]:
            ofi_emb += int(x[16])
            sub_emb += int(x[17])
            slp_emb += int(x[18])
            sl18_emb += int(x[19])
            sl12_emb += int(x[20])

    # Totales por grupo
    total_descanso = ofi_descanso + sub_descanso + slp_descanso + sl18_descanso + sl12_descanso
    total_entrenamiento = ofi_entrenamiento + sub_entrenamiento + slp_entrenamiento + sl18_entrenamiento + sl12_entrenamiento
    total_operaciones = ofi_operaciones + sub_operaciones + slp_operaciones + sl18_operaciones + sl12_operaciones
    total_emb = ofi_emb + sub_emb + slp_emb + sl18_emb + sl12_emb
    total_novedades = total - (total_descanso + total_entrenamiento + total_operaciones + total_emb)
    total_personal = total_descanso + total_entrenamiento + total_novedades + total_operaciones + total_emb

    return {
        "total": total_p,
        "resultados_cuadros": resultados_cuadros,
        "resultados_cuadros_slp": resultados_cuadros_slp,
        "total_descanso": fmt(total_descanso),
        "total_entrenamiento": fmt(total_entrenamiento),
        "total_novedades": fmt(total_novedades),
        "total_operaciones": fmt(total_operaciones),
        "total_emb": fmt(total_emb),
        "total_personal": fmt(total_personal)
    }

def filtrar_datos_por_clave(data, data_2, data_3, data_4_pto, data_5_pto):
    # Unidades principales y sus alias

    unidades = UNIDADES
    resultado_unidades = {}
    for nombre, aliases in unidades.items():
        resultado_unidades[nombre] = {
            "data": [row for row in data if row[9] in aliases],
            "data_2": [row for row in data_2 if row[0] in aliases and "BLICA" not in row[1]],
            "data_3": [row for row in data_3 if row[0] in aliases and "BLICA" not in row[1]]
        }

    # Filtro especial para BLICA
    resultado_unidades["BLICA"] = {
        "data": [row for row in data if row[10] == "BLICA"],
        "data_2": [row for row in data_2 if row[1] == "BLICA"],
        "data_3": [row for row in data_3 if row[1] == "BLICA"]
    }

    # Lista de departamentos
    departamentos = [
        "LA GUAJIRA", "MAGDALENA", "CESAR", "ATLÁNTICO", "BOLÍVAR", "SUCRE", "CÓRDOBA", "ANTIOQUIA",
        "SANTANDER", "NORTE DE SANTANDER", "CHOCÓ", "VALLE DEL CAUCA", "CAUCA", "NARIÑO", "PUTUMAYO",
        "CAQUETÁ", "AMAZONAS", "META", "GUAVIARE", "VAUPÉS", "GUAINÍA", "VICHADA", "ARAUCA", "CASANARE",
        "BOYACÁ", "CALDAS", "CUNDINAMARCA", "HUILA", "QUINDÍO", "RISARALDA", "TOLIMA"
    ]

    resultado_departamentos = {}
    for dpto in departamentos:
        resultado_departamentos[dpto] = {
            "data": [row for row in data if row[3] == dpto],
            "data_4": [row for row in data_4_pto if row[5] == dpto],
            "data_5": [row for row in data_5_pto if row[6] == dpto]
        }

    # Filtros con "in" (como BOGOTÁ, ARCHIPIÉLAGO)
    especiales = {
        "SAN_ANDRES_ISTAS": "ARCHIPIÉLAGO",
        "BOGOTA": "BOGOTÁ"
    }

    for clave, valor in especiales.items():
        resultado_departamentos[clave] = {
            "data": [row for row in data if valor in row[3]],
            "data_4": [row for row in data_4_pto if valor in row[5]],
            "data_5": [row for row in data_5_pto if valor in row[6]]
        }

    return resultado_unidades, resultado_departamentos

def llamado_unidades(dato):
    # Extracción de datos del diccionario
    unidades = UNIDADES
  
    fecha_insitop = dato.get("fecha_insitop", "")
    CODE = dato.get("CODE", "")
    dpto = dato.get("dpto", "").split(",")
    mpio = dato.get("mpio", "").split(",")
    divisiones_filtrados = dato.get("divisiones_filtrados", "").split(",")
    
    # Inicializar cláusulas WHERE
    where_dpto = ""
    where_mpio = ""
    where_unidad = ""

    # Procesar subregión si existe
    item = json.loads(dato.get('subregion', '{}'))

    if isinstance(item, dict) and "nombre" in item:
        valor_1 = validacion_cantidad_tupla([d["nombre"] for d in item["departamentos"]])
        valor_2 = validacion_cantidad_tupla([
            m for d in item["departamentos"] for m in d["municipios"]
        ])
        where_dpto = f"AND DEPARTAMENTO {valor_1}"
        where_mpio = f"AND MUNICIPIO {valor_2}"
    else:
        if dpto and len(dpto[0]) > 1:
            valor_1 = validacion_cantidad_tupla(dpto)
            where_dpto = f"AND DEPARTAMENTO {valor_1}"

        if mpio and len(mpio[0]) > 1:
            valor_2 = validacion_cantidad_tupla(mpio)
            where_mpio = f"AND MUNICIPIO {valor_2}"

    # Diccionario de unidades y filtrado

    unidad_filtrada = []
    if divisiones_filtrados and len(divisiones_filtrados[0]) > 1:
        for y in divisiones_filtrados:
            unidad_filtrada.extend(unidades.get(y, []))
        valor_3 = validacion_cantidad_tupla(unidad_filtrada)
        where_unidad = f"AND SIGLA_DIVISION {valor_3}"

    # Consultas SQL
    if CODE != "" and CODE != "---":
        query = f"""
            SELECT * FROM registro_insitop 
            WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
            AND CODE LIKE '%{CODE}%' 
            {where_dpto} {where_mpio} {where_unidad};
        """
    else:
        query = f"""
            SELECT * FROM registro_insitop 
            WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
            {where_dpto} {where_mpio} {where_unidad};
        """

    query_2 = f"""
        SELECT DISTINCT SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, COMPANIA, ARMA, SUBSTRING(PELOTON, 1, 1)
        FROM registro_insitop 
        WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
        {where_dpto} {where_mpio} {where_unidad};
    """

    query_3 = f"""
        SELECT DISTINCT SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, COMPANIA, ARMA, CODE, SUBSTRING(PELOTON, 1, 1)
        FROM registro_insitop 
        WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
        {where_dpto} {where_mpio} {where_unidad};
    """

    query_4_pto = f"""
        SELECT DISTINCT SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, COMPANIA, ARMA, DEPARTAMENTO, SUBSTRING(PELOTON, 1, 1)
        FROM registro_insitop 
        WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
        {where_dpto} {where_mpio} {where_unidad};
    """

    query_5_pto = f"""
        SELECT DISTINCT SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, COMPANIA, ARMA, CODE, DEPARTAMENTO, SUBSTRING(PELOTON, 1, 1)
        FROM registro_insitop 
        WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
        {where_dpto} {where_mpio} {where_unidad};
    """
    
    query_6_pto = f"""
        SELECT SIGLA_DIVISION, SIGLA_BRIGADA, SIGLA_UNIDD, COMPANIA, ARMA, CODE, DEPARTAMENTO, PELOTON, COMANDANTE , SUBSTRING(COMANDANTE, 1, 2) AS INICIALES_COMANDANTE, CELULAR_COMANDANTE
        FROM registro_insitop 
        WHERE FECHA_INSITOP LIKE '%{fecha_insitop}%' 
        {where_dpto} {where_mpio} {where_unidad};
    """

    # Ejecutar las consultas
    conn = connect()
    cursor = conn.cursor()

    try:
        cursor.execute(query)
        data = cursor.fetchall()

        cursor.execute(query_2)
        data_2 = cursor.fetchall()

        cursor.execute(query_3)
        data_3 = cursor.fetchall()

        cursor.execute(query_4_pto)
        data_4_pto = cursor.fetchall()

        cursor.execute(query_5_pto)
        data_5_pto = cursor.fetchall()

        cursor.execute(query_6_pto)
        data_6_pto = cursor.fetchall()

    finally:
        cursor.close()
        conn.close()
    return[data, data_2, data_3, data_4_pto, data_5_pto, fecha_insitop, data_6_pto]

        
def reporte_insitop(dato):
  
    data, data_2, data_3, data_4_pto, data_5_pto, fecha_insitop, data_6_pto = llamado_unidades(dato)
    unidades_filtradas, departamentos_filtrados = filtrar_datos_por_clave(data, data_2, data_3, data_4_pto, data_5_pto)
    datos_calculados = calcular_resumen_personal(data)

    pdf = PDF(orientation = 'P', unit = 'mm', format='letter')

    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION')

    caligrafia_ingreso( pdf, DIRECION)
    pdf.set_auto_page_break(True, 4)
    pdf.parametros( fondo_pagina = "NO" )
    pdf.add_page()

    #
    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))
    dirercion_archvios = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/"
    import qrcode

    fullname =  dato["fullname"]
    
    inf_qur = "JEFATURA DE ESTADO MAYOR DE OPERACIONES \n"+"DIRECIÓN DE OPERACIONES DEL EJÉRCITO \n" + "Fecha de Informacion \n" +fecha_insitop +"\n"+fullname+"\n"+str(fecha_insitop)
    img = qrcode.make(inf_qur)
    f = open(dirercion_archvios+"QR.png", "wb")
    img.save(f)
    f.close()
    
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/Diapositiva3.jpg",0,0,215.9,279.4)

    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/jemop.png",190,16,15,15)
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/EJC.png",10,255,15,20)

    pdf.image(dirercion_archvios+"QR.png",190,255,24,24)

    pdf.set_fill_color(255, 0, 0)
    #pdf.rounded_rect(5, 5, 206, 270, 1,'D', '1234')
        
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial Narrow', 'B', 11)

    pdf.text(190,7,str("JEMOP-DIROP"))
    pdf.set_text_color(40,40,40)
    pdf.set_font('Arial Narrow', 'B', 14)
 
    fecha_insitop_modificada = fecha(fecha_insitop)
    fecha_insitop_modificada = (str(fecha_insitop_modificada[0])+"-"+str(fecha_insitop_modificada[4])+"-"+str(fecha_insitop_modificada[2])).upper()

    pdf.text(185,57,str(fecha_insitop_modificada))

    pdf.set_text_color(56,87,35)
    pdf.set_font('Arial Black', '', 20)

    pdf.text(148,37,str("EFECTIVOS EN"))
    pdf.text(149,43,str("OPERACIONES"))
    pdf.set_draw_color(140,140,140)
    pdf.set_line_width(0.7)
    pdf.line(180, 45, 207, 45)
        
    pdf.set_text_color(140,140,140)
    pdf.set_font('Arial Narrow', 'B', 14)

    pdf.text(174,50,str("POR DIVISIONES"))

    

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 22)

    pdf.text(160,60,str(datos_calculados["total"]))
    pdf.set_font('BebasNeue', '', 20)
    pdf.set_text_color(56,87,35)
    pdf.text(160,66,str("TOTAL"))
    pdf.text(160,72,str("EN OPERACIONES"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 16)
    pdf.text(160,77,str( datos_calculados["resultados_cuadros"]))
    pdf.text(160,82,str(datos_calculados["resultados_cuadros_slp"]))


    #cantidad por divisiones 
    creaacon(pdf, unidades_filtradas["DIV01"]["data"], "DIV01", 88, 43, "div01")
    creaacon(pdf, unidades_filtradas["DIV02"]["data"], "DIV02", 105, 90, "div02")
    creaacon(pdf, unidades_filtradas["DIV03"]["data"], "DIV03", 45, 155, "div03")
    creaacon(pdf, unidades_filtradas["DIV04"]["data"], "DIV04", 115, 145, "div04")
    creaacon(pdf, unidades_filtradas["DIV05"]["data"], "DIV05", 70, 125, "div05")
    creaacon(pdf, unidades_filtradas["DIV06"]["data"], "DIV06", 120, 210, "div06")

    creaacon(pdf, unidades_filtradas["DIV07"]["data"], "DIV07", 55, 90, "div07")
    creaacon(pdf, unidades_filtradas["DIV08"]["data"], "DIV08", 155, 120, "div08")

    creaacon(pdf, unidades_filtradas["FUTOM"]["data"], "FUTOM", 100, 175, "futco")

    cuadros(pdf, unidades_filtradas["TREJC"]["data"], "TREJC", 10, 200, 200,200,180)
    cuadros(pdf, unidades_filtradas["DAVAA"]["data"], "DAVAA", 10, 213, 200,200,180)
    cuadros(pdf, unidades_filtradas["BLICA"]["data"], "BLICA", 10, 226, 200,200,180)
    cuadros(pdf, unidades_filtradas["CAAID"]["data"], "CAAID", 10, 239, 200,200,180)

    cuadros(pdf, unidades_filtradas["CAOCC"]["data"], "CAOCC", 62.5, 221, 212,215,211)
    cuadros(pdf, unidades_filtradas["CONAT"]["data"], "CONAT", 62.5, 234, 212,215,211)
    cuadros(pdf, unidades_filtradas["DIVFE"]["data"], "DIVFE", 62.5, 247, 212,215,211)
    
    pdf.set_font('BebasNeue', '', 16)
    pdf.set_text_color(56,87,35)
    pdf.text(180,220,str("CONVENCIONES"))

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 14)

    pdf.text(180,225,str("TOTAL PERSONAL"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 12)
    pdf.text(180,230,str("OFI - SUB"))
    pdf.text(180,235,str("SLP - SL18 - SL12"))


    #segundo mapa del estado CODE
    pdf.parametros( fondo_pagina = "NO" )
    pdf.add_page()

    #
    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))

    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/Diapositiva3.jpg",0,0,215.9,279.4)
    
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/jemop.png",190,16,15,15)
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/EJC.png",10,255,15,20)

    #pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",145,50,73,54.5)


    pdf.image(dirercion_archvios+"QR.png",190,255,24,24)

    pdf.set_fill_color(255, 0, 0)
    #pdf.rounded_rect(5, 5, 206, 270, 1,'D', '1234')
        
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial Narrow', 'B', 11)

    pdf.text(190,7,str("JEMOP-DIROP"))
    pdf.set_font('Arial Narrow', 'B', 14)
    pdf.set_text_color(40,40,40)
    pdf.text(185,57,str(fecha_insitop_modificada))

    pdf.set_text_color(56,87,35)
    pdf.set_font('Arial Black', '', 20)

    pdf.text(148,37,str("EFECTIVOS EN"))
    pdf.text(184,43,str("CODE"))
    
    pdf.set_draw_color(140,140,140)
    pdf.set_line_width(0.7)
    pdf.line(180, 45, 207, 45)
        
    pdf.set_text_color(140,140,140)
    pdf.set_font('Arial Narrow', 'B', 14)

    pdf.text(174,50,str("POR DIVISIONES"))


    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 22)

    pdf.text(135,55,"TOTAL. "+str( datos_calculados["total_personal"] ))
    pdf.set_font('BebasNeue', '', 20)
    pdf.set_text_color(56,87,35)
    pdf.text(135,64,str("CODE EJC"))
    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 16)
    pdf.text(135,69,"OPERACIONES: "+str( datos_calculados["total_operaciones"] ))
    pdf.text(135,74,"DESCANSO: "+str( datos_calculados["total_descanso"] ))
    pdf.text(135,79,"ENTRENAMIENTO: "+str( datos_calculados["total_entrenamiento"] ))
    pdf.text(135,84,"ESTADO MAYOR: "+str( datos_calculados["total_emb"] ))
    #pdf.text(135,89,"NOVEDADES: "+str( datos_calculados["total_novedades"]))

    calcular_CODE(pdf, unidades_filtradas["DIV01"]["data"], "DIV01", 88, 43, "div01")
    
    #cantidad por divisiones 
    calcular_CODE(pdf, unidades_filtradas["DIV02"]["data"], "DIV02", 105, 90, "div02")
    calcular_CODE(pdf, unidades_filtradas["DIV03"]["data"], "DIV03", 45, 155, "div03")
    calcular_CODE(pdf, unidades_filtradas["DIV04"]["data"], "DIV04", 115, 145, "div04")
    calcular_CODE(pdf, unidades_filtradas["DIV05"]["data"], "DIV05", 70, 125, "div05")
    calcular_CODE(pdf, unidades_filtradas["DIV06"]["data"], "DIV06", 120, 210, "div06")

    calcular_CODE(pdf, unidades_filtradas["DIV07"]["data"], "DIV07", 55, 90, "div07")
    calcular_CODE(pdf, unidades_filtradas["DIV08"]["data"], "DIV08", 155, 120, "div08")

    calcular_CODE(pdf, unidades_filtradas["FUTOM"]["data"], "FUTOM", 100, 175, "futco")

    cuadros_CODE(pdf, unidades_filtradas["TREJC"]["data"], "TREJC", 2, 197, 212,215,211)
    cuadros_CODE(pdf, unidades_filtradas["DAVAA"]["data"], "DAVAA", 2, 228, 200,200,180)

    cuadros_CODE(pdf, unidades_filtradas["BLICA"]["data"], "BLICA", 24, 197, 212,215,211)
    cuadros_CODE(pdf, unidades_filtradas["CAAID"]["data"], "CAAID", 24, 228, 200,200,180)

    cuadros_CODE(pdf, unidades_filtradas["CAOCC"]["data"], "CAOCC", 46, 197, 212,215,211)
    cuadros_CODE(pdf, unidades_filtradas["CONAT"]["data"], "CONAT", 46, 228, 200,200,180)

    cuadros_CODE(pdf, unidades_filtradas["DIVFE"]["data"], "DIVFE", 68, 228, 200,200,180)

    pdf.set_font('BebasNeue', '', 16)
    pdf.set_text_color(56,87,35)
    #pdf.text(180,220,str("CONVENCIONES"))

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 14)

    #pdf.text(180,225,str("TOTAL PERSONAL"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 12)
    #pdf.text(180,230,str("OPERACIONES"))
    #pdf.text(180,235,str("ENTRENAMIENTO"))
    #pdf.text(180,240,str("DESCANSO"))
    #pdf.text(180,245,str("ESTADO MAYOR"))
    #pdf.text(180,250,str("NOVEDADES"))


    #SEGUNDA DIVISION
    pdf.parametros( fondo_pagina = "NO" )
    pdf.add_page()


    #
    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))

    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/Diapositiva2.jpg",0,0,215.9,279.4)
    
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/jemop.png",190,16,15,15)
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/EJC.png",10,255,15,20)

    pdf.image(dirercion_archvios+"QR.png",190,255,24,24)

    pdf.set_fill_color(255, 0, 0)
    #pdf.rounded_rect(5, 5, 206, 270, 1,'D', '1234')
        
        
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial Narrow', 'B', 11)

    pdf.text(190,7,str("JEMOP-DIROP"))
    pdf.set_font('Arial Narrow', 'B', 14)
    pdf.set_text_color(40,40,40)
    pdf.text(185,57,str(fecha_insitop_modificada))

    pdf.set_text_color(56,87,35)
    pdf.set_font('Arial Black', '', 20)

    pdf.text(148,37,str("EFECTIVOS EN"))
    pdf.text(149,43,str("OPERACIONES"))
    
    pdf.set_draw_color(140,140,140)
    pdf.set_line_width(0.7)
    pdf.line(180, 45, 207, 45)
        
    pdf.set_text_color(140,140,140)
    pdf.set_font('Arial Narrow', 'B', 14)


    pdf.text(160,50,str("POR DEPARTAMENTOS"))

    DIVI(pdf, departamentos_filtrados["LA GUAJIRA"]["data"], 108, 25)
    DIVI(pdf, departamentos_filtrados["MAGDALENA"]["data"], 87, 35)

    DIVI(pdf, departamentos_filtrados["CESAR"]["data"], 101, 48)
    DIVI(pdf, departamentos_filtrados["ATLÁNTICO"]["data"], 73, 20)

    pdf.set_draw_color(93,203,177)
    pdf.set_line_width(0.5)
    pdf.line(77, 33, 82, 40)

    DIVI(pdf, departamentos_filtrados["BOLÍVAR"]["data"], 58, 25)
    pdf.set_draw_color(93,203,200)
    pdf.line(72, 38, 77, 48)
    pdf.set_draw_color(93,203,177)
    DIVI(pdf, departamentos_filtrados["SUCRE"]["data"], 50, 40)
    pdf.line(60, 53, 75, 57)

    DIVI(pdf, departamentos_filtrados["CÓRDOBA"]["data"], 63, 66)
    DIVI(pdf, departamentos_filtrados["ANTIOQUIA"]["data"], 61, 92)
    DIVI(pdf, departamentos_filtrados["SANTANDER"]["data"], 100, 92)
    DIVI(pdf, departamentos_filtrados["NORTE DE SANTANDER"]["data"], 110, 70)
    
    DIVI(pdf, departamentos_filtrados["CHOCÓ"]["data"], 15, 105)
    pdf.line(40, 117, 53, 117)

    DIVI(pdf, departamentos_filtrados["VALLE DEL CAUCA"]["data"], 15, 130)
    pdf.line(40, 142, 60, 142)
    
    DIVI(pdf, departamentos_filtrados["CAUCA"]["data"], 13, 147)
    pdf.line(38, 158, 45, 158)
    
    DIVI(pdf, departamentos_filtrados["NARIÑO"]["data"], 28, 170)

    DIVI(pdf, departamentos_filtrados["CAQUETÁ"]["data"], 85, 180)
    DIVI(pdf, departamentos_filtrados["AMAZONAS"]["data"], 120, 215)
    DIVI(pdf, departamentos_filtrados["META"]["data"], 100, 145)

    DIVI(pdf, departamentos_filtrados["GUAVIARE"]["data"], 112, 165)

    DIVI(pdf, departamentos_filtrados["VAUPÉS"]["data"], 140, 183)
    DIVI(pdf, departamentos_filtrados["GUAINÍA"]["data"], 168, 153)
    DIVI(pdf, departamentos_filtrados["VICHADA"]["data"], 168, 120)
    DIVI(pdf, departamentos_filtrados["ARAUCA"]["data"], 135, 94)
    DIVI(pdf, departamentos_filtrados["CASANARE"]["data"], 126, 110)

    DIVI(pdf, departamentos_filtrados["BOYACÁ"]["data"], 137, 80)
    DIVI(pdf, departamentos_filtrados["SAN_ANDRES_ISTAS"]["data"], 55, 3)
    pdf.line(62, 16, 50, 20)
    pdf.line(126, 98, 135, 88)
    pdf.line(70, 195, 82, 211)
    DIVI(pdf, departamentos_filtrados["PUTUMAYO"]["data"], 79, 211)
    
    cuadros(pdf, departamentos_filtrados["CUNDINAMARCA"]["data"], "CUNDINAMARCA", 162, 60, 217,217,217)
    cuadros(pdf, departamentos_filtrados["TOLIMA"]["data"], "TOLIMA", 162, 73, 218,112,16)
    cuadros(pdf, departamentos_filtrados["HUILA"]["data"], "HUILA", 162, 86, 227,29,70)
    
    cuadros(pdf, departamentos_filtrados["CALDAS"]["data"], "CALDAS", 10, 200, 200,157,207)
    cuadros(pdf, departamentos_filtrados["RISARALDA"]["data"], "RISARALDA", 10, 213, 4,113,181)
    cuadros(pdf, departamentos_filtrados["QUINDÍO"]["data"], "QUINDÍO", 10, 226, 218,237,156)
    cuadros(pdf, departamentos_filtrados["BOGOTA"]["data"], "BOGOTÁ, D.C.", 10, 239, 217,217,217)
    

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 22)

    pdf.text(70,235,str( datos_calculados["total"] ))
    pdf.set_font('BebasNeue', '', 20)
    pdf.set_text_color(56,87,35)
    pdf.text(70,241,str("TOTAL"))
    pdf.text(70,247,str("EN OPERACIONES"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 16)
    pdf.text(70,253,str( datos_calculados["resultados_cuadros"] ))
    pdf.text(70,259,str( datos_calculados["resultados_cuadros_slp"] ))
    
    
    pdf.set_font('BebasNeue', '', 16)
    pdf.set_text_color(56,87,35)
    pdf.text(180,220,str("CONVENCIONES"))

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 14)

    pdf.text(180,225,str("TOTAL PERSONAL"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 12)
    pdf.text(180,230,str("OFI - SUB"))
    pdf.text(180,235,str("SLP - SL18 - SL12"))


    #--------------------
    #--------------------
    
    #SEGUNDA DIVISION
    pdf.parametros( fondo_pagina = "NO" )
    pdf.add_page()


    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/Diapositiva3.jpg",0,0,215.9,279.4)
    
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/jemop.png",190,16,15,15)
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/EJC.png",10,255,15,20)



    pdf.image(dirercion_archvios+"QR.png",190,255,24,24)

    pdf.set_fill_color(255, 0, 0)
    #pdf.rounded_rect(5, 5, 206, 270, 1,'D', '1234')
        
        
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial Narrow', 'B', 11)

    pdf.text(190,7,str("JEMOP-DIROP"))
    pdf.set_font('Arial Narrow', 'B', 14)
    pdf.set_text_color(40,40,40)
    pdf.text(185,57,str(fecha_insitop_modificada))

    pdf.set_text_color(56,87,35)
    pdf.set_font('Arial Black', '', 20)

    pdf.text(145,37,str("PELOTONES EN"))
    pdf.text(184,43,str("CODE"))
    
       
    pdf.set_draw_color(140,140,140)
    pdf.set_line_width(0.7)
    pdf.line(170, 45, 207, 45)
        
    pdf.set_text_color(140,140,140)
    pdf.set_font('Arial Narrow', 'B', 14)

    pdf.text(174,50,str("POR DIVISIONES"))

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 22)

    total_pel  = pelotones_total(data_2, data_3)

    pdf.text(150,60,str(total_pel[0]))
    pdf.set_font('BebasNeue', '', 20)
    pdf.set_text_color(56,87,35)
    pdf.text(150,66,str("TOTAL"))
    pdf.text(150,72,str("PELOTONES"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 16)

    pdf.line(150, 73, 190, 73)

    pdf.text(150,79,"DISPOSITIVO")
    pdf.text(150,84,"OPERACIONES: "+str(total_pel[1]))
    pdf.text(150,89,"DESCANSO: "+str(total_pel[3]))
    pdf.text(150,94,"ENTRENAMIENTO: "+str(total_pel[2]))


    pelotones(pdf, unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"], "DIV01", 88, 43, "div01")
    pelotones(pdf, unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"], "DIV02", 105, 90, "div02")
    pelotones(pdf, unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"], "DIV03", 45, 155, "div03")
    pelotones(pdf, unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"], "DIV04", 115, 145, "div04")
    pelotones(pdf, unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"], "DIV05", 70, 125, "div05")
    pelotones(pdf, unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"], "DIV06", 120, 210, "div06")
    pelotones(pdf, unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"], "DIV07", 55, 90, "div07")
    pelotones(pdf, unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"], "DIV08", 155, 120, "div08")
    pelotones(pdf, unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"], "FUTOM", 100, 175, "futco")

    
    cuadros_peloton(pdf, unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"], "TREJC", 2, 197, 212,215,211)
    cuadros_peloton(pdf, unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"], "DAVAA", 2, 228, 200,200,180)

    cuadros_peloton(pdf, unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"], "BLICA", 24, 197, 212,215,211)
    cuadros_peloton(pdf, unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"], "CAAID", 24, 228, 200,200,180)

    cuadros_peloton(pdf, unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"], "CAOCC", 46, 197, 212,215,211)
    cuadros_peloton(pdf, unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"], "CONAT", 46, 228, 200,200,180)

    cuadros_peloton(pdf, unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"], "DIVFE", 68, 228, 200,200,180)


    #SEGUNDA DIVISION
    pdf.parametros( fondo_pagina = "NO" )
    pdf.add_page()


    #
    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))

    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/Diapositiva2.jpg",0,0,215.9,279.4)
    
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/jemop.png",190,16,15,15)
    pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/insitop/EJC.png",10,255,15,20)

    pdf.image(dirercion_archvios+"QR.png",190,255,24,24)

    pdf.set_fill_color(255, 0, 0)
    #pdf.rounded_rect(5, 5, 206, 270, 1,'D', '1234')
        
        
    pdf.set_text_color(255,255,255)
    pdf.set_font('Arial Narrow', 'B', 11)

    pdf.text(190,7,str("JEMOP-DIROP"))
    pdf.set_font('Arial Narrow', 'B', 14)
    pdf.set_text_color(40,40,40)
    pdf.text(185,57,str(fecha_insitop_modificada))

    pdf.set_text_color(56,87,35)
    pdf.set_font('Arial Black', '', 20)

    pdf.text(148,37,str("PELOTONES EN"))
    pdf.text(184,43,str("CODE"))
    
    pdf.set_draw_color(140,140,140)
    pdf.set_line_width(0.7)
    pdf.line(180, 45, 207, 45)
        
    pdf.set_text_color(140,140,140)
    pdf.set_font('Arial Narrow', 'B', 14)

    pdf.text(160,50,str("POR DEPARTAMENTOS"))

    DIVI_PEL(pdf, departamentos_filtrados["LA GUAJIRA"]["data_4"], departamentos_filtrados["LA GUAJIRA"]["data_5"], 108, 25)
    DIVI_PEL(pdf, departamentos_filtrados["MAGDALENA"]["data_4"], departamentos_filtrados["MAGDALENA"]["data_5"], 87, 35)

    DIVI_PEL(pdf, departamentos_filtrados["CESAR"]["data_4"], departamentos_filtrados["CESAR"]["data_5"], 101, 48)
    DIVI_PEL(pdf, departamentos_filtrados["ATLÁNTICO"]["data_4"], departamentos_filtrados["ATLÁNTICO"]["data_5"], 73, 20)

    pdf.set_draw_color(93,203,177)
    pdf.set_line_width(0.5)
    pdf.line(77, 33, 82, 40)

    DIVI_PEL(pdf, departamentos_filtrados["BOLÍVAR"]["data_4"], departamentos_filtrados["BOLÍVAR"]["data_5"], 58, 25)
    pdf.set_draw_color(93,203,200)
    pdf.line(72, 38, 77, 48)
    pdf.set_draw_color(93,203,177)
    DIVI_PEL(pdf, departamentos_filtrados["SUCRE"]["data_4"], departamentos_filtrados["SUCRE"]["data_5"], 50, 40)
    pdf.line(60, 53, 75, 57)

    DIVI_PEL(pdf, departamentos_filtrados["CÓRDOBA"]["data_4"], departamentos_filtrados["CÓRDOBA"]["data_5"], 63, 66)
    DIVI_PEL(pdf, departamentos_filtrados["ANTIOQUIA"]["data_4"], departamentos_filtrados["ANTIOQUIA"]["data_5"], 61, 92)
    DIVI_PEL(pdf, departamentos_filtrados["SANTANDER"]["data_4"], departamentos_filtrados["SANTANDER"]["data_5"], 100, 92)
    DIVI_PEL(pdf, departamentos_filtrados["NORTE DE SANTANDER"]["data_4"], departamentos_filtrados["NORTE DE SANTANDER"]["data_5"], 110, 70)
    
    DIVI_PEL(pdf, departamentos_filtrados["CHOCÓ"]["data_4"], departamentos_filtrados["CHOCÓ"]["data_5"], 15, 105)
    pdf.line(40, 117, 53, 117)

    DIVI_PEL(pdf, departamentos_filtrados["VALLE DEL CAUCA"]["data_4"], departamentos_filtrados["VALLE DEL CAUCA"]["data_5"], 15, 130)
    pdf.line(40, 142, 60, 142)
    
    DIVI_PEL(pdf, departamentos_filtrados["CAUCA"]["data_4"], departamentos_filtrados["CAUCA"]["data_5"], 13, 147)
    pdf.line(38, 158, 45, 158)
    
    DIVI_PEL(pdf, departamentos_filtrados["NARIÑO"]["data_4"], departamentos_filtrados["NARIÑO"]["data_5"], 28, 170)

    DIVI_PEL(pdf, departamentos_filtrados["CAQUETÁ"]["data_4"], departamentos_filtrados["CAQUETÁ"]["data_5"], 85, 180)
    DIVI_PEL(pdf, departamentos_filtrados["AMAZONAS"]["data_4"], departamentos_filtrados["AMAZONAS"]["data_5"], 120, 215)
    DIVI_PEL(pdf, departamentos_filtrados["META"]["data_4"], departamentos_filtrados["META"]["data_5"], 100, 145)

    DIVI_PEL(pdf, departamentos_filtrados["GUAVIARE"]["data_4"], departamentos_filtrados["GUAVIARE"]["data_5"], 112, 165)

    DIVI_PEL(pdf, departamentos_filtrados["VAUPÉS"]["data_4"], departamentos_filtrados["VAUPÉS"]["data_5"], 140, 183)
    DIVI_PEL(pdf, departamentos_filtrados["GUAINÍA"]["data_4"], departamentos_filtrados["GUAINÍA"]["data_5"], 168, 153)
    DIVI_PEL(pdf, departamentos_filtrados["VICHADA"]["data_4"], departamentos_filtrados["VICHADA"]["data_5"], 168, 120)
    DIVI_PEL(pdf, departamentos_filtrados["ARAUCA"]["data_4"], departamentos_filtrados["ARAUCA"]["data_5"], 135, 94)
    DIVI_PEL(pdf, departamentos_filtrados["CASANARE"]["data_4"], departamentos_filtrados["CASANARE"]["data_5"], 126, 110)

    DIVI_PEL(pdf, departamentos_filtrados["BOYACÁ"]["data_4"], departamentos_filtrados["BOYACÁ"]["data_5"], 137, 80)
    pdf.line(126, 98, 135, 88)
    
    cuadros_PEL(pdf, departamentos_filtrados["CUNDINAMARCA"]["data_4"], departamentos_filtrados["CUNDINAMARCA"]["data_5"], "CUNDINAMARCA", 162, 60, 217,217,217)
    cuadros_PEL(pdf, departamentos_filtrados["TOLIMA"]["data_4"], departamentos_filtrados["TOLIMA"]["data_5"], "TOLIMA", 162, 73, 218,112,16)
    cuadros_PEL(pdf, departamentos_filtrados["HUILA"]["data_4"], departamentos_filtrados["HUILA"]["data_5"], "HUILA", 162, 86, 227,29,70)
    
    cuadros_PEL(pdf, departamentos_filtrados["CALDAS"]["data_4"], departamentos_filtrados["CALDAS"]["data_5"], "CALDAS", 10, 200, 200,157,207)
    cuadros_PEL(pdf, departamentos_filtrados["RISARALDA"]["data_4"], departamentos_filtrados["RISARALDA"]["data_5"], "RISARALDA", 10, 213, 4,113,181)
    cuadros_PEL(pdf, departamentos_filtrados["QUINDÍO"]["data_4"], departamentos_filtrados["QUINDÍO"]["data_5"], "QUINDÍO", 10, 226, 218,237,156)
    cuadros_PEL(pdf, departamentos_filtrados["BOGOTA"]["data_4"], departamentos_filtrados["BOGOTA"]["data_5"], "BOGOTÁ, D.C.", 10, 239, 217,217,217)

    pdf.set_draw_color(93,203,177)
    pdf.set_line_width(0.5)
    pdf.line(62, 16, 50, 20)
    DIVI_PEL(pdf, departamentos_filtrados["SAN_ANDRES_ISTAS"]["data_4"], departamentos_filtrados["SAN_ANDRES_ISTAS"]["data_5"], 55, 3)

    pdf.line(70, 195, 82, 211)
    DIVI_PEL(pdf, departamentos_filtrados["PUTUMAYO"]["data_4"], departamentos_filtrados["PUTUMAYO"]["data_5"], 79, 211)

    
    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 22)

    total_pel  = pelotones_total(data_2, data_3)

    pdf.text(70,235,str(total_pel[0]))
    pdf.set_font('BebasNeue', '', 20)
    pdf.set_text_color(56,87,35)
    pdf.text(70,241,str("TOTAL"))
    pdf.text(70,247,str("PELOTONES"))
    PELOTONES = "OPE. "+str(total_pel[1])
    PELOTONES_2 = "ENT. "+str(total_pel[2]) + " - DES. "+str(total_pel[3])
    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 16)
    pdf.text(70,253,str(PELOTONES))
    pdf.text(70,259,str(PELOTONES_2))
    
    pdf.set_font('BebasNeue', '', 16)
    pdf.set_text_color(56,87,35)
    pdf.text(175,220,str("CONVENCIONES"))

    pdf.set_text_color(125,0,0)
    pdf.set_font('BebasNeue', '', 14)

    pdf.text(175,225,str("TOTAL PELOTONES"))

    pdf.set_text_color(50,50,50)
    pdf.set_font('BebasNeue', '', 12)
    pdf.text(175,230,str("OPERACIONES"))
    pdf.text(175,235,str("ENTRENAMIENTO - DESCANSO "))




    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION_3')
    titulo = "Efectivos en Operaciones {}".format(fecha_insitop)

    direcion =  LINK+"Efectivos en Operaciones {}.pdf".format(fecha_insitop)
    pdf.output(direcion, 'F')


    direcion= DIRECION+str("Efectivos en Operaciones {}.pdf").format(fecha_insitop)


    # print(direcion)
    return [direcion, titulo]

#--------------------------------------
#--------pelotones por unidades -------
#--------------------------------------
def reporte_insitop_pelotones(dato):

    data, data_2, data_3, data_4_pto, data_5_pto, fecha_insitop, data_6_pto = llamado_unidades(dato)
    unidades_filtradas, departamentos_filtrados = filtrar_datos_por_clave(data, data_2, data_3, data_4_pto, data_5_pto)


    pdf = PDF(orientation = 'L', unit = 'mm', format='letter')

    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION')

    caligrafia_ingreso( pdf, DIRECION)
    pdf.set_auto_page_break(True, 4)
    

    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))
    dirercion_archvios = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/"
    import qrcode

    #------- CANTIDAD DE PELOTONES POR DEPARTAMENTOS 

    fullname =  dato["fullname"]
    
    inf_qur = "JEFATURA DE ESTADO MAYOR DE OPERACIONES \n"+"DIRECIÓN DE OPERACIONES DEL EJÉRCITO \n" + "Fecha de Informacion \n" +fecha_insitop +"\n"+fullname+"\n"+str(fecha_insitop)
    img = qrcode.make(inf_qur)
    f = open(dirercion_archvios+"QR.png", "wb")
    img.save(f)
    f.close()

    total_pel  = pelotones_TOTAL_2_gaula(data_2, data_3)

    if total_pel[4] > 0:
        titulo = "CANTIDAD DE PELOTONES - {}".format(fecha_insitop)
        pdf.parametros( fondo_pagina = "SI" , titulo = titulo, imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
    
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)

                
        try:

            # declaring data 
            data = [total_pel[5], total_pel[6], total_pel[7]] 
            keys = ['OPE.', 'ENT.', 'DES.'] 
            colors = ["#c80000", "#1094df", "#73e72c"]

            import matplotlib as mpl
            mpl.rcParams['font.size'] = 9.0
        
            # declaring exploding pie 
            explode = [0.05, 0.05, 0.05]
            # define Seaborn color palette to use 
            palette_color = sns.color_palette('dark') 
            
            # plotting data on chart 
            font2 = {'family':'serif','color':'darkred','size':15}

            #sns.set(font_scale = 1.2)
            #plt.figure(figsize=(8,8))
            ax = plt.pie(data, labels=keys, colors=colors, 
                explode=explode, autopct='%.0f%%', pctdistance=0.8, labeldistance=1.05, textprops={'fontsize':20, 'weight':'bold'}, shadow=True ) 
            


            sns.despine()
            sns.color_palette("rocket")
            locs, labels = plt.xticks()
            plt.rcParams["figure.figsize"] = (250, 150)
            font1 = {'family':'serif','color':'darkred','size':20}
            

            #plt.title("CODE", fontdict = font1)


            plt.setp(labels)
            plt.tight_layout()
            
            # plotting data on chart 
            #plt.pie(data, labels=keys, colors=palette_color, autopct='%.0f%%')
            plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",  transparent=True )
            plt.clf()
        except:
            plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",  transparent=True )
            plt.clf()
    


        CANTIDAD_2=[]
        ANIO=[]
        columns=[]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV01" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
    
        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV02" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

                
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV03" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

                
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV04" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

                
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV05" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

                        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV06" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        
                        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV07" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
                        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV08" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "FUTOM" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
   

        Palette = ["#7F7F7F", "#c80000", "#1094df", "#73e72c"] #define your preference
        sns.set()
        
        #f, axs = plt.subplots(1, 2, figsize=(8, 4), gridspec_kw=dict(width_ratios=[4, 3]))
        ax = sns.barplot(x = columns, y=CANTIDAD_2 , hue = ANIO,  palette=Palette)
        
        ax.bar_label(ax.containers[3], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[2], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[1], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[0], fontsize=11, color="black",rotation=90,padding=3) 
        sns.despine()
        sns.color_palette("rocket")
        locs, labels = plt.xticks()
        plt.rcParams["figure.figsize"] = (250, 100)
        font1 = {'family':'serif','color':'black','size':20}
        font2 = {'family':'serif','color':'darkred','size':15}


        #plt.xlabel("Afectaciones a la Amenaza", fontdict = font2)
        plt.rc('xtick', labelsize=30) 
        plt.rc('ytick', labelsize=30) 

        plt.setp(labels, rotation=45)
       
        plt.tight_layout()

        sns.despine(top=True, right=True, left=True, bottom=True)
        
        plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo1.png", transparent=True)
        ax.get_figure().clf()

        
        CANTIDAD_2=[]
        ANIO=[]
        columns=[]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DAVAA" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CONAT" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIVFE" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "BLICA" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "TREJC" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CAAID" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CAOCC" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

          
        Palette = ["#7F7F7F", "#c80000", "#1094df", "#73e72c"] #define your preference
        sns.set()
        
        #f, axs = plt.subplots(1, 2, figsize=(8, 4), gridspec_kw=dict(width_ratios=[4, 3]))
        ax = sns.barplot(x = columns, y=CANTIDAD_2 , hue = ANIO, palette=Palette)
        
        ax.bar_label(ax.containers[3], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[2], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[1], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[0], fontsize=11, color="black",rotation=90,padding=3)
        sns.despine()
        sns.color_palette("rocket")
        locs, labels = plt.xticks()
        plt.rcParams["figure.figsize"] = (250, 100)
        font1 = {'family':'serif','color':'black','size':20}
        font2 = {'family':'serif','color':'darkred','size':15}


        #plt.xlabel("Afectaciones a la Amenaza", fontdict = font2)
        plt.rc('xtick', labelsize=30) 
        plt.rc('ytick', labelsize=30) 

        plt.setp(labels, rotation=45)
       
        plt.tight_layout()

        sns.despine(top=True, right=True, left=True, bottom=True)
        
        plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo2.png", transparent=True)
        ax.get_figure().clf()


        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo1.png",60,35,200,80)
        
        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo2.png",65,120,200,80)

            
        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",1,110,73,54.5)

    
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV01", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
    
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV02", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV03", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV04", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV05", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV06", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV07", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV08", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA FUTOM", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DAVAA", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CONAT", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIVFE", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA BLICA", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"], 30, 44, 5)

        
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA TREJC", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
        
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"], 30, 44, 5)
        
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CAAID", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"], 30, 44, 5)
    
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CAOCC", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        cuadros_divisiones_pelotones_pelotones(pdf, unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"], 30, 44, 5)

    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION_3')
    titulo = "Efectivos en Operaciones {}".format(fecha_insitop)

    direcion =  LINK+"Efectivos en Operaciones {}.pdf".format(fecha_insitop)
    pdf.output(direcion, 'F')

    direcion= DIRECION+str("Efectivos en Operaciones {}.pdf").format(fecha_insitop)


    # print(direcion)
    return [direcion, titulo]



    #cuadros_peloton(pdf, CAOCC_2, CAOCC_3, "CAOCC", 46, 197, 212,215,211)



#--------------------------------------
#--------pelotones por unidades -------
#--------------------------------------
def reporte_insitop_pelotones_divisiones(dato):
    
    data, data_2, data_3, data_4_pto, data_5_pto, fecha_insitop, data_6_pto = llamado_unidades(dato)
    unidades_filtradas, departamentos_filtrados = filtrar_datos_por_clave(data, data_2, data_3, data_4_pto, data_5_pto)


    pdf = PDF(orientation = 'L', unit = 'mm', format='letter')

    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION')

    caligrafia_ingreso( pdf, DIRECION)
    pdf.set_auto_page_break(True, 4)
    

    #pdf.text(35,20,str("EJERCITO NACIONAL DE COLOMBIA"))
    #pdf.text(35,25,str("DIRECCIÓN DE OPERACIONES"))
    dirercion_archvios = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/"
    import qrcode

    #------- CANTIDAD DE PELOTONES POR DEPARTAMENTOS 


    fullname =  dato["fullname"]
    
    inf_qur = "JEFATURA DE ESTADO MAYOR DE OPERACIONES \n"+"DIRECIÓN DE OPERACIONES DEL EJÉRCITO \n" + "Fecha de Informacion \n" +fecha_insitop +"\n"+fullname+"\n"+str(fecha_insitop)
    img = qrcode.make(inf_qur)
    f = open(dirercion_archvios+"QR.png", "wb")
    img.save(f)
    f.close()

    total_pel  = pelotones_TOTAL_2_gaula(data_2, data_3)

    if total_pel[4] > 0:
        titulo = "CANTIDAD DE PELOTONES - {}".format(fecha_insitop)
        pdf.parametros( fondo_pagina = "SI" , titulo = titulo, imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
    
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)

                
        try:

            # declaring data 
            data = [total_pel[5], total_pel[6], total_pel[7]] 
            keys = ['OPE.', 'ENT.', 'DES.'] 
            colors = ["#c80000", "#1094df", "#73e72c"]

            import matplotlib as mpl
            mpl.rcParams['font.size'] = 9.0
        
            # declaring exploding pie 
            explode = [0.05, 0.05, 0.05]
            # define Seaborn color palette to use 
            palette_color = sns.color_palette('dark') 
            
            # plotting data on chart 
            font2 = {'family':'serif','color':'darkred','size':15}

            #sns.set(font_scale = 1.2)
            #plt.figure(figsize=(8,8))
            ax = plt.pie(data, labels=keys, colors=colors, 
                explode=explode, autopct='%.0f%%', pctdistance=0.8, labeldistance=1.05, textprops={'fontsize':20, 'weight':'bold'}, shadow=True ) 
            


            sns.despine()
            sns.color_palette("rocket")
            locs, labels = plt.xticks()
            plt.rcParams["figure.figsize"] = (250, 150)
            font1 = {'family':'serif','color':'darkred','size':20}
            

            #plt.title("CODE", fontdict = font1)


            plt.setp(labels)
            plt.tight_layout()
            
            # plotting data on chart 
            #plt.pie(data, labels=keys, colors=palette_color, autopct='%.0f%%')
            plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",  transparent=True )
            plt.clf()
        except:
            plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",  transparent=True )
            plt.clf()
    


        CANTIDAD_2=[]
        ANIO=[]
        columns=[]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV01" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
    
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV02" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
  
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV03" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
     
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV04" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
 
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV05" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
         
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV06" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
              
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV07" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
                        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIV08" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "FUTOM" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
   

        Palette = ["#7F7F7F", "#c80000", "#1094df", "#73e72c"] #define your preference
        sns.set()
        
        #f, axs = plt.subplots(1, 2, figsize=(8, 4), gridspec_kw=dict(width_ratios=[4, 3]))
        ax = sns.barplot(x = columns, y=CANTIDAD_2 , hue = ANIO,  palette=Palette)
        
        ax.bar_label(ax.containers[3], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[2], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[1], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[0], fontsize=11, color="black",rotation=90,padding=3) 
        sns.despine()
        sns.color_palette("rocket")
        locs, labels = plt.xticks()
        plt.rcParams["figure.figsize"] = (250, 100)
        font1 = {'family':'serif','color':'black','size':20}
        font2 = {'family':'serif','color':'darkred','size':15}


        #plt.xlabel("Afectaciones a la Amenaza", fontdict = font2)
        plt.rc('xtick', labelsize=30) 
        plt.rc('ytick', labelsize=30) 

        plt.setp(labels, rotation=45)
       
        plt.tight_layout()

        sns.despine(top=True, right=True, left=True, bottom=True)
        
        plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo1.png", transparent=True)
        ax.get_figure().clf()

        
        CANTIDAD_2=[]
        ANIO=[]
        columns=[]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DAVAA" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CONAT" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]


        total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "DIVFE" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

        total_pel  = pelotones_TOTAL_2(unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "BLICA" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "TREJC" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CAAID" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]
        total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"])
        datos_matriz = llenar_matriz(CANTIDAD_2, total_pel, columns, ANIO, "CAOCC" )
        CANTIDAD_2=datos_matriz[0]
        columns=datos_matriz[1]
        ANIO=datos_matriz[2]

          
        Palette = ["#7F7F7F", "#c80000", "#1094df", "#73e72c"] #define your preference
        sns.set()
        
        #f, axs = plt.subplots(1, 2, figsize=(8, 4), gridspec_kw=dict(width_ratios=[4, 3]))
        ax = sns.barplot(x = columns, y=CANTIDAD_2 , hue = ANIO, palette=Palette)
        
        ax.bar_label(ax.containers[3], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[2], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[1], fontsize=11, color="black",rotation=90,padding=3)
        ax.bar_label(ax.containers[0], fontsize=11, color="black",rotation=90,padding=3)
        sns.despine()
        sns.color_palette("rocket")
        locs, labels = plt.xticks()
        plt.rcParams["figure.figsize"] = (250, 100)
        font1 = {'family':'serif','color':'black','size':20}
        font2 = {'family':'serif','color':'darkred','size':15}


        #plt.xlabel("Afectaciones a la Amenaza", fontdict = font2)
        plt.rc('xtick', labelsize=30) 
        plt.rc('ytick', labelsize=30) 

        plt.setp(labels, rotation=45)
       
        plt.tight_layout()

        sns.despine(top=True, right=True, left=True, bottom=True)
        
        plt.savefig("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo2.png", transparent=True)
        ax.get_figure().clf()


        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo1.png",60,35,200,80)
        
        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/Ejemplo2.png",65,120,200,80)

            
        pdf.image("C:/Users/nicolas.valbuena/Documents/programacion 2023/server/procesamiento_insitop/QR/torta_code.png",1,110,73,54.5)

    
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV01", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
    
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV01"]["data_2"], unidades_filtradas["DIV01"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV02", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV02"]["data_2"], unidades_filtradas["DIV02"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV03", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV03"]["data_2"], unidades_filtradas["DIV03"]["data_3"], 30, 44, 5)



    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV04", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV04"]["data_2"], unidades_filtradas["DIV04"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV05", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV05"]["data_2"], unidades_filtradas["DIV05"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV06", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf,unidades_filtradas["DIV06"]["data_2"], unidades_filtradas["DIV06"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV07", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV07"]["data_2"], unidades_filtradas["DIV07"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIV08", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIV08"]["data_2"], unidades_filtradas["DIV08"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA FUTOM", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["FUTOM"]["data_2"], unidades_filtradas["FUTOM"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DAVAA", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DAVAA"]["data_2"], unidades_filtradas["DAVAA"]["data_3"], 30, 44, 5)

    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CONAT", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["CONAT"]["data_2"], unidades_filtradas["CONAT"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA DIVFE", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["DIVFE"]["data_2"], unidades_filtradas["DIVFE"]["data_3"], 30, 44, 5)



    total_pel  = pelotones_TOTAL_2(unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA BLICA", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["BLICA"]["data_2"], unidades_filtradas["BLICA"]["data_3"], 30, 44, 5)


    total_pel  = pelotones_TOTAL_2(unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA TREJC", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)
        
        pdf.add_page()
        pdf.set_font('Arial Narrow', 'B', 11)
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["TREJC"]["data_2"], unidades_filtradas["TREJC"]["data_3"], 30, 44, 5)
        
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CAAID", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["CAAID"]["data_2"], unidades_filtradas["CAAID"]["data_3"], 30, 44, 5)
    
    total_pel  = pelotones_TOTAL_2(unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"])
    if total_pel[4] > 0:
        pdf.parametros( fondo_pagina = "SI" , titulo = "CANTIDAD DE PELOTONES DE LA CAOCC", imagen = "C:/Users/nicolas.valbuena/Documents/programacion 2023/server/static/img/carta/Diapositiva18.jpg",  img_qr =dirercion_archvios, total_pel=total_pel)

        pdf.add_page()
        cuadros_divisiones_pelotones(pdf, unidades_filtradas["CAOCC"]["data_2"], unidades_filtradas["CAOCC"]["data_3"], 30, 44, 5)

    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION_3')
    titulo = "Efectivos en Operaciones {}".format(fecha_insitop)

    direcion =  LINK+"Efectivos en Operaciones {}.pdf".format(fecha_insitop)
    pdf.output(direcion, 'F')

    direcion= DIRECION+str("Efectivos en Operaciones {}.pdf").format(fecha_insitop)


    # print(direcion)
    return [direcion, titulo]



    #cuadros_peloton(pdf, CAOCC_2, CAOCC_3, "CAOCC", 46, 197, 212,215,211)


#--------------------------------------
#--------cantidad de cdte pelotones -------
#--------------------------------------
def reporte_cantidad_cdte_pelotones(dato):
    data, data_2, data_3, data_4_pto, data_5_pto, fecha_insitop, data_6_pto = llamado_unidades(dato)

    jerarquia = ["GR", "MGR", "BGR", "CR", "TC", "MY", "CT", "TE", "ST", "SMC", "SM", "SP", "SV", "SS", "CP", "CS", "C3"]
    jerarquia_valor = {grado: i for i, grado in enumerate(jerarquia)}

    # Filtrar datos excluidos
    filtrados = [fila for fila in data_6_pto if fila[3].strip() not in EXCLUIR_ESTADOS]

    # Conteo por grado
    from collections import defaultdict
    conteo_grado = defaultdict(int)
    for fila in filtrados:
        grado = fila[9].strip()
        conteo_grado[grado] += 1

    # Agrupar por pelotón
    pelotones_dict = defaultdict(list)
    for fila in filtrados:
        clave = tuple(fila[i].strip() for i in [0, 1, 2, 3])
        num = fila[7].strip()
        peloton = num[0] if num else ''
        pelotones_dict[clave + (peloton,)].append(fila)

    # Conteo jerárquico por pelotón
    conteo_peloton_por_grado = defaultdict(set)
    conteo_por_code_grado = defaultdict(lambda: defaultdict(int))
    conteo_por_division_grado = defaultdict(lambda: defaultdict(int))

    for key, filas in pelotones_dict.items():
        division = key[0]
        code = filas[0][5].strip()
        grados = [f[9].strip() for f in filas]
        if not grados:
            continue
        grado_mas_alto = sorted(grados, key=lambda g: jerarquia_valor.get(g, 999))[0]
        conteo_peloton_por_grado[grado_mas_alto].add(key)
        conteo_por_code_grado[code][grado_mas_alto] += 1
        conteo_por_division_grado[division][grado_mas_alto] += 1

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import os

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Consolidado"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def escribir_tabla(hoja, datos, titulo, col_inicio):
        grados_presentes = sorted({g for d in datos.values() for g in d.keys()}, key=lambda g: jerarquia_valor.get(g, 999))
        encabezado = [titulo] + grados_presentes + ["Total"]

        for i, valor in enumerate(encabezado):
            cell = hoja.cell(row=1, column=col_inicio + i, value=valor)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for fila_idx, (clave, subdict) in enumerate(datos.items(), start=2):
            hoja.cell(row=fila_idx, column=col_inicio, value=clave)
            hoja.cell(row=fila_idx, column=col_inicio).alignment = center_align
            hoja.cell(row=fila_idx, column=col_inicio).border = border

            for col_idx, grado in enumerate(grados_presentes, start=1):
                val = subdict.get(grado, 0)
                cell = hoja.cell(row=fila_idx, column=col_inicio + col_idx, value=val)
                cell.alignment = center_align
                cell.border = border

            total = sum(subdict.values())
            hoja.cell(row=fila_idx, column=col_inicio + len(grados_presentes) + 1, value=total).border = border

    def escribir_recuento_grado(hoja, datos, col_inicio):
        hoja.cell(row=1, column=col_inicio, value="Grado que reflejan mando en el INSITOP").font = header_font
        hoja.cell(row=1, column=col_inicio + 1, value="Cantidad").font = header_font
        hoja.cell(row=1, column=col_inicio).fill = header_fill
        hoja.cell(row=1, column=col_inicio + 1).fill = header_fill
        hoja.cell(row=1, column=col_inicio).alignment = center_align
        hoja.cell(row=1, column=col_inicio + 1).alignment = center_align

        for i, grado in enumerate(jerarquia, start=2):
            hoja.cell(row=i, column=col_inicio, value=grado).border = border
            hoja.cell(row=i, column=col_inicio + 1, value=datos.get(grado, 0)).border = border
            hoja.cell(row=i, column=col_inicio).alignment = center_align
            hoja.cell(row=i, column=col_inicio + 1).alignment = center_align

    # Escribir las tres tablas en la hoja
    escribir_recuento_grado(ws, conteo_grado, col_inicio=1)
    escribir_tabla(ws, conteo_por_division_grado, "cdte pelotón por División", col_inicio=5)
    escribir_tabla(ws, conteo_por_code_grado, "cdte pelotón por Code", col_inicio=25)

    # --------------------------------------
    # Crear hoja adicional de Comandantes (solo grado más alto por pelotón)
    # --------------------------------------
    hoja_cmd = wb.create_sheet(title="Comandantes")
    encabezados_cmd = ['DIVISIÓN', 'BRIGADA', 'UNIDAD', 'COMPAÑÍA', 'PELOTÓN', 'GRADO', 'COMANDANTE', 'CELULAR']

    for i, encabezado in enumerate(encabezados_cmd, start=1):
        cell = hoja_cmd.cell(row=1, column=i, value=encabezado)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    fila_excel = 2
    for key, filas in pelotones_dict.items():
        division, brigada, unidad, compania, peloton = key
        grados = [f[9].strip() for f in filas]
        if not grados:
            continue
        grado_mas_alto = sorted(grados, key=lambda g: jerarquia_valor.get(g, 999))[0]

        for fila in filas:
            if fila[9].strip() == grado_mas_alto:
                comandante = fila[8].strip()
                celular = fila[10].strip()
                valores = [division, brigada, unidad, compania, peloton, grado_mas_alto, comandante, celular]

                for col_idx, val in enumerate(valores, start=1):
                    cell = hoja_cmd.cell(row=fila_excel, column=col_idx, value=val)
                    cell.alignment = center_align
                    cell.border = border

                fila_excel += 1
                break  # solo el primero que cumpla con el grado más alto

    # --------------------------------------
    # Crear hoja adicional de Excluidos
    # --------------------------------------
    hoja_excluidos = wb.create_sheet(title="CDT UNIDADES Y PLANAS MAYORES")
    encabezados_excluidos = ['DIVISIÓN', 'BRIGADA', 'UNIDAD', 'COMPAÑÍA', 'PELOTÓN', 'GRADO', 'COMANDANTE', 'CELULAR', 'ESTADO']

    for i, encabezado in enumerate(encabezados_excluidos, start=1):
        cell = hoja_excluidos.cell(row=1, column=i, value=encabezado)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    fila_excluidos = 2
    for fila in data_6_pto:
        estado = fila[3].strip()
        if estado in EXCLUIR_ESTADOS:
            division = fila[0].strip()
            brigada = fila[1].strip()
            unidad = fila[2].strip()
            compania = fila[4].strip()
            peloton = fila[7].strip()[0] if fila[7].strip() else ""
            grado = fila[9].strip()
            comandante = fila[8].strip()
            celular = fila[10].strip()

            valores = [division, brigada, unidad, compania, peloton, grado, comandante, celular, estado]
            for col_idx, val in enumerate(valores, start=1):
                cell = hoja_excluidos.cell(row=fila_excluidos, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = border

            fila_excluidos += 1


    # Guardar archivo
    LINK = os.getenv('DIRECION_3_B')
    DIRECION = os.getenv('DIRECION_3')
    titulo = f"resumen_grado_desde_codigo {fecha_insitop}"
    excel_path = f"{LINK}resumen_grado_desde_codigo {fecha_insitop}.xlsx"
    direcion = f"{DIRECION}resumen_grado_desde_codigo {fecha_insitop}.xlsx"
    wb.save(excel_path)
    print(f"✅ Archivo guardado como: {excel_path}")
    return [direcion, titulo]
