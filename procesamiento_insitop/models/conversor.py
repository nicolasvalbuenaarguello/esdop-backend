import openpyxl
from models.base_datos_posgrest import *
#from funciones.configuracion import *
import sys
def transformar(dato):
    dato = str(dato)
    dato = dato.replace("datetime.time(",  " '")
    dato = dato.replace("?",  ' ')
    dato = dato.replace("#",  ' ')
    # dato = dato.replace("/",  ' ')
    dato = dato.replace("*",  ' ')
    dato = dato.replace(",",  '-')
    dato = dato.replace("datetime.datetime(", " '")
    # dato = dato.replace("(",  ' ')
    dato = dato.replace("),",  "',")
    # dato = dato.replace(")",  ' ')
    dato = dato.replace("$",  ' ')
    dato = dato.replace("%",  ' ')
    dato = dato.replace("{",  ' ')
    dato = dato.replace("}",  ' ')
    dato = dato.replace("[",  ' ')
    dato = dato.replace("]",  ' ')
    dato = dato.replace("<",  ' ')
    dato = dato.replace(">",  ' ')
    dato = dato.replace("¨",  ' ')
    dato = dato.replace("^",  ' ')
    dato = dato.replace("~",  ' ')
    dato = dato.replace("`",  ' ')
    dato = dato.replace("'",  '')
    dato = dato.replace('"',  ' ')
    
    
    return dato
    
def reemplazar(dato_1,dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32, dato_33, dato_34, dato_35, dato_36, dato_37, dato_38, dato_39, dato_40, dato_41, dato_42, dato_43, dato_44, dato_45, dato_46, dato_47, dato_48, dato_49, dato_50):
    
    dato_1=transformar(str( dato_1.value))
    dato_2=transformar(str( dato_2.value))
    dato_3=transformar(str( dato_3.value))
    dato_4=transformar(str( dato_4.value))
    dato_5=transformar(str( dato_5.value))
    dato_6=transformar(str( dato_6.value))
    dato_7=transformar(str( dato_7.value))
    dato_8=transformar(str( dato_8.value))
    dato_9=transformar(str( dato_9.value))
    dato_10=transformar(str(dato_10.value))
    dato_11=transformar(str(dato_11.value))
    dato_12=transformar(str(dato_12.value))
    dato_13=transformar(str(dato_13.value))
    dato_14=transformar(str(dato_14.value))
    dato_15=transformar(str(dato_15.value))
    dato_16=transformar(str(dato_16.value))
    dato_17=transformar(str(dato_17.value))
    dato_18=transformar(str(dato_18.value))
    dato_19=transformar(str(dato_19.value))
    dato_20=transformar(str(dato_20.value))
    dato_21=transformar(str(dato_21.value))
    dato_22=transformar(str(dato_22.value))
    dato_23=transformar(str(dato_23.value))
    dato_24=transformar(str(dato_24.value))
    dato_25=transformar(str(dato_25.value))
    dato_26=transformar(str(dato_26.value))
    dato_27=transformar(str(dato_27.value))
    dato_28=transformar(str(dato_28.value))
    dato_29=transformar(str(dato_29.value))
    dato_30=transformar(str(dato_30.value))
    dato_31=transformar(str(dato_31.value))
    dato_32=transformar(str(dato_32.value))
    dato_33=transformar(str(dato_33.value))
    dato_34=transformar(str(dato_34.value))
    dato_35=transformar(str(dato_35.value))
    dato_36=transformar(str(dato_36.value))
    dato_37=transformar(str(dato_37.value))
    dato_38=transformar(str(dato_38.value))
    dato_39=transformar(str(dato_39.value))
    dato_40=transformar(str(dato_40.value))
    dato_41=transformar(str(dato_41.value))
    dato_42=transformar(str(dato_42.value))
    dato_43=transformar(str(dato_43.value))
    dato_44=transformar(str(dato_44.value))
    dato_45=transformar(str(dato_45.value))
    dato_46=transformar(str(dato_46.value))
    dato_47=transformar(str(dato_47.value))
    dato_48=transformar(str(dato_48.value))
    dato_49=transformar(str(dato_49.value))
    dato_50=transformar(str(dato_50.value))
    
    dato = [dato_1, dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32, dato_33, dato_34, dato_35, dato_36, dato_37, dato_38, dato_39, dato_40, dato_41, dato_42, dato_43, dato_44, dato_45, dato_46, dato_47, dato_48, dato_49, dato_50]

    return dato
  
      
def reemplazar_hechos(dato_1,dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32):
    
    dato_1=transformar(str( dato_1.value))
    dato_2=transformar(str( dato_2.value))
    dato_3=transformar(str( dato_3.value))
    dato_4=transformar(str( dato_4.value))
    dato_5=transformar(str( dato_5.value))
    dato_6=transformar(str( dato_6.value))
    dato_7=transformar(str( dato_7.value))
    dato_8=transformar(str( dato_8.value))
    dato_9=transformar(str( dato_9.value))
    dato_10=transformar(str(dato_10.value))
    dato_11=transformar(str(dato_11.value))
    dato_12=transformar(str(dato_12.value))
    dato_13=transformar(str(dato_13.value))
    dato_14=transformar(str(dato_14.value))
    dato_15=transformar(str(dato_15.value))
    dato_16=transformar(str(dato_16.value))
    dato_17=transformar(str(dato_17.value))
    dato_18=transformar(str(dato_18.value))
    dato_19=transformar(str(dato_19.value))
    dato_20=transformar(str(dato_20.value))
    dato_21=transformar(str(dato_21.value))
    dato_22=transformar(str(dato_22.value))
    dato_23=transformar(str(dato_23.value))
    dato_24=transformar(str(dato_24.value))
    dato_25=transformar(str(dato_25.value))
    dato_26=transformar(str(dato_26.value))
    dato_27=transformar(str(dato_27.value))
    dato_28=transformar(str(dato_28.value))
    dato_29=transformar(str(dato_29.value))
    dato_30=transformar(str(dato_30.value))
    dato_31=transformar(str(dato_31.value))
    dato_32=transformar(str(dato_32.value))


    
    dato = [dato_1, dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12,dato_13,dato_14,dato_15,dato_16,dato_17,dato_18,dato_19,dato_20,dato_21,dato_22,dato_23,dato_24,dato_25,dato_26,dato_27,dato_28,dato_29,dato_30, dato_31, dato_32]

    return dato 
def reemplazar_erradicacion(dato_1,dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12):
    
    dato_1=transformar(str( dato_1.value))
    dato_2=transformar(str( dato_2.value))
    dato_3=transformar(str( dato_3.value))
    dato_4=transformar(str( dato_4.value))
    dato_5=transformar(str( dato_5.value))
    dato_6=transformar(str( dato_6.value))
    dato_7=transformar(str( dato_7.value))
    dato_8=transformar(str( dato_8.value))
    dato_9=transformar(str( dato_9.value))
    dato_10=transformar(str(dato_10.value))
    dato_11=transformar(str(dato_11.value))
    dato_12=transformar(str(dato_12.value))

    
    
    dato = [dato_1, dato_2,dato_3,dato_4 ,dato_5,dato_6,dato_7,dato_8,dato_9,dato_10,dato_11,dato_12]

    return dato 

#coding='utf8'
class conversion ():
    def __init__(self):
        pass
           
    def leer_excel(fichero):
        global hojas
        global wb
        wb = openpyxl.load_workbook(fichero)
        hojas = wb.get_sheet_names()
        
                
            
    def convertir(libro, hoja, resultado):
        global a_1
        global dato_1
        global dato_2
        global fila
        
        



        # wb = openpyxl.load_workbook(libro)
        sheet = wb[hoja] # Obtenga el libro de trabajo actualmente activo
        fila = sheet.max_row # Número máximo de líneas
        fila = fila +1
        columna = sheet.max_column # Número máximo de columnas

        
        
        
        a_1 = ""
        dato_1 = []
        dato_2 = ""
        
        

        if resultado == "ReporteDispositivos" and columna == 32:
           # ri =2

            a_1 =()
            for ri in range(2, fila ):
               
                a_1 = reemplazar_hechos(sheet.cell(row = ri, column = 1),sheet.cell(row = ri, column = 2),sheet.cell(row = ri, column = 3),sheet.cell(row = ri, column = 4),sheet.cell(row = ri, column = 5),sheet.cell(row = ri, column = 6),sheet.cell(row = ri, column = 7),sheet.cell(row = ri, column = 8),sheet.cell(row = ri, column = 9),sheet.cell(row = ri, column = 10),sheet.cell(row = ri, column = 11),sheet.cell(row = ri, column = 12),sheet.cell(row = ri, column = 13),sheet.cell(row = ri, column = 14),sheet.cell(row = ri, column = 15),sheet.cell(row = ri, column = 16),sheet.cell(row = ri, column = 17),sheet.cell(row = ri, column = 18),sheet.cell(row = ri, column = 19),sheet.cell(row = ri, column = 20), sheet.cell(row = ri, column = 21),sheet.cell(row = ri, column = 22),sheet.cell(row = ri, column = 23),sheet.cell(row = ri, column = 24),sheet.cell(row = ri, column = 25),sheet.cell(row = ri, column = 26),sheet.cell(row = ri, column = 27),sheet.cell(row = ri, column = 28),sheet.cell(row = ri, column = 29),sheet.cell(row = ri, column = 30), sheet.cell(row = ri, column = 31),sheet.cell(row = ri, column = 32))  

                dato_1.append(a_1)
 
            # print(a_1)         
                # print( str(ri) + " completado de " + str(fila) , end='\r')
                # sys.stdout.flush()
            FECHA = a_1[1]
            dato_1 = str(dato_1)
            dato_1 = dato_1.replace(']',  '),')
            dato_1 = dato_1.replace('[',  '(')
            dato_1 = dato_1.replace('("',  '(')
            dato_1 = dato_1.replace('")',  ')')
            dato_1 = dato_1.replace(',,',  ', ') 
            dato_1 = dato_1.replace('", "',  '), (') 

            dato_1 =dato_1[:-3]
            dato_1 =dato_1[1:]



        
        save_data(dato_1, resultado, FECHA)
            

            
            # print(dato_2)
 


 
 
           
    def convertir_unidad(libro, hoja, resultado):
        global a_1
        global dato_1
        global dato_2
        global fila
        
        



        # wb = openpyxl.load_workbook(libro)
        sheet = wb[hoja] # Obtenga el libro de trabajo actualmente activo
        fila = sheet.max_row # Número máximo de líneas
        fila = fila +1
        columna = sheet.max_column # Número máximo de columnas

        
        a_1 = ""
        dato_1 = []
        dato_2 = ""
        
    
        if resultado == "ReporteDispositivos" and columna == 32:
           # ri =2

            a_1 =()
            for ri in range(2, fila ):
               
                a_1 = reemplazar_hechos(sheet.cell(row = ri, column = 1),sheet.cell(row = ri, column = 2),sheet.cell(row = ri, column = 3),sheet.cell(row = ri, column = 4),sheet.cell(row = ri, column = 5),sheet.cell(row = ri, column = 6),sheet.cell(row = ri, column = 7),sheet.cell(row = ri, column = 8),sheet.cell(row = ri, column = 9),sheet.cell(row = ri, column = 10),sheet.cell(row = ri, column = 11),sheet.cell(row = ri, column = 12),sheet.cell(row = ri, column = 13),sheet.cell(row = ri, column = 14),sheet.cell(row = ri, column = 15),sheet.cell(row = ri, column = 16),sheet.cell(row = ri, column = 17),sheet.cell(row = ri, column = 18),sheet.cell(row = ri, column = 19),sheet.cell(row = ri, column = 20), sheet.cell(row = ri, column = 21),sheet.cell(row = ri, column = 22),sheet.cell(row = ri, column = 23),sheet.cell(row = ri, column = 24),sheet.cell(row = ri, column = 25),sheet.cell(row = ri, column = 26),sheet.cell(row = ri, column = 27),sheet.cell(row = ri, column = 28),sheet.cell(row = ri, column = 29),sheet.cell(row = ri, column = 30), sheet.cell(row = ri, column = 31),sheet.cell(row = ri, column = 32))  

                dato_1.append(a_1)
 
            # print(a_1)         
                # print( str(ri) + " completado de " + str(fila) , end='\r')
                # sys.stdout.flush()
            FECHA = a_1[1]
            dato_1 = str(dato_1)
            dato_1 = dato_1.replace(']',  '),')
            dato_1 = dato_1.replace('[',  '(')
            dato_1 = dato_1.replace('("',  '(')
            dato_1 = dato_1.replace('")',  ')')
            dato_1 = dato_1.replace(',,',  ', ') 
            dato_1 = dato_1.replace('", "',  '), (') 

            dato_1 =dato_1[:-3]
            dato_1 =dato_1[1:]



        
        save_data(dato_1, resultado, FECHA)
            

            
            # print(dato_2)
 


 
 
 