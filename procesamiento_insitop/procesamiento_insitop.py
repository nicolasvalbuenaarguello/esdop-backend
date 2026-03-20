from typing import Annotated, List
from __init__ import *
from fastapi import FastAPI, File, UploadFile, Form

router = APIRouter()
import os
import shutil
import pandas as pd
from fastapi import Request, Depends, HTTPException
from sqlalchemy import create_engine, text
from fastapi.security import OAuth2PasswordBearer
import traceback

from fastapi.responses import StreamingResponse
from io import BytesIO
from geopy.distance import geodesic
from sklearn.neighbors import BallTree
import numpy as np

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile

from sklearn.neighbors import BallTree
import numpy as np
import pandas as pd
router = APIRouter()
# --- Configuración de conexión a PostgreSQL ---
DATABASE_URL = "postgresql://postgres:NICval10**@localhost:5432/dirop"
engine = create_engine(DATABASE_URL)
#importaciones a modulos propios

app = FastAPI(title="DIROP",
              description="api para el manejo de los resultados",
              version='1.0.1')

origins = [
"*",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=IP_SEGURITY,
    allow_credentials=True,
    expose_headers=[],
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme_2 = OAuth2PasswordBearer("/")
db = Databa_bases_2.conexion_directa()


# --- Función para limpiar carpeta ---
def limpiar_carpeta(ruta):
    for nombre_archivo in os.listdir(ruta):
        ruta_archivo = os.path.join(ruta, nombre_archivo)
        try:
            if os.path.isfile(ruta_archivo) or os.path.islink(ruta_archivo):
                os.unlink(ruta_archivo)
            elif os.path.isdir(ruta_archivo):
                shutil.rmtree(ruta_archivo)
        except Exception as e:
            print(f"Error eliminando {ruta_archivo}. Detalle: {e}")

# --- Endpoint para cargar archivo Excel de unidades ---
@router.post("/cargar_inistop_url_unidades", tags=["Unidades"])
async def guarda_datos_unidades(
    input: Request,
    token: str = Depends(oauth2_scheme_2)
):
    try:
        # --- Leer archivo del formulario ---
        contents = await input.form()
        archi_insitop = contents["archivo"]

        # --- Validación de extensión ---
        if not archi_insitop.filename.endswith((".xls", ".xlsx")):
            raise HTTPException(status_code=400, detail="⚠️ El archivo debe tener extensión .xls o .xlsx")

        # --- Preparar rutas ---
        ruta_base = os.getenv('DIRECION', os.getcwd())
        ruta_destino = os.path.join(ruta_base, 'procesamiento_insitop', 'documentos')
        os.makedirs(ruta_destino, exist_ok=True)
        limpiar_carpeta(ruta_destino)

        # --- Guardar archivo en carpeta temporal ---
        ruta_guardar = os.path.join(ruta_destino, "archivo.xlsx")
        with open(ruta_guardar, "wb") as myfile:
            archivo_bytes = await archi_insitop.read()
            myfile.write(archivo_bytes)

        # --- Leer archivo Excel con pandas ---
        df = pd.read_excel(ruta_guardar, engine='openpyxl')
        df.columns = df.columns.str.strip().str.lower()

        # --- Validar columnas requeridas ---
        columnas_esperadas = ['sigla_div', 'sigla_br', 'ut', 'nombre', 'departamento', 'ciudad', 'disposicion']
        columnas_faltantes = [col for col in columnas_esperadas if col not in df.columns]
        if columnas_faltantes:
            raise HTTPException(
                status_code=422,
                detail={
                    "mensaje": "Columnas requeridas faltantes",
                    "esperadas": columnas_esperadas,
                    "faltantes": columnas_faltantes
                }
            )

        # --- Preprocesamiento del DataFrame ---
        df.dropna(subset=['sigla_div', 'ut', 'nombre'], inplace=True)
        df = df.fillna("")

        # --- Verificar conexión a base de datos ---
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))

        # --- Eliminar datos antiguos y guardar nuevos ---
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM unidades"))
            df.to_sql('unidades', con=conn, if_exists='append', index=False, method='multi')

        return {
            "mensaje": "✅ Información cargada correctamente en la base de datos",
            "registros_insertados": len(df),
            "unidades": df
        }

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail="❌ Error al procesar el archivo. Detalle: " + str(e)
        )

  # --- Endpoint para leer unidades desde la base de datos ---

@router.get("/cargar_inistop_leer_unidades")
async def leer_unidades(token: str = Depends(oauth2_scheme_2)):
    try:
        with engine.connect() as conn:
            resultado = conn.execute(text("SELECT * FROM unidades"))
            columnas = resultado.keys()
            datos = [dict(zip(columnas, fila)) for fila in resultado.fetchall()]

        return {
            "mensaje": "✅ Unidades obtenidas correctamente",
            "total_registros": len(datos),
            "unidades": datos
        }

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"❌ Error al leer las unidades. Detalle: {e}"
        )
  

@router.post("/cargar_inistop_url")
async def guarda_datos(
    input: Request,
    token: str = Depends(oauth2_scheme_2)
):
    try:
        # --- Procesar archivo Excel recibido ---
        contents = await input.form()
        archi_insitop = contents["archi_insitop"]

        # --- Definir rutas ---
        ruta_base = os.getenv('DIRECION', getcwd())
        ruta_destino = os.path.join(ruta_base, 'procesamiento_insitop', 'documentos')
        ruta_guardar = os.path.join(getcwd(), "documentos", "INSITOP.xlsx")

        # --- Limpiar archivos antiguos ---
        if os.path.exists(ruta_destino):
            for file in os.listdir(ruta_destino):
                path = os.path.join(ruta_destino, file)
                try:
                    if os.path.isdir(path):
                        shutil.rmtree(path)
                    else:
                        os.remove(path)
                except Exception as e:
                    print(f"Error eliminando {path}: {e}")

        # --- Guardar nuevo archivo ---
        with open(ruta_guardar, "wb") as myfile:
            archivo_bytes = await archi_insitop.read()
            myfile.write(archivo_bytes)

        print("Archivo guardado correctamente en:", ruta_guardar)

        # --- Procesar el archivo (defínelo tú) ---
        await cargar_resultados()

        return {
            "informacion_insitop": "Información cargada en la base de datos"
        }

    except Exception as e:
        print("Error:", e)
        raise HTTPException(
            status_code=500,
            detail="Error al procesar el archivo. Detalle: " + str(e)
        )


@router.post("/cargar_inistop_url_calculo")
async def cargar_inistop_url_calculo(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = obtener_operaciones(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        return {"informacion_insitop_calculo": contentem}


@router.post("/reporte_inistop_url")
async def guarda_datos(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = reporte_insitop(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "link":contentem[0],
                "nombre":contentem[1]
            }
        return dir


#--------------------------------------
#--------cantidad de pelotones -------
#--------------------------------------
@router.post("/cantidad_cdte_url")
async def cantidad_cdte_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = reporte_cantidad_cdte_pelotones(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "link":contentem[0],
                "nombre":contentem[1]
            }
        return dir


#--------------------------------------
#--------pelotones por unidades -------
#--------------------------------------
@router.post("/reporte_inistop_url_pelotones")
async def reporte_inistop_pelotones_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = reporte_insitop_pelotones(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "link":contentem[0],
                "nombre":contentem[1]
            }
        return dir


#--------------------------------------
#--------pelotones por divisiones -----
#--------------------------------------
@router.post("/reporte_inistop_pelotones_url")
async def reporte_inistop_url_pelotones(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = reporte_insitop_pelotones_divisiones(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "link":contentem[0],
                "nombre":contentem[1]
            }
        return dir


@router.post("/listado_inistop_url")
async def listado_inistop_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = listado_inistop(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']

        # print(hechos_file) 



        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "listado_insitop":contentem[0],

            }
        return dir


@router.post("/grafico_listado_inistop_url")
async def grafico_listado_inistop_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = graficas(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']


        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "labels_insitop":contentem[0],
                "data_insitop":contentem[1],
                "valor_div01":contentem[2],
                "labe_unidades":contentem[3],
                "tipo_filtro":contentem[4]
  
            }
        return dir

@router.post("/grafico_estado_mayor_inistop_url")
async def grafico_estado_mayor_inistop_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = graficas_estado_mayor(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']


        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "total_descanso_data":contentem[0],
                "total_entrenamiento_data":contentem[1],
                "total_operaciones_data":contentem[2],
                "total_emb_data":contentem[3],
                "total_novedades_data":contentem[4],
                "label":contentem[5],
                "tipo_filtro":"SIGLA_DIVISION"

            }
        return dir


@router.post("/grafico_estado_mayor_inistop_unidad_url")
async def grafico_estado_mayor_inistop_unidad_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = graficas_estado_mayor_unidades(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']


        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "total_descanso_data":contentem[0],
                "total_entrenamiento_data":contentem[1],
                "total_operaciones_data":contentem[2],
                "total_emb_data":contentem[3],
                "total_novedades_data":contentem[4],
                "label":contentem[5],
                "tipo_filtro":contentem[6],
                "unidad_f":contentem[7]

            }
        return dir

@router.post("/grafico_unidades_listado_inistop_unidad_url")
async def grafico_unidades_listado_inistop_unidad_url(input: Request, token: str =  Depends(oauth2_scheme_2)):

    # token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    # print(token_decode.get("exp"))

    # user_name =  token_decode.get("sub")
    # exp = token_decode.get("exp")
    # code =  token_decode.get("code")
    # tiempo = datetime.utcfromtimestamp(exp)
    # tiempo_2 = tiempo-timedelta(minutes=10)
    # now = datetime.utcnow()+timedelta()
    # # print(now)
    # if now > tiempo_2:
    #          raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # try:
    #     token_decode = jwt.decode(token, key=SECRETE_KEY, algorithms=[ALGORITHM])
    #     user_name =  token_decode.get("sub")
    #     code =  token_decode.get("code")
    #     exp = token_decode.get("exp")

    #     # tiempo = datetime.utcfromtimestamp(exp)

    #     if user_name == None:
    #          raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError:
    #     raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})
    # except JWTError.ExpiredSignatureError:
    #      raise HTTPException(status_code=401, detail="su seccion expiro",headers={"WWW-Authenticate":"Bearer"})
    
    # user = User(user_name = user_name )
    # logged_user = ModelUser.loguin(db, user)

    # if not logged_user:
    #      raise HTTPException(status_code=401, detail="no se puede verifcar credenciales",headers={"WWW-Authenticate":"Bearer"})

    # if user.disabled:
    #       raise HTTPException(status_code=400, detail="expiro credenciales",headers={"WWW-Authenticate":"Bearer"})
    # else:
         
    #con esta variable se cambia el link de los documentos en pdf, se debe cambiar de acuerdo al entrono que se encuentre
    # link = "C:/inetpub/wwwroot/pagina_de_dirop/dociments/boletin/"#produccion 
    
        # ruta = "C:/Users/nicolas.valbuena/Documents/baken_dirop/"#ruta_absulota

        contents = await input.form()
        contentem = graficas_listado_dia_unidades(contents)

        # hechos_file = contents['hechos_file']
        # leer_resultados = input_json['leer_resultados']


        # if leer_herradicacion.filename != "":
                
        #     with open(getcwd() + "/documentos/ERRADICACION APLICATIVO.xlsx", "wb")as myfile:
        #         resultados = await leer_herradicacion.read()
        #         myfile.write(resultados)
        #         myfile.close()



        

        # await actualizar_datos()
        dir = {
                "labels_insitop":contentem[0],
                "data_insitop":contentem[1],
                "valor_div01":contentem[2],
                "labe_unidades":contentem[3],
                "tipo_filtro":contentem[4],
                "unidad_f":contentem[5],

            }
        return dir


import re

def limpiar_texto(x):
    if pd.isna(x):
        return ""
    return re.sub(r"[^\w-]", "", str(x).strip().upper())

@router.post("/unidades_no_cargadas")
async def unidades_no_cargadas(input: Request, token: str = Depends(oauth2_scheme_2)):
    form_data = await input.form()
    fecha_str = form_data.get("fecha", None)

    # 🔹 Validación de fecha enviada
    fecha = None
    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Fecha inválida: {fecha_str}")

    with engine.connect() as conn:
        # 🔹 Cargar unidades oficiales
        unidades_df = pd.read_sql("SELECT sigla_div, sigla_br, ut, nombre FROM unidades", conn)

        # 🔹 Cargar registro_insitop
        if fecha:
            # Traer todos los registros y filtrar en pandas
            registro_df = pd.read_sql("SELECT DISTINCT sigla_unidd, fecha_insitop FROM registro_insitop", conn)
            registro_df = registro_df[registro_df["fecha_insitop"].notna()]

            # Convertir a fecha en pandas (ignora errores)
            registro_df["fecha_parsed"] = pd.to_datetime(registro_df["fecha_insitop"], errors='coerce').dt.date

            # Filtrar por la fecha deseada
            registro_df = registro_df[registro_df["fecha_parsed"] == fecha]
            registro_df = registro_df[["sigla_unidd"]]
        else:
            registro_df = pd.read_sql("SELECT DISTINCT sigla_unidd FROM registro_insitop", conn)


    if "sigla_unidd" not in registro_df.columns:
        raise HTTPException(status_code=500, detail="Columna 'sigla_unidd' no encontrada en registro_insitop")

    # 🔹 Limpiar y normalizar
    unidades_df["ut"] = unidades_df["ut"].apply(limpiar_texto)
    registro_df["sigla_unidd"] = registro_df["sigla_unidd"].apply(limpiar_texto)

    # 🔹 Detectar unidades no cargadas usando merge
    merged = unidades_df.merge(
        registro_df.rename(columns={"sigla_unidd": "ut"}),
        on="ut",
        how="left",
        indicator=True
    )
    no_cargadas_df = merged[merged["_merge"] == "left_only"]

    # 🔹 Generar resumen
    resumen = []
    divisiones = no_cargadas_df["sigla_div"].dropna().unique()

    for div in divisiones:
        df_div = no_cargadas_df[no_cargadas_df["sigla_div"] == div]
        brigadas = df_div["sigla_br"].dropna().unique()
        unidades = df_div["ut"].dropna().unique()

        detalle = []
        for br in brigadas:
            df_bri = df_div[df_div["sigla_br"] == br]
            batallones = df_bri[["ut", "nombre"]].to_dict(orient="records")

            detalle.append({
                "sigla_br": br,
                "batallones": batallones
            })

        resumen.append({
            "sigla_div": div,
            "total_brigadas": len(brigadas),
            "total_unidades": len(unidades),
            "brigadas": sorted(brigadas),
            "unidades": sorted(unidades),
            "detalle": detalle
        })

    return {
        "fecha": fecha_str,
        "total_divisiones": len(resumen),
        "resumen": resumen
    }




@router.post("/reporte_unidades_no_cargadas_excel")
async def unidades_no_cargadas_excel(input: Request, token: str = Depends(oauth2_scheme_2)):
    form_data = await input.form()
    fecha_str = form_data.get("fecha", None)
    fecha = None
    if fecha_str:
        try:
            fecha = pd.to_datetime(fecha_str).date()
        except Exception:
            raise HTTPException(status_code=400, detail=f"Fecha inválida: {fecha_str}")

    with engine.connect() as conn:
        unidades_df = pd.read_sql("SELECT sigla_div, sigla_br, ut, nombre FROM unidades", conn)
        registro_df = pd.read_sql("SELECT DISTINCT sigla_unidd, fecha_insitop FROM registro_insitop", conn)

    # Filtrar registros válidos y convertir a fecha
    registro_df = registro_df[registro_df["fecha_insitop"].notna()]
    registro_df["fecha_parsed"] = pd.to_datetime(registro_df["fecha_insitop"], errors='coerce').dt.date

    if fecha:
        registro_df = registro_df[registro_df["fecha_parsed"] == fecha]

    registro_df = registro_df[["sigla_unidd"]]

    # Normalización
    unidades_df["ut"] = unidades_df["ut"].astype(str).str.strip().str.upper().str.replace("\u200b", "")
    registro_df["sigla_unidd"] = registro_df["sigla_unidd"].astype(str).str.strip().str.upper().str.replace("\u200b", "")

    # Filtrar no cargadas
    no_cargadas_df = unidades_df[~unidades_df["ut"].isin(registro_df["sigla_unidd"])]

    if no_cargadas_df.empty:
        raise HTTPException(status_code=404, detail="No hay unidades no cargadas para la fecha especificada.")

    # Exportar a Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        no_cargadas_df.to_excel(writer, index=False, sheet_name="Unidades No Cargadas")
    output.seek(0)

    filename = f"unidades_no_cargadas_{fecha_str or 'todas'}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@router.post("/unidades_aisladas")
async def unidades_aisladas(input: Request, token: str = Depends(oauth2_scheme_2)):
    form_data = await input.form()

    distancia_minima = float(form_data.get("distancia", 3000))  # en metros
    fecha = form_data.get("fecha", None)
    departamento = form_data.get("departamento", None)
    municipio = form_data.get("municipio", None)

    query = """
        SELECT sigla_division, sigla_brigada, sigla_unidd, compania, peloton,
               latitud, longitud, departamento, municipio
        FROM registro_insitop
        WHERE latitud IS NOT NULL AND longitud IS NOT NULL
          AND latitud != '' AND longitud != ''
    """

    filtros = []
    if fecha:
        filtros.append(
            f"fecha_insitop IS NOT NULL "
            f"AND fecha_insitop != '' "
            f"AND fecha_insitop != 'None' "
            f"AND fecha_insitop::date = '{fecha}'"
        )
    if departamento:
        filtros.append(f"departamento ILIKE '{departamento}'")
    if municipio:
        filtros.append(f"municipio ILIKE '{municipio}'")

    if filtros:
        query += " AND " + " AND ".join(filtros)

    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)

    df["latitud"] = pd.to_numeric(df["latitud"], errors="coerce")
    df["longitud"] = pd.to_numeric(df["longitud"], errors="coerce")
    df = df.dropna(subset=["latitud", "longitud"])

    if df.empty:
        return {
            "fecha": fecha,
            "filtros": {"departamento": departamento, "municipio": municipio},
            "distancia_umbral": distancia_minima,
            "total_aisladas": 0,
            "unidades_aisladas": []
        }

    # Convertir coordenadas a radianes [lat, lon]
    coords_rad = np.radians(df[["latitud", "longitud"]].astype(float).values)

    # Crear BallTree con métrica haversine
    tree = BallTree(coords_rad, metric="haversine")

    # Buscar 5 vecinos para tener mejor referencia
    dist, ind = tree.query(coords_rad, k=5)

    # Distancia al vecino más cercano distinto de sí mismo (índice 1)
    dist_metros = dist[:, 1] * 6371000  # radio tierra en metros

    df["distancia_minima"] = dist_metros

    # 🔍 Log de verificación: muestra las primeras 10 unidades y sus distancias
    print(df[["sigla_unidd", "latitud", "longitud", "distancia_minima"]].head(10))

    # Filtrar las aisladas
    aisladas_df = df[df["distancia_minima"] > distancia_minima]
    aisladas = aisladas_df.to_dict(orient="records")

    return {
        "fecha": fecha,
        "filtros": {"departamento": departamento, "municipio": municipio},
        "distancia_umbral": distancia_minima,
        "total_aisladas": len(aisladas),
        "unidades_aisladas": aisladas
    }



@router.post("/unidades_cercanas")
async def unidades_aisladas(input: Request, token: str = Depends(oauth2_scheme_2)):
    form_data = await input.form()

    fecha = form_data.get("fecha", None)
    departamento = form_data.get("departamento", None)
    municipio = form_data.get("municipio", None)

    query = """
        SELECT sigla_division, sigla_brigada, sigla_unidd, compania, peloton,
               latitud, longitud, departamento, municipio
        FROM registro_insitop
        WHERE latitud IS NOT NULL AND longitud IS NOT NULL
          AND latitud != '' AND longitud != ''
    """

    filtros = []
    if fecha:
        filtros.append(
            f"fecha_insitop IS NOT NULL "
            f"AND fecha_insitop != '' "
            f"AND fecha_insitop != 'None' "
            f"AND fecha_insitop::date = '{fecha}'"
        )
    if departamento:
        filtros.append(f"departamento ILIKE '{departamento}'")
    if municipio:
        filtros.append(f"municipio ILIKE '{municipio}'")

    if filtros:
        query += " AND " + " AND ".join(filtros)

    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)

    df["latitud"] = pd.to_numeric(df["latitud"], errors="coerce")
    df["longitud"] = pd.to_numeric(df["longitud"], errors="coerce")
    df = df.dropna(subset=["latitud", "longitud"])

    if df.empty:
        return {
            "fecha": fecha,
            "filtros": {"departamento": departamento, "municipio": municipio},
            "total_aisladas": 0,
            "unidades_aisladas": []
        }

    # Coordenadas en radianes
    coords_rad = np.radians(df[["latitud", "longitud"]].astype(float).values)

    # Crear árbol
    tree = BallTree(coords_rad, metric="haversine")

    # Buscar 2 vecinos (el mismo y el más cercano real)
    dist, ind = tree.query(coords_rad, k=2)

    dist_metros = dist[:, 1] * 6371000  # en metros
    vecinos_idx = ind[:, 1]

    # Armar unidades con su más cercana
    unidades = []
    for i, row in df.iterrows():
        vecino = df.iloc[vecinos_idx[i]]
        #print(vecino)
        unidades.append({
            "sigla_division": row["sigla_division"],
            "sigla_brigada": row["sigla_brigada"],
            "sigla_unidd": row["sigla_unidd"],
            "compania": row["compania"],  # o row["compañia"] si tiene tilde
            "peloton": row["peloton"],    # o row["Pelotón"]
            "departamento": row["departamento"],
            "municipio": row["municipio"],
            "latitud": row["latitud"],
            "longitud": row["longitud"],
            "unidad_cercana": vecino["sigla_unidd"], 
            "compania_cercana": vecino.get("compania") or vecino.get("compañia"),
            "peloton_cercana": vecino.get("peloton") or vecino.get("Pelotón"),
            "distancia_minima": float(dist_metros[i])
        })


    return {
        "fecha": fecha,
        "filtros": {"departamento": departamento, "municipio": municipio},
        "total_aisladas": len(unidades),
        "unidades_aisladas": unidades
    }


@router.post("/unidades_resumen_completo")
async def unidades_resumen_completo(input: Request, token: str = Depends(oauth2_scheme_2)):
    form_data = await input.form()
    
    fecha_inicio = form_data.get("fecha_inicio")
    fecha_fin = form_data.get("fecha_fin")
    distancia_minima_aisladas = float(form_data.get("distancia_aisladas", 3000))
    distancia_minima_movimiento = float(form_data.get("distancia_movimiento", 999))
    
    companias_excluidas = {"NOV", "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM", "P. MAYOR", "STAFF"}
    
    # ---------- Traer datos ----------
    query = f"""
        SELECT *
        FROM registro_insitop
        WHERE fecha_insitop BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
          AND compania NOT IN ({','.join([f"'{c}'" for c in companias_excluidas])})
          AND latitud IS NOT NULL AND longitud IS NOT NULL
          AND latitud != '' AND longitud != ''
        ORDER BY sigla_division, sigla_brigada, sigla_unidd, compania, peloton, fecha_insitop
    """
    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)
    
    # ---------- Limpiar ----------
    df["latitud"] = pd.to_numeric(df["latitud"], errors="coerce")
    df["longitud"] = pd.to_numeric(df["longitud"], errors="coerce")
    df["total_soldados"] = pd.to_numeric(df["total_soldados"], errors="coerce")
    df = df.dropna(subset=["latitud", "longitud", "total_soldados"])

    if df.empty:
        return {"mensaje": "No hay datos para las fechas o filtros seleccionados"}
    
    # ---------- Calcular aisladas ----------
    coords_rad = np.radians(df[["latitud", "longitud"]].values)
    tree = BallTree(coords_rad, metric="haversine")
    dist, ind = tree.query(coords_rad, k=5)
    df["distancia_minima"] = dist[:,1] * 6371000
    df["aislada"] = df["distancia_minima"] > distancia_minima_aisladas
    df["dias_aisladas"] = df.groupby(["sigla_division","sigla_brigada","sigla_unidd","compania","peloton"])["aislada"].cumsum()
    
    # ---------- Detectar sin movimiento ----------
    df = df.sort_values(["sigla_division","sigla_brigada","sigla_unidd","compania","peloton","fecha_insitop"])
    df["distancia_prev"] = df.groupby(["sigla_division","sigla_brigada","sigla_unidd","compania","peloton"])["distancia_minima"].shift(1)
    df["sin_movimiento"] = (df["distancia_minima"] - df["distancia_prev"]).abs() < distancia_minima_movimiento
    df["dias_sin_movimiento"] = df.groupby(["sigla_division","sigla_brigada","sigla_unidd","compania","peloton"])["sin_movimiento"].cumsum()
    
    # ---------- Detectar efectivos disminuidos ----------
    def soldados_minimo(peloton, total):
        return total < 28 if len(str(peloton)) == 1 else total < 22
    df["efectivos_disminuidos"] = df.apply(lambda x: soldados_minimo(x["peloton"], x["total_soldados"]), axis=1)
    
    # ---------- Detectar movimiento 0 ----------
    df["movimiento_0"] = df["distancia_minima"] == 0
    
    # ---------- Construir JSON unificado con resumen completo ----------
    unidades = {}
    for _, row in df.iterrows():
        key = (row["sigla_division"], row["sigla_brigada"], row["sigla_unidd"], row["compania"], row["peloton"])
        if key not in unidades:
            unidades[key] = {
                "division": row["sigla_division"],
                "brigada": row["sigla_brigada"],
                "unidad": row["sigla_unidd"],
                "compania": row["compania"],
                "peloton": row["peloton"],
                "comandante": row["comandante"],
                "tarea": row["tareas_accion_decisiva_ttf"],
                "aisladas": 0,
                "sin_movimiento": 0,
                "efectivos_disminuidos": 0,
                "movimiento_0": 0,
                "resumen": ""
            }
        # Actualizar valores según el registro actual
        if row["aislada"]:
            unidades[key]["aisladas"] = max(unidades[key]["aisladas"], row["dias_aisladas"])
        if row["sin_movimiento"]:
            unidades[key]["sin_movimiento"] = max(unidades[key]["sin_movimiento"], row["dias_sin_movimiento"])
        if row["efectivos_disminuidos"]:
            unidades[key]["efectivos_disminuidos"] = row["total_soldados"]
        if row["movimiento_0"]:
            unidades[key]["movimiento_0"] = max(unidades[key]["movimiento_0"], row["dias_sin_movimiento"])
        
        # Actualizar resumen completo con comandante y tarea
        resumen = []
        if unidades[key]["aisladas"]:
            resumen.append(f"Aisladas: {unidades[key]['aisladas']} días")
        if unidades[key]["sin_movimiento"]:
            resumen.append(f"Sin movimiento: {unidades[key]['sin_movimiento']} días")
        if unidades[key]["efectivos_disminuidos"]:
            resumen.append(f"Efectivos disminuidos: {unidades[key]['efectivos_disminuidos']}")
        if unidades[key]["movimiento_0"]:
            resumen.append(f"Movimiento 0: {unidades[key]['movimiento_0']} días")
        
        resumen.append(f"Tarea: {unidades[key]['tarea']}")
        resumen.append(f"Comandante: {unidades[key]['comandante']}")
        unidades[key]["resumen"] = ", ".join(resumen)
    
    # Convertir a lista ordenada
    result = sorted(unidades.values(), key=lambda x: (x["division"], x["brigada"], x["unidad"], x["compania"], x["peloton"]))
    
    return {"unidades": result}



@router.post("/descargar_reporte_excel")
async def unidades_resumen_completo(input: Request, token: str = Depends(oauth2_scheme_2)):
    import re, os, tempfile
    from collections import defaultdict, Counter
    import numpy as np
    from sklearn.neighbors import BallTree
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    form_data = await input.form()
    fecha_inicio = form_data.get("fecha_inicio")
    fecha_fin = form_data.get("fecha_fin")
    distancia_minima_aisladas = float(form_data.get("distancia_aisladas", 5000))
    distancia_minima_movimiento = float(form_data.get("distancia_movimiento", 999))

    companias_excluidas = {"NOV", "CDO", "EM", "EMB", "EMD", "PLM", "CDOYPM", "P. MAYOR", "STAFF"}

    # ---------- Traer datos ----------
    query = f"""
        SELECT *
        FROM registro_insitop
        WHERE fecha_insitop BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
          AND compania NOT IN ({','.join([f"'{c}'" for c in companias_excluidas])})
          AND latitud IS NOT NULL AND longitud IS NOT NULL
          AND latitud != '' AND longitud != ''
        ORDER BY sigla_division, sigla_brigada, sigla_unidd, compania, peloton, fecha_insitop
    """
    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)

    # ---------- Validaciones / limpieza ----------
    if df.empty:
        return {"mensaje": "No hay datos para las fechas o filtros seleccionados"}

    df["latitud"] = pd.to_numeric(df.get("latitud"), errors="coerce")
    df["longitud"] = pd.to_numeric(df.get("longitud"), errors="coerce")
    df["total_soldados"] = pd.to_numeric(df.get("total_soldados"), errors="coerce")
    df["fecha_insitop"] = pd.to_datetime(df.get("fecha_insitop"), errors="coerce")
    df = df.dropna(subset=["latitud", "longitud", "total_soldados", "fecha_insitop"])

    # ---------- Calcular distancias (kms) ----------
    coords_rad = np.radians(df[["latitud", "longitud"]].values)
    tree = BallTree(coords_rad, metric="haversine")
    dist, ind = tree.query(coords_rad, k=2)
    df["distancia_minima"] = dist[:, 1] * 6371.0  # kms

    # ---------- Flags / reglas ----------
    df["aislada"] = df["distancia_minima"] > (distancia_minima_aisladas / 1000)
    df["mov_menor_1km"] = df["distancia_minima"] < 1
    df["sin_movimiento"] = df["distancia_minima"] == 0

    def soldados_minimo(peloton, total):
        # lógica que indicaste
        return total < 30 if len(str(peloton)) == 1 else total < 15
    df["efectivos_disminuidos"] = df.apply(lambda x: soldados_minimo(x["peloton"], x["total_soldados"]), axis=1)

    # formatos
    df["distancia_km_fmt"] = df["distancia_minima"].apply(lambda x: f"{x:.2f} km")
    df["total_soldados_fmt"] = df["total_soldados"].apply(lambda x: f"{int(x):,}".replace(",", "."))

    # utilidades de texto
    def extract_grade(texto):
        if not isinstance(texto, str) or not texto.strip():
            return ""
        first = texto.strip().split()[0]
        return re.sub(r'[^A-Z0-9]', '', first.upper())

    def first_non_empty(dct, keys):
        for k in keys:
            if k in dct and pd.notna(dct[k]) and str(dct[k]).strip() != "":
                return dct[k]
        return ""

    # ---------- Preparar Excel ----------
    wb = Workbook()
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center")

    def escribir_hoja(nombre, columnas, datos):
        ws = wb.create_sheet(title=nombre)

        # ---------- Encabezados ----------
        for i, col in enumerate(columnas, start=1):
            c = ws.cell(row=1, column=i, value=col)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center_align, border

        # ---------- Filas ----------
        for r_idx, fila in enumerate(datos, start=2):
            for c_idx, valor in enumerate(fila, start=1):
                if valor is None or (isinstance(valor, float) and np.isnan(valor)) or str(valor).strip().lower() == "none":
                    valor = ""
                cell = ws.cell(row=r_idx, column=c_idx, value=valor)
                cell.alignment, cell.border = center_align, border

        # ---------- Ajuste de anchos ----------
        for idx, col in enumerate(ws.columns, start=1):
            col_letter = col[0].column_letter
            col_name = columnas[idx - 1]  # nombre de la columna

            if col_name.strip().upper() == "OBSERVACIONES":
                # Fijar ancho para Observaciones sin expandirse
                ws.column_dimensions[col_letter].width = 30
            else:
                # Autoajuste normal
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

        # ---------- Autofiltro ----------
        ws.auto_filter.ref = ws.dimensions

    # ---------- 1) No movimiento (agrupado: días sin movimiento) ----------
    dias_sin_mov = (
        df[df["sin_movimiento"]]
        .groupby([
            "sigla_division","sigla_brigada","sigla_unidd","compania","peloton",
            "lugar","municipio","comandante","tareas_accion_decisiva_ttf","code", "observaciones"
        ])
        .agg(dias_sin_mov=("fecha_insitop", lambda x: x.dt.normalize().nunique()))
        .reset_index()
    )
    dias_sin_mov = dias_sin_mov[[
        "sigla_division","sigla_brigada","sigla_unidd","compania","peloton",
        "lugar","municipio","comandante","dias_sin_mov","tareas_accion_decisiva_ttf","code", "observaciones"
    ]]
    escribir_hoja(
        "No movimiento",
        ["DIV","BR","BAT","CP","PEL","LUGAR","MUNICIPIO","COMANDANTE","DÍAS SIN MOVIMIENTO","TAREAS_ACCION_DECISIVA_TTF", "CODE", "OBSERVACIONES"],
        dias_sin_mov.values
    )

    # Llaves base para agrupar unidades (sin tarea/lugar para consolidar)
    unidad_keys = ["sigla_division","sigla_brigada","sigla_unidd","compania","peloton"]

    # ---------- 2) Mov <1km (un registro por unidad, dias únicos + distancia MÁX y datos asociados) ----------
    df_mov = df[df["mov_menor_1km"]].copy()
    mov_rows = []
    if not df_mov.empty:
        for keys, g in df_mov.groupby(unidad_keys):
            dias_unicos = g["fecha_insitop"].dt.normalize().nunique()
            # elegir fila con distancia máxima
            idx = g["distancia_minima"].idxmax()
            row = g.loc[idx]
            dist_max_fmt = f"{row['distancia_minima']:.2f} km" if pd.notna(row["distancia_minima"]) else ""
            tarea = (g["tareas_accion_decisiva_ttf"].dropna().mode().iloc[0]) if not g["tareas_accion_decisiva_ttf"].dropna().empty else ""
            code = (g["code"].dropna().mode().iloc[0]) if not g["code"].dropna().empty else ""
            # 🔑 concatenar observaciones (todas las no vacías del grupo)
            obs_series = g.get("observaciones")
            if obs_series is not None:
                observaciones = "; ".join(obs_series.dropna().astype(str).unique())
            else:
                observaciones = ""
            mov_rows.append([
                row.get("sigla_division",""), row.get("sigla_brigada",""), row.get("sigla_unidd",""),
                row.get("compania",""), row.get("peloton",""),
                row.get("lugar",""), row.get("municipio",""), row.get("comandante",""),
                dist_max_fmt, int(dias_unicos), tarea, code, observaciones
            ])
    escribir_hoja(
        "Mov <1km",
        ["DIV","BR","BAT","CP","PEL","LUGAR","MUNICIPIO","COMANDANTE","DIST MAX (KMs)","DÍAS","TAREAS_ACCION_DECISIVA_TTF", "CODE", "OBSERVACIONES"],
        mov_rows
    )

    # ---------- 3) Efectivos disminuidos (un registro por unidad, dias únicos + total MAX) ----------
    df_eff = df[df["efectivos_disminuidos"]].copy()
    eff_rows = []
    if not df_eff.empty:
        for keys, g in df_eff.groupby(unidad_keys):
            dias_unicos = g["fecha_insitop"].dt.normalize().nunique()
            idx = g["total_soldados"].idxmax()
            row = g.loc[idx]
            total_fmt = f"{int(row['total_soldados']):,}".replace(",", ".") if pd.notna(row["total_soldados"]) else ""
            tarea = (g["tareas_accion_decisiva_ttf"].dropna().mode().iloc[0]) if not g["tareas_accion_decisiva_ttf"].dropna().empty else ""
            code = (g["code"].dropna().mode().iloc[0]) if not g["code"].dropna().empty else ""
            obs_series = g.get("observaciones")
            if obs_series is not None:
                observaciones = "; ".join(obs_series.dropna().astype(str).unique())
            else:
                observaciones = ""
            eff_rows.append([
                row.get("sigla_division",""), row.get("sigla_brigada",""), row.get("sigla_unidd",""),
                row.get("compania",""), row.get("peloton",""),
                row.get("lugar",""), row.get("municipio",""), row.get("comandante",""),
                total_fmt, int(dias_unicos), tarea, code, observaciones
            ])
    escribir_hoja(
        "Efectivos disminuidos",
        ["DIV","BRI","BAT","CP","PEL","LUGAR","MUNICIPIO","COMANDANTE","TOTAL MAX HOMBRES","DÍAS","TAREAS_ACCION_DECISIVA_TTF", "CODE", "OBSERVACIONES"],
        eff_rows
    )



    def calcular_aisladas(df, distancia_umbral=3000):
        # Asegurar coordenadas válidas
        df["latitud"] = pd.to_numeric(df["latitud"], errors="coerce")
        df["longitud"] = pd.to_numeric(df["longitud"], errors="coerce")
        df = df.dropna(subset=["latitud", "longitud"]).copy()

        if df.empty:
            return []

        # Calcular distancia mínima con BallTree (haversine)
        coords_rad = np.radians(df[["latitud", "longitud"]].astype(float).values)
        tree = BallTree(coords_rad, metric="haversine")

        # k=2: [0] es la propia unidad, [1] el vecino más cercano
        dist, ind = tree.query(coords_rad, k=2)
        dist_metros = dist[:, 1] * 6371000  # radio tierra en metros
        df["distancia_minima"] = dist_metros

        # Filtrar aisladas
        df_aisl = df[df["distancia_minima"] > distancia_umbral].copy()
        if df_aisl.empty:
            return []

        aisl_rows = []
        for _, row in df_aisl.iterrows():
            dist_min_fmt = f"{row['distancia_minima']:.2f} m"
            aisl_rows.append([
                row.get("sigla_division",""), row.get("sigla_brigada",""), row.get("sigla_unidd",""),
                row.get("compania",""), row.get("peloton",""),
                row.get("lugar",""), row.get("comandante",""),
                dist_min_fmt,
                1 if "fecha_insitop" in row and pd.notna(row["fecha_insitop"]) else 0,
                row.get("tareas_accion_decisiva_ttf","") or "",
                row.get("code","") or "",
                row.get("observaciones","") or ""
            ])

        return aisl_rows

        
    
    aisl_rows = calcular_aisladas(df, distancia_umbral=distancia_minima_aisladas)

    escribir_hoja(
        "Aisladas",
        ["DIV","BR","BAT","CP","PEL","LUGAR","COMANDANTE",
        "DIST MIN (m)","DÍAS","TAREAS_ACCION_DECISIVA_TTF","CODE","OBSERVACIONES"],
        aisl_rows
    )


    # ---------- 5) Comandantes Cabos (jerarquía + dias como cabo) ----------
    # agrupar por pelotón similar a tu lógica
    pelotones_dict = defaultdict(list)
    for _, row in df.iterrows():
        pel_label = str(row.get("peloton","")).strip()
        peloton_label = pel_label[0] if pel_label else ""
        key = (row.get("sigla_division",""), row.get("sigla_brigada",""), row.get("sigla_unidd",""), row.get("compania",""), peloton_label)
        pelotones_dict[key].append(row.to_dict())

    CABO_SET = {"CP","CS","C3"}
    jerarquia = ["GR","MGR","BGR","CR","TC","MY","CT","TE","ST","SMC","SM","SP","SV","SS","CP","CS","C3"]
    jerarquia_valor = {g:i for i,g in enumerate(jerarquia)}

    cabos_rows = []
    for key, filas in pelotones_dict.items():
        division, brigada, unidad, compania, peloton_label = key
        grados = [ extract_grade(f.get("comandante","")) for f in filas if f.get("comandante") ]
        grados = [g for g in grados if g]
        if not grados: 
            continue
        grado_mas_alto = sorted(grados, key=lambda g: jerarquia_valor.get(g, 999))[0]
        if grado_mas_alto not in CABO_SET:
            continue
        filas_grado = [f for f in filas if extract_grade(f.get("comandante","")) == grado_mas_alto]
        if not filas_grado: 
            continue
        nombres = [ (f.get("comandante") or "").strip() for f in filas_grado ]
        nombre_comandante = Counter(nombres).most_common(1)[0][0] if nombres else ""
        celulares = [ first_non_empty(f, ["celular_comandante","CELULAR_COMANDANTE","celular","CELULAR"]) for f in filas_grado ]
        celular = Counter([c for c in celulares if c]).most_common(1)[0][0] if any(celulares) else ""
        fechas = [ pd.to_datetime(f.get("fecha_insitop")) for f in filas_grado if pd.notna(f.get("fecha_insitop")) ]
        dias_unicos = len(pd.Series(fechas).dropna().dt.normalize().unique()) if fechas else 0
        lugar = Counter([ first_non_empty(f, ["lugar","LUGAR"]) for f in filas_grado if first_non_empty(f, ["lugar","LUGAR"]) ]).most_common(1)
        municipio = Counter([ first_non_empty(f, ["municipio","MUNICIPIO"]) for f in filas_grado if first_non_empty(f, ["municipio","MUNICIPIO"]) ]).most_common(1)
        tarea = Counter([ first_non_empty(f, ["tareas_accion_decisiva_ttf","TAREAS_ACCION_DECISIVA_TTF"]) for f in filas_grado if first_non_empty(f, ["tareas_accion_decisiva_ttf","TAREAS_ACCION_DECISIVA_TTF"]) ]).most_common(1)
        code = Counter([ first_non_empty(f, ["code","CODE"]) for f in filas_grado if first_non_empty(f, ["code","CODE"]) ]).most_common(1)
        observaciones = Counter([ first_non_empty(f, ["observaciones","OBSERVACIONES"]) for f in filas_grado if first_non_empty(f, ["observaciones","OBSERVACIONES"]) ]).most_common(1)
        cabos_rows.append([
            division, brigada, unidad, compania, peloton_label,
            lugar[0][0] if lugar else "", municipio[0][0] if municipio else "",
            grado_mas_alto, nombre_comandante, celular, int(dias_unicos), tarea[0][0] if tarea else "", code[0][0] if code else "", observaciones[0][0] if observaciones else ""
        ])
    escribir_hoja(
        "Comandantes Cabos",
        ["DIV","BR","BAT","CP","PEL","LUGAR","MUNICIPIO","GRADO","COMANDANTE","CELULAR","DÍAS_COMO_CABO","TAREAS_ACCION_DECISIVA_TTF","CODE","OBSERVACIONES"],
        cabos_rows
    )

    # ---------- 6) CONTROL: todos los comandantes, días y pelotón resumido ----------
    control_rows = []
    # agrupamos por comandante (y unidad)
    grupos = df.groupby(["sigla_division","sigla_brigada","sigla_unidd","compania","comandante"])
    for (division, brigada, unidad, compania, comandante), subdf in grupos:
        if not comandante or str(comandante).strip() == "":
            continue
        # peloton más frecuente y peloton resumido (primer dígito)
        pel_common = subdf["peloton"].dropna().astype(str).str.strip()
        pel_most = pel_common.mode().iloc[0] if not pel_common.empty else ""
        pel_resum = str(pel_most).strip()[0] if str(pel_most).strip() else ""
        # días únicos
        dias_unicos = subdf["fecha_insitop"].dt.normalize().nunique()
        lugar = Counter([str(x) for x in subdf["lugar"] if pd.notna(x) and str(x).strip()]).most_common(1)
        municipio = Counter([str(x) for x in subdf["municipio"] if pd.notna(x) and str(x).strip()]).most_common(1)
        tarea = Counter([str(x) for x in subdf["tareas_accion_decisiva_ttf"] if pd.notna(x) and str(x).strip()]).most_common(1)
        control_rows.append([
            division, brigada, unidad, compania, pel_most,
            lugar[0][0] if lugar else "", municipio[0][0] if municipio else "",
            comandante, int(dias_unicos), pel_resum, tarea[0][0] if tarea else ""
        ])
    escribir_hoja(
        "CONTROL",
        ["DIV","BR","BAT","CP","PEL","LUGAR","MUNICIPIO","COMANDANTE","DÍAS","PEL_RESUMIDO","TAREAS_ACCION_DECISIVA_TTF"],
        control_rows
    )

    # Quitar sheet por defecto si existe
    try:
        del wb["Sheet"]
    except KeyError:
        pass

    # Guardar en archivo temporal
    tmpdir = tempfile.gettempdir()
    filename = f"resumen_unidades_{fecha_inicio}_{fecha_fin}.xlsx"
    filepath = os.path.join(tmpdir, filename)
    wb.save(filepath)

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )



app = FastAPI(title="DIROP",
              description="api para el manejo de los resultados",
              version='1.0.1')

origins = [
"*",
]
app.include_router(router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


import uvicorn
IP = os.getenv('IP')

PUERTO = int(os.getenv('PUERTO'))

if __name__ == "__main__":
     config = uvicorn.Config("procesamiento_insitop:app", port=PUERTO+10, host = IP, log_level="info", reload=True)
     server = uvicorn.Server(config)
     server.run()





